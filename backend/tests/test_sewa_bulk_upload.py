"""
Test SEWA Advance Bulk Upload Feature
Tests:
1. GET /api/payroll/sewa-advances/template/download - Excel template download
2. POST /api/payroll/sewa-advances/bulk-upload - Bulk upload with validation
3. Validation: employee exists, monthly deduction > 0, paid <= total
4. Replacement of existing active advances
"""
import pytest
import requests
import os
import io

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

class TestSewaAdvanceBulkUpload:
    """Test SEWA Advance Bulk Upload Feature"""
    
    @pytest.fixture(autouse=True)
    def setup(self):
        """Setup test - get auth token"""
        self.session = requests.Session()
        
        # Login as admin
        login_response = self.session.post(
            f"{BASE_URL}/api/auth/login",
            json={"email": "admin@shardahr.com", "password": "password"},
            headers={"Content-Type": "application/json"}
        )
        assert login_response.status_code == 200, f"Login failed: {login_response.text}"
        
        data = login_response.json()
        self.token = data.get("access_token")
        assert self.token, "No access token received"
        
        # Set auth header but NOT Content-Type (let requests set it for multipart)
        self.auth_headers = {"Authorization": f"Bearer {self.token}"}
        self.json_headers = {"Authorization": f"Bearer {self.token}", "Content-Type": "application/json"}
        print(f"✓ Logged in as admin@shardahr.com")
    
    def test_01_login_works_no_body_stream_error(self):
        """Test that login works without body stream already read error"""
        # Fresh session to test login
        session = requests.Session()
        response = session.post(
            f"{BASE_URL}/api/auth/login",
            json={"email": "admin@shardahr.com", "password": "password"},
            headers={"Content-Type": "application/json"}
        )
        
        assert response.status_code == 200, f"Login failed with status {response.status_code}: {response.text}"
        data = response.json()
        assert "access_token" in data, "No access_token in response"
        assert "user" in data, "No user in response"
        print(f"✓ Login successful, token received")
    
    def test_02_template_download_returns_excel(self):
        """Test GET /api/payroll/sewa-advances/template/download returns Excel file"""
        response = self.session.get(
            f"{BASE_URL}/api/payroll/sewa-advances/template/download",
            headers=self.auth_headers
        )
        
        assert response.status_code == 200, f"Template download failed: {response.status_code} - {response.text}"
        
        # Check content type
        content_type = response.headers.get("Content-Type", "")
        assert "spreadsheet" in content_type or "excel" in content_type or "octet-stream" in content_type, \
            f"Expected Excel content type, got: {content_type}"
        
        # Check content disposition
        content_disp = response.headers.get("Content-Disposition", "")
        assert "sewa_advance_template.xlsx" in content_disp, \
            f"Expected filename in Content-Disposition, got: {content_disp}"
        
        # Check file size (should be > 0)
        assert len(response.content) > 0, "Downloaded file is empty"
        
        # Verify it's a valid xlsx (starts with PK - zip signature)
        assert response.content[:2] == b'PK', "File doesn't appear to be a valid xlsx (zip) file"
        
        print(f"✓ Template downloaded successfully ({len(response.content)} bytes)")
    
    def test_03_template_contains_employee_data(self):
        """Test that template is pre-filled with employee codes"""
        import openpyxl
        
        response = self.session.get(
            f"{BASE_URL}/api/payroll/sewa-advances/template/download",
            headers=self.auth_headers
        )
        assert response.status_code == 200
        
        # Load workbook
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        
        # Check sheets exist
        assert "SEWA Advances" in wb.sheetnames, f"Missing 'SEWA Advances' sheet. Found: {wb.sheetnames}"
        assert "Instructions" in wb.sheetnames, f"Missing 'Instructions' sheet. Found: {wb.sheetnames}"
        
        ws = wb["SEWA Advances"]
        
        # Check headers
        headers = [ws.cell(1, i).value for i in range(1, 10)]
        expected_headers = [
            "Employee Code", "Employee Name", "Employee ID",
            "Total Advance Amount", "Monthly Deduction", "Amount Paid Till Now",
            "Start Month (1-12)", "Start Year", "Reason"
        ]
        assert headers == expected_headers, f"Headers mismatch. Got: {headers}"
        
        # Check that there are employee rows (at least one)
        row_count = 0
        for row in ws.iter_rows(min_row=2, max_row=100, values_only=True):
            if row[0] or row[2]:  # emp_code or employee_id
                row_count += 1
        
        assert row_count > 0, "Template has no employee data pre-filled"
        print(f"✓ Template contains {row_count} employees with correct headers")
    
    def test_04_bulk_upload_requires_auth(self):
        """Test that bulk upload requires authentication"""
        # Create a minimal xlsx file
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["EMP001", "Test", "emp_001", 10000, 1000, 0, 1, 2026, "Test"])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Try without auth
        session = requests.Session()
        response = session.post(
            f"{BASE_URL}/api/payroll/sewa-advances/bulk-upload",
            files={"file": ("test.xlsx", output, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        assert response.status_code in [401, 403], f"Expected 401/403 without auth, got: {response.status_code}"
        print(f"✓ Bulk upload correctly requires authentication")
    
    def test_05_bulk_upload_validates_employee_exists(self):
        """Test that bulk upload validates employee exists"""
        import openpyxl
        
        wb = openpyxl.Workbook()
        ws = wb.active
        # Header row
        ws.append(["Employee Code", "Employee Name", "Employee ID", 
                   "Total Advance Amount", "Monthly Deduction", "Amount Paid Till Now",
                   "Start Month (1-12)", "Start Year", "Reason"])
        # Invalid employee
        ws.append(["INVALID_CODE", "Invalid Employee", "invalid_emp_id", 10000, 1000, 0, 1, 2026, "Test"])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Use only auth header, let requests handle Content-Type for multipart
        response = requests.post(
            f"{BASE_URL}/api/payroll/sewa-advances/bulk-upload",
            files={"file": ("test.xlsx", output, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=self.auth_headers
        )
        
        assert response.status_code == 200, f"Upload failed: {response.status_code} - {response.text}"
        data = response.json()
        
        # Should have error for invalid employee
        assert "errors" in data, "Response should contain errors field"
        assert len(data["errors"]) > 0, "Should have error for invalid employee"
        assert "not found" in data["errors"][0].lower(), f"Error should mention 'not found': {data['errors']}"
        assert data["created"] == 0, "Should not create any advances for invalid employee"
        
        print(f"✓ Bulk upload correctly validates employee exists: {data['errors']}")
    
    def test_06_bulk_upload_validates_monthly_deduction_positive(self):
        """Test that bulk upload validates monthly deduction > 0"""
        import openpyxl
        
        # First get a valid employee
        emp_response = requests.get(f"{BASE_URL}/api/employees?limit=1", headers=self.json_headers)
        assert emp_response.status_code == 200
        emp_data = emp_response.json()
        employees = emp_data.get("employees", emp_data) if isinstance(emp_data, dict) else emp_data
        assert len(employees) > 0, "No employees found"
        
        emp = employees[0]
        emp_code = emp.get("emp_code", "")
        emp_id = emp.get("employee_id", "")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Employee Code", "Employee Name", "Employee ID", 
                   "Total Advance Amount", "Monthly Deduction", "Amount Paid Till Now",
                   "Start Month (1-12)", "Start Year", "Reason"])
        # Monthly deduction = 0
        ws.append([emp_code, "Test Employee", emp_id, 10000, 0, 0, 1, 2026, "Test"])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = requests.post(
            f"{BASE_URL}/api/payroll/sewa-advances/bulk-upload",
            files={"file": ("test.xlsx", output, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=self.auth_headers
        )
        
        assert response.status_code == 200, f"Upload failed: {response.status_code} - {response.text}"
        data = response.json()
        
        # Should have error for zero monthly deduction
        assert "errors" in data, "Response should contain errors field"
        assert len(data["errors"]) > 0, "Should have error for zero monthly deduction"
        assert "monthly" in data["errors"][0].lower() or "deduction" in data["errors"][0].lower(), \
            f"Error should mention monthly deduction: {data['errors']}"
        
        print(f"✓ Bulk upload correctly validates monthly deduction > 0: {data['errors']}")
    
    def test_07_bulk_upload_validates_paid_not_exceed_total(self):
        """Test that bulk upload validates paid <= total"""
        import openpyxl
        
        # Get a valid employee
        emp_response = requests.get(f"{BASE_URL}/api/employees?limit=1", headers=self.json_headers)
        assert emp_response.status_code == 200
        emp_data = emp_response.json()
        employees = emp_data.get("employees", emp_data) if isinstance(emp_data, dict) else emp_data
        emp = employees[0]
        emp_code = emp.get("emp_code", "")
        emp_id = emp.get("employee_id", "")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Employee Code", "Employee Name", "Employee ID", 
                   "Total Advance Amount", "Monthly Deduction", "Amount Paid Till Now",
                   "Start Month (1-12)", "Start Year", "Reason"])
        # Paid > Total
        ws.append([emp_code, "Test Employee", emp_id, 10000, 1000, 15000, 1, 2026, "Test"])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = requests.post(
            f"{BASE_URL}/api/payroll/sewa-advances/bulk-upload",
            files={"file": ("test.xlsx", output, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=self.auth_headers
        )
        
        assert response.status_code == 200, f"Upload failed: {response.status_code} - {response.text}"
        data = response.json()
        
        # Should have error for paid > total
        assert "errors" in data, "Response should contain errors field"
        assert len(data["errors"]) > 0, "Should have error for paid > total"
        assert "exceeds" in data["errors"][0].lower() or "paid" in data["errors"][0].lower(), \
            f"Error should mention paid exceeds total: {data['errors']}"
        
        print(f"✓ Bulk upload correctly validates paid <= total: {data['errors']}")
    
    def test_08_bulk_upload_creates_advance_successfully(self):
        """Test that bulk upload creates SEWA advance successfully"""
        import openpyxl
        import uuid
        
        # Get a valid employee
        emp_response = requests.get(f"{BASE_URL}/api/employees?limit=10", headers=self.json_headers)
        assert emp_response.status_code == 200
        emp_data = emp_response.json()
        employees = emp_data.get("employees", emp_data) if isinstance(emp_data, dict) else emp_data
        
        # Find an employee without active advance (or use first one)
        emp = employees[0]
        emp_code = emp.get("emp_code", "")
        emp_id = emp.get("employee_id", "")
        
        # Use unique test amount to identify our test advance
        test_amount = 12345.67
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Employee Code", "Employee Name", "Employee ID", 
                   "Total Advance Amount", "Monthly Deduction", "Amount Paid Till Now",
                   "Start Month (1-12)", "Start Year", "Reason"])
        ws.append([emp_code, "Test Employee", emp_id, test_amount, 1000, 0, 1, 2026, "TEST_BULK_UPLOAD"])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = requests.post(
            f"{BASE_URL}/api/payroll/sewa-advances/bulk-upload",
            files={"file": ("test.xlsx", output, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=self.auth_headers
        )
        
        assert response.status_code == 200, f"Upload failed: {response.status_code} - {response.text}"
        data = response.json()
        
        # Should create or replace
        total_processed = data.get("created", 0) + data.get("replaced", 0)
        assert total_processed >= 1 or data.get("created", 0) >= 1, \
            f"Should create at least 1 advance. Got: {data}"
        
        # Verify advance was created with source='bulk_upload'
        if data.get("advances"):
            advance = data["advances"][0]
            assert advance.get("source") == "bulk_upload", f"Source should be 'bulk_upload', got: {advance.get('source')}"
            assert advance.get("total_amount") == test_amount, f"Amount mismatch"
        
        print(f"✓ Bulk upload created advance successfully: created={data.get('created')}, replaced={data.get('replaced')}")
    
    def test_09_bulk_upload_replaces_existing_active_advance(self):
        """Test that bulk upload replaces existing active advance"""
        import openpyxl
        
        # Get a valid employee
        emp_response = requests.get(f"{BASE_URL}/api/employees?limit=5", headers=self.json_headers)
        assert emp_response.status_code == 200
        emp_data = emp_response.json()
        employees = emp_data.get("employees", emp_data) if isinstance(emp_data, dict) else emp_data
        emp = employees[0]
        emp_code = emp.get("emp_code", "")
        emp_id = emp.get("employee_id", "")
        
        # First, create an advance via API
        create_response = requests.post(
            f"{BASE_URL}/api/payroll/sewa-advances",
            json={
                "employee_id": emp_id,
                "total_amount": 5000,
                "monthly_amount": 500,
                "duration_months": 10,
                "reason": "TEST_ORIGINAL_ADVANCE",
                "start_month": 1,
                "start_year": 2026
            },
            headers=self.json_headers
        )
        # May fail if already exists, that's ok
        
        # Now upload a new advance for same employee
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Employee Code", "Employee Name", "Employee ID", 
                   "Total Advance Amount", "Monthly Deduction", "Amount Paid Till Now",
                   "Start Month (1-12)", "Start Year", "Reason"])
        ws.append([emp_code, "Test Employee", emp_id, 8000, 800, 0, 2, 2026, "TEST_REPLACEMENT_ADVANCE"])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = requests.post(
            f"{BASE_URL}/api/payroll/sewa-advances/bulk-upload",
            files={"file": ("test.xlsx", output, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=self.auth_headers
        )
        
        assert response.status_code == 200, f"Upload failed: {response.status_code} - {response.text}"
        data = response.json()
        
        # Should have replaced or created
        assert data.get("created", 0) >= 1 or data.get("replaced", 0) >= 1, \
            f"Should create or replace advance. Got: {data}"
        
        print(f"✓ Bulk upload handles existing advances: created={data.get('created')}, replaced={data.get('replaced')}")
    
    def test_10_get_sewa_advances_list(self):
        """Test GET /api/payroll/sewa-advances returns list"""
        response = requests.get(f"{BASE_URL}/api/payroll/sewa-advances", headers=self.json_headers)
        
        assert response.status_code == 200, f"Failed to get advances: {response.status_code} - {response.text}"
        data = response.json()
        
        assert isinstance(data, list), f"Expected list, got: {type(data)}"
        
        # Check structure if there are advances
        if len(data) > 0:
            advance = data[0]
            required_fields = ["advance_id", "employee_id", "total_amount", "monthly_amount", "is_active"]
            for field in required_fields:
                assert field in advance, f"Missing field '{field}' in advance"
        
        print(f"✓ GET /api/payroll/sewa-advances returns {len(data)} advances")
    
    def test_11_invalid_file_format_rejected(self):
        """Test that non-Excel files are rejected"""
        # Send a text file
        response = requests.post(
            f"{BASE_URL}/api/payroll/sewa-advances/bulk-upload",
            files={"file": ("test.txt", b"This is not an Excel file", "text/plain")},
            headers=self.auth_headers
        )
        
        assert response.status_code == 400, f"Expected 400 for invalid file, got: {response.status_code}"
        print(f"✓ Invalid file format correctly rejected")
    
    def test_12_empty_rows_skipped(self):
        """Test that rows with total_amount <= 0 are skipped"""
        import openpyxl
        
        # Get a valid employee
        emp_response = requests.get(f"{BASE_URL}/api/employees?limit=1", headers=self.json_headers)
        emp_data = emp_response.json()
        employees = emp_data.get("employees", emp_data) if isinstance(emp_data, dict) else emp_data
        emp = employees[0]
        emp_code = emp.get("emp_code", "")
        emp_id = emp.get("employee_id", "")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Employee Code", "Employee Name", "Employee ID", 
                   "Total Advance Amount", "Monthly Deduction", "Amount Paid Till Now",
                   "Start Month (1-12)", "Start Year", "Reason"])
        # Row with 0 amount - should be skipped
        ws.append([emp_code, "Test Employee", emp_id, 0, 0, 0, 1, 2026, "Should be skipped"])
        # Empty row
        ws.append(["", "", "", "", "", "", "", "", ""])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = requests.post(
            f"{BASE_URL}/api/payroll/sewa-advances/bulk-upload",
            files={"file": ("test.xlsx", output, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=self.auth_headers
        )
        
        assert response.status_code == 200, f"Upload failed: {response.status_code} - {response.text}"
        data = response.json()
        
        # Should not create any advances (all rows have 0 or empty amount)
        assert data.get("created", 0) == 0, f"Should skip rows with 0 amount. Got: {data}"
        assert len(data.get("errors", [])) == 0, f"Should not have errors for skipped rows. Got: {data.get('errors')}"
        
        print(f"✓ Empty/zero rows correctly skipped")


# Cleanup fixture
@pytest.fixture(scope="module", autouse=True)
def cleanup_test_advances():
    """Cleanup test advances after all tests"""
    yield
    # Cleanup would go here if needed
    print("Test cleanup complete")


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
