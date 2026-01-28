"""Payroll API Routes"""
from fastapi import APIRouter, HTTPException, Request, UploadFile, File
from fastapi.responses import StreamingResponse
from typing import List, Optional
from datetime import datetime, timezone
from calendar import monthrange
import uuid
import io
from motor.motor_asyncio import AsyncIOMotorClient
import os

# Import payroll calculation helpers
from routes.payroll_v2 import (
    process_employee_salary,
    get_calendar_days_in_month,
    generate_payroll_export_data
)

router = APIRouter(prefix="/payroll", tags=["Payroll"])

# Get DB from environment
mongo_url = os.environ.get('MONGO_URL', 'mongodb://localhost:27017')
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ.get('DB_NAME', 'test_database')]


async def get_current_user(request: Request) -> dict:
    """Get current user from session"""
    from server import get_current_user as auth_get_user
    return await auth_get_user(request)


# ==================== SALARY COMPONENTS ====================

@router.get("/components")
async def list_salary_components(request: Request):
    """List all salary components"""
    await get_current_user(request)
    components = await db.salary_components.find({"is_active": True}, {"_id": 0}).to_list(100)
    return components


@router.post("/components")
async def create_salary_component(data: dict, request: Request):
    """Create salary component"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    data["component_id"] = f"comp_{uuid.uuid4().hex[:12]}"
    data["created_at"] = datetime.now(timezone.utc).isoformat()
    data["is_active"] = True
    
    await db.salary_components.insert_one(data)
    data.pop('_id', None)
    return data


# ==================== SALARY TEMPLATES ====================

@router.get("/templates")
async def list_salary_templates(request: Request):
    """List salary templates"""
    await get_current_user(request)
    templates = await db.salary_templates.find({"is_active": True}, {"_id": 0}).to_list(50)
    return templates


@router.post("/templates")
async def create_salary_template(data: dict, request: Request):
    """Create salary template"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    data["template_id"] = f"tmpl_{uuid.uuid4().hex[:12]}"
    data["created_at"] = datetime.now(timezone.utc).isoformat()
    data["is_active"] = True
    
    await db.salary_templates.insert_one(data)
    data.pop('_id', None)
    return data


# ==================== EMPLOYEE SALARY ====================

@router.get("/employee/{employee_id}")
async def get_employee_salary(employee_id: str, request: Request):
    """Get employee salary structure"""
    user = await get_current_user(request)
    
    # Check permission
    if user.get("role") not in ["super_admin", "hr_admin", "finance"]:
        if user.get("employee_id") != employee_id:
            raise HTTPException(status_code=403, detail="Not authorized")
    
    salary = await db.employee_salaries.find_one(
        {"employee_id": employee_id, "is_active": True}, {"_id": 0}
    )
    return salary


@router.post("/employee/{employee_id}")
async def assign_employee_salary(employee_id: str, data: dict, request: Request):
    """Assign salary structure to employee"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Deactivate existing salary
    await db.employee_salaries.update_many(
        {"employee_id": employee_id, "is_active": True},
        {"$set": {"is_active": False, "effective_to": data.get("effective_from")}}
    )
    
    data["salary_id"] = f"sal_{uuid.uuid4().hex[:12]}"
    data["employee_id"] = employee_id
    data["created_at"] = datetime.now(timezone.utc).isoformat()
    data["is_active"] = True
    
    await db.employee_salaries.insert_one(data)
    data.pop('_id', None)
    return data


# ==================== PAYROLL RUN ====================

@router.get("/runs")
async def list_payroll_runs(request: Request, year: Optional[int] = None):
    """List payroll runs"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    query = {}
    if year:
        query["year"] = year
    
    runs = await db.payroll_runs.find(query, {"_id": 0}).sort([("year", -1), ("month", -1)]).to_list(24)
    return runs


@router.post("/runs")
async def create_payroll_run(month: int, year: int, request: Request):
    """Initialize payroll run for a month"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Check if already exists
    existing = await db.payroll_runs.find_one({"month": month, "year": year})
    if existing:
        raise HTTPException(status_code=400, detail="Payroll run already exists for this month")
    
    payroll_run = {
        "payroll_id": f"pr_{uuid.uuid4().hex[:12]}",
        "month": month,
        "year": year,
        "status": "draft",
        "total_employees": 0,
        "total_gross": 0,
        "total_deductions": 0,
        "total_net": 0,
        "total_pf": 0,
        "total_esi": 0,
        "total_pt": 0,
        "total_tds": 0,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.payroll_runs.insert_one(payroll_run)
    payroll_run.pop('_id', None)
    return payroll_run


@router.post("/runs/{payroll_id}/process")
async def process_payroll(payroll_id: str, request: Request):
    """
    Process payroll - calculate all payslips using the new salary structure template logic
    
    New calculation logic:
    - Uses calendar days (28-31) not fixed 26 days
    - Pro-rates each salary component individually
    - Applies WFH at configurable percentage (default 50%)
    - Calculates EPF, ESI, SEWA based on formulas
    - Handles SEWA Advance for applicable employees
    - Applies late deduction rules
    """
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    payroll = await db.payroll_runs.find_one({"payroll_id": payroll_id}, {"_id": 0})
    if not payroll:
        raise HTTPException(status_code=404, detail="Payroll run not found")
    
    if payroll["status"] == "locked":
        raise HTTPException(status_code=400, detail="Payroll is locked")
    
    month, year = payroll["month"], payroll["year"]
    total_days = get_calendar_days_in_month(year, month)
    month_str = f"{year}-{str(month).zfill(2)}"
    
    # Get payroll config with defaults
    config = await db.payroll_config.find_one({"is_active": True}, {"_id": 0})
    payroll_rules = await db.payroll_rules.find_one({"is_active": True}, {"_id": 0})
    
    # Merge config and rules
    payroll_config = {
        "epf_employee_percentage": float((payroll_rules or {}).get("epf_employee_percentage", 12.0)),
        "epf_wage_ceiling": float((config or {}).get("pf_wage_ceiling", 15000)),
        "esi_employee_percentage": float((payroll_rules or {}).get("esi_employee_percentage", 0.75)),
        "esi_wage_ceiling": float((config or {}).get("esi_wage_ceiling", 21000)),
        "sewa_percentage": float((payroll_rules or {}).get("sewa_percentage", 2.0)),
        "wfh_pay_percentage": float((payroll_rules or {}).get("wfh_pay_percentage", 50.0)),
        "late_deduction_enabled": (payroll_rules or {}).get("late_deduction_enabled", True),
        "late_count_threshold": int((payroll_rules or {}).get("late_count_threshold", 2)),
    }
    
    # Get salary data from BOTH collections and merge
    salary_data_map = {}
    
    # First, get from employee_salaries (has fixed_components)
    emp_salaries = await db.employee_salaries.find(
        {"is_active": True}, {"_id": 0}
    ).to_list(1000)
    for sal in emp_salaries:
        emp_id = sal.get("employee_id")
        if emp_id:
            salary_data_map[emp_id] = sal
    
    # Then, get from salary_structures (fallback)
    sal_structures = await db.salary_structures.find(
        {"is_active": True}, {"_id": 0}
    ).to_list(1000)
    for sal in sal_structures:
        emp_id = sal.get("employee_id")
        if emp_id and emp_id not in salary_data_map:
            salary_data_map[emp_id] = sal
    
    # Get all active SEWA advances
    sewa_advances = await db.sewa_advances.find(
        {"is_active": True}, {"_id": 0}
    ).to_list(1000)
    sewa_advance_map = {sa.get("employee_id"): sa for sa in sewa_advances}
    
    # Get one-time deductions for this month
    one_time_deductions = await db.one_time_deductions.find(
        {"month": month, "year": year, "is_active": True}, {"_id": 0}
    ).to_list(1000)
    ot_deductions_map = {}
    for otd in one_time_deductions:
        emp_id = otd.get("employee_id")
        if emp_id not in ot_deductions_map:
            ot_deductions_map[emp_id] = []
        ot_deductions_map[emp_id].append(otd)
    
    # Get holidays for the month (including half-day flag)
    holidays = await db.holidays.find({
        "date": {"$regex": f"^{month_str}"}
    }, {"_id": 0}).to_list(31)
    holiday_dates = {h.get("date"): h for h in holidays}  # Store full holiday object
    
    employees_with_salary = list(salary_data_map.values())
    
    total_gross = 0
    total_deductions = 0
    total_net = 0
    total_pf = 0
    total_esi = 0
    total_sewa = 0
    payslips_created = 0
    
    for emp_salary in employees_with_salary:
        employee_id = emp_salary.get("employee_id")
        if not employee_id:
            continue
            
        # Skip if employee is not active
        employee = await db.employees.find_one({"employee_id": employee_id, "is_active": True})
        if not employee:
            continue
        
        # Get attendance for the month
        attendance = await db.attendance.find(
            {"employee_id": employee_id, "date": {"$regex": f"^{month_str}"}}
        ).to_list(50)
        
        # Calculate attendance breakdown
        office_days = 0
        wfh_days = 0
        leave_days = 0
        late_count = 0
        half_day_count = 0  # NEW: Count half-day attendance
        
        for att in attendance:
            status = att.get("status", "").lower()
            if status in ["present", "tour"]:
                office_days += 1
            elif status == "wfh":
                wfh_days += 1
            elif status in ["leave", "absent"]:
                leave_days += 1
            elif status in ["half_day", "hd", "half-day"]:  # NEW: Half-day status
                half_day_count += 1
            
            if att.get("is_late"):
                late_count += 1
        
        # Calculate Sundays, holidays, and second Saturdays in the month
        sundays_holidays = 0
        half_day_holidays = 0
        second_saturday_count = 0
        
        from routes.payroll_v2 import is_second_saturday
        
        for day in range(1, total_days + 1):
            date_str = f"{year}-{str(month).zfill(2)}-{str(day).zfill(2)}"
            try:
                from datetime import date as dt_date
                d = dt_date(year, month, day)
                
                # Check if it's a second Saturday (half-day work, half-day pay)
                if is_second_saturday(year, month, day):
                    second_saturday_count += 1
                elif d.weekday() == 6:  # Sunday (full day off)
                    sundays_holidays += 1
                elif date_str in holiday_dates:
                    holiday = holiday_dates[date_str]
                    if holiday.get("is_half_day"):
                        half_day_holidays += 1  # Half-day holiday
                    else:
                        sundays_holidays += 1  # Full-day holiday
            except Exception:
                pass
        
        # Build attendance data
        attendance_data = {
            "office_days": office_days,
            "sundays_holidays": sundays_holidays,
            "leave_days": leave_days,
            "wfh_days": wfh_days,
            "late_count": late_count,
            "half_day_count": half_day_count + half_day_holidays,  # Combine half-day attendance + half-day holidays
            "second_saturday_count": second_saturday_count
        }
        
        # Get SEWA advance if applicable
        sewa_advance_info = sewa_advance_map.get(employee_id)
        
        # Get one-time deductions
        emp_ot_deductions = ot_deductions_map.get(employee_id, [])
        
        # Process salary using new calculation
        try:
            payslip_data = process_employee_salary(
                employee_salary=emp_salary,
                attendance_data=attendance_data,
                payroll_config=payroll_config,
                month=month,
                year=year,
                sewa_advance_info=sewa_advance_info,
                one_time_deductions=emp_ot_deductions
            )
        except Exception as e:
            print(f"Error processing salary for {employee_id}: {e}")
            continue
        
        # Skip if no valid salary
        if payslip_data.get("gross_salary", 0) <= 0:
            continue
        
        emp_name = f"{employee.get('first_name', '')} {employee.get('last_name', '')}".strip()
        
        # Create payslip document
        payslip = {
            "payslip_id": f"ps_{uuid.uuid4().hex[:12]}",
            "payroll_id": payroll_id,
            "employee_id": employee_id,
            "emp_code": employee.get("emp_code", ""),
            "employee_name": emp_name,
            "department": employee.get("department_name") or employee.get("department", ""),
            "designation": employee.get("designation") or employee.get("job_title", ""),
            "month": month,
            "year": year,
            
            # New structure - full breakdown
            "fixed_components": payslip_data["fixed_components"],
            "attendance": payslip_data["attendance"],
            "earnings": payslip_data["earnings"],
            "deductions": payslip_data["deductions"],
            
            # Summary fields (for backwards compatibility)
            "working_days": total_days,
            "present_days": office_days + wfh_days,
            "lwp_days": leave_days,
            "paid_days": payslip_data["attendance"]["total_earned_days"],
            "basic": payslip_data["earnings"]["basic_da_earned"],
            "hra": payslip_data["earnings"]["hra_earned"],
            "special_allowance": payslip_data["earnings"]["other_allowance_earned"] + payslip_data["earnings"]["medical_allowance_earned"],
            "gross_salary": payslip_data["gross_salary"],
            "pf_employee": payslip_data["deductions"]["epf"],
            "esi_employee": payslip_data["deductions"]["esi"],
            "sewa": payslip_data["deductions"]["sewa"],
            "sewa_advance": payslip_data["deductions"]["sewa_advance"],
            "other_deduction": payslip_data["deductions"]["other_deduction"],
            "total_deductions": payslip_data["total_deductions"],
            "net_salary": payslip_data["net_payable"],
            
            "status": "draft",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "config_used": payslip_data["config_used"]
        }
        
        # Upsert payslip
        await db.payslips.update_one(
            {"employee_id": employee_id, "month": month, "year": year},
            {"$set": payslip},
            upsert=True
        )
        
        # Update SEWA advance tracking if applicable
        if sewa_advance_info and sewa_advance_info.get("is_active"):
            sewa_amount = payslip_data["deductions"]["sewa_advance"]
            new_paid = float(sewa_advance_info.get("total_paid", 0)) + sewa_amount
            new_remaining = float(sewa_advance_info.get("total_amount", 0)) - new_paid
            
            update_data = {
                "total_paid": new_paid,
                "remaining_amount": max(0, new_remaining),
                "last_deduction_month": month,
                "last_deduction_year": year,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
            
            # Auto-complete if fully paid
            if new_remaining <= 0:
                update_data["is_active"] = False
                update_data["status"] = "completed"
                update_data["completed_at"] = datetime.now(timezone.utc).isoformat()
            
            await db.sewa_advances.update_one(
                {"employee_id": employee_id, "is_active": True},
                {"$set": update_data}
            )
        
        total_gross += payslip_data["gross_salary"]
        total_deductions += payslip_data["total_deductions"]
        total_net += payslip_data["net_payable"]
        total_pf += payslip_data["deductions"]["epf"]
        total_esi += payslip_data["deductions"]["esi"]
        total_sewa += payslip_data["deductions"]["sewa"] + payslip_data["deductions"]["sewa_advance"]
        payslips_created += 1
    
    # Update payroll run
    await db.payroll_runs.update_one(
        {"payroll_id": payroll_id},
        {"$set": {
            "status": "processed",
            "total_employees": payslips_created,
            "total_gross": round(total_gross, 0),
            "total_deductions": round(total_deductions, 0),
            "total_net": round(total_net, 0),
            "total_pf": round(total_pf, 0),
            "total_esi": round(total_esi, 0),
            "total_sewa": round(total_sewa, 0),
            "calendar_days": total_days,
            "processed_by": user["user_id"],
            "processed_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": f"Processed {payslips_created} payslips", "payroll_id": payroll_id}


@router.post("/runs/{payroll_id}/lock")
async def lock_payroll(payroll_id: str, request: Request):
    """Lock payroll after verification"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    await db.payroll_runs.update_one(
        {"payroll_id": payroll_id},
        {"$set": {
            "status": "locked",
            "locked_by": user["user_id"],
            "locked_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "Payroll locked successfully"}


@router.delete("/runs/{payroll_id}")
async def delete_payroll_run(payroll_id: str, request: Request):
    """Delete a payroll run and all associated payslips (processed or locked)"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Check if payroll exists
    payroll = await db.payroll_runs.find_one({"payroll_id": payroll_id})
    if not payroll:
        raise HTTPException(status_code=404, detail="Payroll run not found")
    
    # Delete all associated payslips
    delete_result = await db.payslips.delete_many({"payroll_id": payroll_id})
    payslips_deleted = delete_result.deleted_count
    
    # Delete the payroll run
    await db.payroll_runs.delete_one({"payroll_id": payroll_id})
    
    return {
        "message": "Payroll run deleted successfully",
        "payslips_deleted": payslips_deleted,
        "month": payroll.get("month"),
        "year": payroll.get("year")
    }


@router.get("/runs/{payroll_id}")
async def get_payroll_run_details(payroll_id: str, request: Request):
    """Get detailed payroll run with all payslips"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get payroll run
    payroll = await db.payroll_runs.find_one({"payroll_id": payroll_id}, {"_id": 0})
    if not payroll:
        raise HTTPException(status_code=404, detail="Payroll run not found")
    
    month, year = payroll["month"], payroll["year"]
    
    # Get all payslips for this payroll run
    payslips = await db.payslips.find(
        {"payroll_id": payroll_id}, {"_id": 0}
    ).to_list(1000)
    
    # Also try to find by month/year if payroll_id not set on all payslips
    if len(payslips) < payroll.get("total_employees", 0):
        additional_payslips = await db.payslips.find(
            {"month": month, "year": year, "payroll_id": {"$exists": False}}, {"_id": 0}
        ).to_list(1000)
        payslips.extend(additional_payslips)
    
    # Enrich with employee details
    enriched_payslips = []
    for slip in payslips:
        employee = await db.employees.find_one(
            {"employee_id": slip["employee_id"]}, {"_id": 0}
        )
        if employee:
            slip["employee_name"] = f"{employee.get('first_name', '')} {employee.get('last_name', '')}".strip()
            slip["emp_code"] = employee.get("emp_code", slip["employee_id"])
            slip["department"] = employee.get("department_name") or employee.get("department", "")
            slip["designation"] = employee.get("designation") or employee.get("job_title", "")
        enriched_payslips.append(slip)
    
    # Sort by employee name
    enriched_payslips.sort(key=lambda x: x.get("employee_name", ""))
    
    return {
        "payroll": payroll,
        "payslips": enriched_payslips,
        "summary": {
            "total_employees": len(enriched_payslips),
            "total_gross": sum(p.get("gross_salary", 0) for p in enriched_payslips),
            "total_deductions": sum(p.get("total_deductions", 0) for p in enriched_payslips),
            "total_net": sum(p.get("net_salary", 0) for p in enriched_payslips),
            "total_pf": sum(p.get("pf_employee", 0) for p in enriched_payslips),
            "total_esi": sum(p.get("esi_employee", 0) for p in enriched_payslips),
            "total_pt": sum(p.get("professional_tax", 0) for p in enriched_payslips),
        }
    }


# ==================== PAYSLIPS ====================

@router.get("/payslips")
async def list_payslips(
    request: Request,
    month: Optional[int] = None,
    year: Optional[int] = None,
    employee_id: Optional[str] = None
):
    """List payslips"""
    user = await get_current_user(request)
    
    query = {}
    if month:
        query["month"] = month
    if year:
        query["year"] = year
    
    # Regular employees can only see their own
    if user.get("role") not in ["super_admin", "hr_admin", "finance"]:
        query["employee_id"] = user.get("employee_id")
    elif employee_id:
        query["employee_id"] = employee_id
    
    payslips = await db.payslips.find(query, {"_id": 0}).sort([("year", -1), ("month", -1)]).to_list(100)
    return payslips


@router.get("/payslips/{payslip_id}")
async def get_payslip(payslip_id: str, request: Request):
    """Get single payslip"""
    user = await get_current_user(request)
    
    payslip = await db.payslips.find_one({"payslip_id": payslip_id}, {"_id": 0})
    if not payslip:
        raise HTTPException(status_code=404, detail="Payslip not found")
    
    # Check access
    if user.get("role") not in ["super_admin", "hr_admin", "finance"]:
        if payslip["employee_id"] != user.get("employee_id"):
            raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get employee details
    employee = await db.employees.find_one({"employee_id": payslip["employee_id"]}, {"_id": 0})
    payslip["employee"] = employee
    
    return payslip


@router.get("/my-payslips")
async def get_my_payslips(request: Request):
    """Get current user's payslips"""
    user = await get_current_user(request)
    employee_id = user.get("employee_id")
    
    if not employee_id:
        return []
    
    payslips = await db.payslips.find(
        {"employee_id": employee_id}, {"_id": 0}
    ).sort([("year", -1), ("month", -1)]).to_list(24)
    
    return payslips


@router.get("/employee/{employee_id}")
async def get_employee_salary_details_v2(employee_id: str, request: Request):
    """Get salary structure for an employee"""
    user = await get_current_user(request)
    
    # Check access - HR/Admin or self
    if user.get("role") not in ["super_admin", "hr_admin", "finance", "hr_executive"]:
        if user.get("employee_id") != employee_id:
            raise HTTPException(status_code=403, detail="Not authorized")
    
    # First try employee_salaries collection
    salary = await db.employee_salaries.find_one(
        {"employee_id": employee_id, "is_active": True}, {"_id": 0}
    )
    
    if not salary:
        # Try salary_structures collection
        salary = await db.salary_structures.find_one(
            {"employee_id": employee_id, "is_active": True}, {"_id": 0}
        )
    
    if not salary:
        # Try to build from employee record
        employee = await db.employees.find_one({"employee_id": employee_id}, {"_id": 0})
        if employee and employee.get("salary_info"):
            salary = employee.get("salary_info")
    
    return salary or {}


# ==================== EMPLOYEE SALARY STRUCTURE MANAGEMENT ====================

@router.put("/employee/{employee_id}/salary")
async def update_employee_salary(employee_id: str, data: dict, request: Request):
    """Update or create salary structure for an employee (creates change request if approval required)"""
    user = await get_current_user(request)
    
    # Only HR/Admin/Finance can update salary
    if user.get("role") not in ["super_admin", "hr_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Check if employee exists
    employee = await db.employees.find_one({"employee_id": employee_id})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Get current salary for history
    current_salary = await db.employee_salaries.find_one(
        {"employee_id": employee_id, "is_active": True}, {"_id": 0}
    )
    
    # Build the new salary structure based on the Excel format
    new_salary = {
        "salary_id": f"sal_{uuid.uuid4().hex[:12]}",
        "employee_id": employee_id,
        "emp_code": data.get("emp_code") or employee.get("emp_code"),
        "employee_name": data.get("employee_name") or f"{employee.get('first_name', '')} {employee.get('last_name', '')}".strip(),
        
        # Fixed Components (Earnings)
        "fixed_components": {
            "basic": float(data.get("basic", 0)),
            "da": float(data.get("da", 0)),
            "hra": float(data.get("hra", 0)),
            "conveyance": float(data.get("conveyance", 0)),
            "grade_pay": float(data.get("grade_pay", 0)),
            "other_allowance": float(data.get("other_allowance", 0)),
            "medical_allowance": float(data.get("medical_allowance", 0)),
        },
        "total_fixed": float(data.get("total_fixed", 0)) or (
            float(data.get("basic", 0)) + float(data.get("da", 0)) + 
            float(data.get("hra", 0)) + float(data.get("conveyance", 0)) +
            float(data.get("grade_pay", 0)) + float(data.get("other_allowance", 0)) +
            float(data.get("medical_allowance", 0))
        ),
        
        # Deduction Configuration
        "deduction_config": {
            "epf_applicable": data.get("epf_applicable", True),
            "epf_percentage": float(data.get("epf_percentage", 12)),  # 12% of Basic
            "esi_applicable": data.get("esi_applicable", True),
            "esi_percentage": float(data.get("esi_percentage", 0.75)),  # 0.75% of Gross
            "esi_ceiling": float(data.get("esi_ceiling", 21000)),  # ESI only if gross < 21000
            "sewa_applicable": data.get("sewa_applicable", True),
            "sewa_amount": float(data.get("sewa_amount", 0)),
            "sewa_percentage": float(data.get("sewa_percentage", 2)),  # 2% of Basic
            "professional_tax": float(data.get("professional_tax", 0)),
            "lwf": float(data.get("lwf", 0)),
        },
        
        # Fixed deductions
        "fixed_deductions": {
            "sewa_advance": float(data.get("sewa_advance", 0)),
            "other_deduction": float(data.get("other_deduction", 0)),
        },
        
        "effective_from": data.get("effective_from") or datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "is_active": True,
        "created_by": user.get("user_id"),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    
    # Check if approval is required (based on payroll rules)
    payroll_rules = await db.payroll_rules.find_one({"is_active": True})
    requires_approval = payroll_rules.get("salary_change_requires_approval", True) if payroll_rules else True
    
    if requires_approval and user.get("role") != "super_admin":
        # Create a salary change request
        change_request = {
            "request_id": f"scr_{uuid.uuid4().hex[:12]}",
            "employee_id": employee_id,
            "emp_code": new_salary["emp_code"],
            "employee_name": new_salary["employee_name"],
            "current_salary": current_salary,
            "new_salary": new_salary,
            "requested_by": user.get("user_id"),
            "requested_by_name": user.get("name"),
            "requested_at": datetime.now(timezone.utc).isoformat(),
            "status": "pending",
            "reason": data.get("reason", "Salary structure update"),
        }
        
        await db.salary_change_requests.insert_one(change_request)
        change_request.pop("_id", None)
        
        return {
            "message": "Salary change request submitted for approval",
            "request": change_request
        }
    else:
        # Direct update (super_admin or no approval required)
        # Deactivate old salary
        if current_salary:
            await db.employee_salaries.update_one(
                {"salary_id": current_salary.get("salary_id")},
                {"$set": {"is_active": False, "deactivated_at": datetime.now(timezone.utc).isoformat()}}
            )
            
            # Save to history
            history_entry = {
                "history_id": f"sh_{uuid.uuid4().hex[:12]}",
                "employee_id": employee_id,
                "old_salary": current_salary,
                "new_salary": new_salary,
                "changed_by": user.get("user_id"),
                "changed_by_name": user.get("name"),
                "changed_at": datetime.now(timezone.utc).isoformat(),
                "reason": data.get("reason", "Salary structure update"),
            }
            await db.salary_change_history.insert_one(history_entry)
        
        # Insert new salary
        await db.employee_salaries.insert_one(new_salary)
        new_salary.pop("_id", None)
        
        return {
            "message": "Salary structure updated successfully",
            "salary": new_salary
        }


@router.get("/salary-change-requests")
async def get_salary_change_requests(
    request: Request,
    status: Optional[str] = None,
    skip: int = 0,
    limit: int = 50
):
    """Get pending salary change requests for approval"""
    user = await get_current_user(request)
    
    if user.get("role") not in ["super_admin", "hr_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    query = {}
    if status:
        query["status"] = status
    
    requests = await db.salary_change_requests.find(query, {"_id": 0}).sort("requested_at", -1).skip(skip).to_list(limit)
    total = await db.salary_change_requests.count_documents(query)
    
    return {"total": total, "requests": requests}


@router.put("/salary-change-requests/{request_id}/approve")
async def approve_salary_change(request_id: str, request: Request):
    """Approve a salary change request"""
    user = await get_current_user(request)
    
    # Only super_admin or finance can approve
    if user.get("role") not in ["super_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Only Super Admin or Finance can approve salary changes")
    
    # Get the request
    change_request = await db.salary_change_requests.find_one({"request_id": request_id})
    if not change_request:
        raise HTTPException(status_code=404, detail="Change request not found")
    
    if change_request.get("status") != "pending":
        raise HTTPException(status_code=400, detail=f"Request is already {change_request.get('status')}")
    
    employee_id = change_request["employee_id"]
    new_salary = change_request["new_salary"]
    current_salary = change_request.get("current_salary")
    
    # Deactivate old salary
    if current_salary:
        await db.employee_salaries.update_one(
            {"salary_id": current_salary.get("salary_id")},
            {"$set": {"is_active": False, "deactivated_at": datetime.now(timezone.utc).isoformat()}}
        )
    
    # Save to history
    history_entry = {
        "history_id": f"sh_{uuid.uuid4().hex[:12]}",
        "employee_id": employee_id,
        "old_salary": current_salary,
        "new_salary": new_salary,
        "changed_by": change_request.get("requested_by"),
        "changed_by_name": change_request.get("requested_by_name"),
        "approved_by": user.get("user_id"),
        "approved_by_name": user.get("name"),
        "changed_at": datetime.now(timezone.utc).isoformat(),
        "reason": change_request.get("reason"),
    }
    await db.salary_change_history.insert_one(history_entry)
    
    # Insert new salary
    new_salary["approved_by"] = user.get("user_id")
    new_salary["approved_at"] = datetime.now(timezone.utc).isoformat()
    await db.employee_salaries.insert_one(new_salary)
    
    # Update request status
    await db.salary_change_requests.update_one(
        {"request_id": request_id},
        {"$set": {
            "status": "approved",
            "approved_by": user.get("user_id"),
            "approved_by_name": user.get("name"),
            "approved_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "Salary change approved and applied successfully"}


@router.put("/salary-change-requests/{request_id}/reject")
async def reject_salary_change(request_id: str, data: dict, request: Request):
    """Reject a salary change request"""
    user = await get_current_user(request)
    
    if user.get("role") not in ["super_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Only Super Admin or Finance can reject salary changes")
    
    change_request = await db.salary_change_requests.find_one({"request_id": request_id})
    if not change_request:
        raise HTTPException(status_code=404, detail="Change request not found")
    
    if change_request.get("status") != "pending":
        raise HTTPException(status_code=400, detail=f"Request is already {change_request.get('status')}")
    
    await db.salary_change_requests.update_one(
        {"request_id": request_id},
        {"$set": {
            "status": "rejected",
            "rejected_by": user.get("user_id"),
            "rejected_by_name": user.get("name"),
            "rejected_at": datetime.now(timezone.utc).isoformat(),
            "rejection_reason": data.get("reason", "")
        }}
    )
    
    return {"message": "Salary change request rejected"}


@router.get("/employee/{employee_id}/salary-history")
async def get_employee_salary_history(employee_id: str, request: Request, limit: int = 20):
    """Get salary change history for an employee"""
    user = await get_current_user(request)
    
    if user.get("role") not in ["super_admin", "hr_admin", "finance", "hr_executive"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    history = await db.salary_change_history.find(
        {"employee_id": employee_id}, {"_id": 0}
    ).sort("changed_at", -1).to_list(limit)
    
    return history


@router.get("/payslips")
async def get_payslips_by_employee(
    request: Request,
    employee_id: Optional[str] = None,
    limit: int = 24
):
    """Get payslips for a specific employee (HR/Admin only)"""
    user = await get_current_user(request)
    
    if user.get("role") not in ["super_admin", "hr_admin", "finance", "hr_executive"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    query = {}
    if employee_id:
        query["employee_id"] = employee_id
    
    payslips = await db.payslips.find(query, {"_id": 0}).sort([("year", -1), ("month", -1)]).to_list(limit)
    return payslips


# ==================== ALL EMPLOYEES PAY INFO (HR/Admin only) ====================

@router.get("/all-employees-pay")
async def get_all_employees_pay_info(
    request: Request,
    month: Optional[int] = None,
    year: Optional[int] = None
):
    """Get pay info for all employees (HR/Admin only)"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Not authorized - HR/Admin only")
    
    # Default to current month/year if not provided
    if not month:
        month = datetime.now().month
    if not year:
        year = datetime.now().year
    
    # Get all payslips for the month
    payslips = await db.payslips.find(
        {"month": month, "year": year}, {"_id": 0}
    ).to_list(1000)
    
    # Enrich with employee details
    enriched_payslips = []
    for slip in payslips:
        employee = await db.employees.find_one(
            {"employee_id": slip["employee_id"]}, {"_id": 0}
        )
        if employee:
            slip["employee_name"] = f"{employee.get('first_name', '')} {employee.get('last_name', '')}".strip()
            slip["employee_code"] = employee.get("employee_code", slip["employee_id"])
            slip["department"] = employee.get("department", "")
            slip["designation"] = employee.get("designation", "")
        enriched_payslips.append(slip)
    
    return enriched_payslips


@router.get("/all-salary-structures")
async def get_all_salary_structures(
    request: Request,
    department_id: Optional[str] = None,
    search: Optional[str] = None,
    skip: int = 0,
    limit: int = 100
):
    """Get all employees' salary structures (HR/Admin only)"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "finance", "hr_executive"]:
        raise HTTPException(status_code=403, detail="Not authorized - HR/Admin only")
    
    # Get all active employees first
    emp_query = {"is_active": True}
    if department_id and department_id != "all":
        emp_query["department_id"] = department_id
    
    employees = await db.employees.find(emp_query, {"_id": 0}).to_list(1000)
    emp_map = {e.get("employee_id"): e for e in employees}
    emp_ids = list(emp_map.keys())
    
    # Get salary data from both collections
    salary_data = {}
    
    # From employee_salaries
    emp_salaries = await db.employee_salaries.find(
        {"employee_id": {"$in": emp_ids}, "is_active": True}, {"_id": 0}
    ).to_list(1000)
    for sal in emp_salaries:
        emp_id = sal.get("employee_id")
        if emp_id:
            salary_data[emp_id] = sal
    
    # From salary_structures (for those not in employee_salaries)
    sal_structures = await db.salary_structures.find(
        {"employee_id": {"$in": emp_ids}, "is_active": True}, {"_id": 0}
    ).to_list(1000)
    for sal in sal_structures:
        emp_id = sal.get("employee_id")
        if emp_id and emp_id not in salary_data:
            salary_data[emp_id] = sal
    
    # Build result with employee details
    result = []
    for emp_id, emp in emp_map.items():
        sal = salary_data.get(emp_id, {})
        
        # Calculate gross from different structures
        gross = sal.get("total_fixed") or sal.get("gross") or (sal.get("ctc", 0) / 12)
        basic = None
        if sal.get("fixed_components"):
            basic = sal["fixed_components"].get("basic")
        else:
            basic = sal.get("basic")
        
        emp_name = f"{emp.get('first_name', '')} {emp.get('last_name', '')}".strip()
        
        # Apply search filter
        if search:
            search_lower = search.lower()
            if not (search_lower in emp_name.lower() or 
                    search_lower in (emp.get("emp_code") or "").lower() or
                    search_lower in (emp.get("email") or "").lower()):
                continue
        
        result.append({
            "employee_id": emp_id,
            "emp_code": emp.get("emp_code"),
            "employee_name": emp_name,
            "department": emp.get("department_name") or emp.get("department"),
            "designation": emp.get("designation") or emp.get("job_title"),
            "employment_type": emp.get("employment_type"),
            "gross_salary": round(gross, 0) if gross else None,
            "basic_salary": round(basic, 0) if basic else None,
            "ctc": sal.get("ctc") or sal.get("annual_ctc"),
            "has_salary_data": bool(sal),
            "effective_from": sal.get("effective_from"),
            "salary_source": "fixed_components" if sal.get("fixed_components") else ("gross" if sal.get("gross") else ("ctc" if sal.get("ctc") else None))
        })
    
    # Sort by employee name
    result.sort(key=lambda x: x.get("employee_name", ""))
    
    # Apply pagination
    total = len(result)
    result = result[skip:skip + limit]
    
    return {
        "total": total,
        "skip": skip,
        "limit": limit,
        "data": result
    }


@router.get("/employee-salary-details/{employee_id}")
async def get_employee_salary_details(employee_id: str, request: Request):
    """Get detailed salary info for an employee (HR/Admin only)"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Not authorized - HR/Admin only")
    
    # Get employee
    employee = await db.employees.find_one({"employee_id": employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Get salary structure
    salary = await db.employee_salaries.find_one(
        {"employee_id": employee_id, "is_active": True}, {"_id": 0}
    )
    
    # Get recent payslips (last 12 months)
    payslips = await db.payslips.find(
        {"employee_id": employee_id}, {"_id": 0}
    ).sort([("year", -1), ("month", -1)]).to_list(12)
    
    return {
        "employee": employee,
        "salary_structure": salary,
        "payslips": payslips
    }


@router.get("/employee-breakdown/{employee_id}")
async def get_employee_salary_breakdown(
    employee_id: str, 
    month: int, 
    year: int,
    request: Request
):
    """Get detailed salary breakdown for an employee for a specific month"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Not authorized - HR/Admin only")
    
    # Get employee
    employee = await db.employees.find_one({"employee_id": employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Get payslip for this month
    payslip = await db.payslips.find_one(
        {"employee_id": employee_id, "month": month, "year": year}, {"_id": 0}
    )
    
    # Get salary structure
    salary_structure = await db.employee_salaries.find_one(
        {"employee_id": employee_id, "is_active": True}, {"_id": 0}
    )
    
    # Get attendance records for the month
    start_date = f"{year}-{month:02d}-01"
    if month == 12:
        end_date = f"{year + 1}-01-01"
    else:
        end_date = f"{year}-{month + 1:02d}-01"
    
    attendance_records = await db.attendance.find({
        "employee_id": employee_id,
        "date": {"$gte": start_date, "$lt": end_date}
    }, {"_id": 0}).to_list(31)
    
    # Analyze attendance
    present_days = len([a for a in attendance_records if a.get("status") == "present"])
    absent_days = len([a for a in attendance_records if a.get("status") == "absent"])
    half_days = len([a for a in attendance_records if a.get("status") == "half_day"])
    late_arrivals = len([a for a in attendance_records if a.get("is_late")])
    early_departures = len([a for a in attendance_records if a.get("is_early_departure")])
    
    # Get leave records for the month
    leave_records = await db.leave_requests.find({
        "employee_id": employee_id,
        "status": "approved",
        "$or": [
            {"start_date": {"$gte": start_date, "$lt": end_date}},
            {"end_date": {"$gte": start_date, "$lt": end_date}}
        ]
    }, {"_id": 0}).to_list(20)
    
    # Categorize leaves
    paid_leave_days = 0
    unpaid_leave_days = 0
    leave_breakdown = []
    
    for leave in leave_records:
        leave_type = await db.leave_types.find_one(
            {"leave_type_id": leave.get("leave_type_id")}, {"_id": 0}
        )
        days = leave.get("total_days", 1)
        leave_info = {
            "leave_type": leave_type.get("name") if leave_type else "Unknown",
            "start_date": leave.get("start_date"),
            "end_date": leave.get("end_date"),
            "days": days,
            "is_paid": leave_type.get("is_paid", False) if leave_type else False
        }
        leave_breakdown.append(leave_info)
        
        if leave_type and leave_type.get("is_paid", False):
            paid_leave_days += days
        else:
            unpaid_leave_days += days
    
    # Get payroll config for rules
    config = await db.payroll_config.find_one({"is_active": True}, {"_id": 0})
    working_days = config.get("working_days_per_month", 26) if config else 26
    
    # Calculate deductions breakdown
    gross_salary = payslip.get("gross_salary", 0) if payslip else (salary_structure.get("gross", 0) if salary_structure else 0)
    per_day_salary = gross_salary / working_days if working_days > 0 else 0
    
    deductions_breakdown = {
        "statutory": {
            "pf_employee": payslip.get("pf_employee", 0) if payslip else 0,
            "esi_employee": payslip.get("esi_employee", 0) if payslip else 0,
            "professional_tax": payslip.get("professional_tax", 0) if payslip else 0,
            "tds": payslip.get("tds", 0) if payslip else 0,
        },
        "attendance_based": {
            "lwp_deduction": (unpaid_leave_days * per_day_salary) if unpaid_leave_days > 0 else 0,
            "absent_deduction": (absent_days * per_day_salary) if absent_days > 0 else 0,
            "half_day_deduction": (half_days * per_day_salary * 0.5) if half_days > 0 else 0,
        },
        "other": payslip.get("other_deductions", 0) if payslip else 0
    }
    
    earnings_breakdown = {
        "basic": payslip.get("basic", 0) if payslip else 0,
        "hra": payslip.get("hra", 0) if payslip else 0,
        "special_allowance": payslip.get("special_allowance", 0) if payslip else 0,
        "conveyance": payslip.get("conveyance", 0) if payslip else 0,
        "medical": payslip.get("medical", 0) if payslip else 0,
        "overtime": payslip.get("overtime_pay", 0) if payslip else 0,
        "bonus": payslip.get("bonus", 0) if payslip else 0,
    }
    
    return {
        "employee": {
            "employee_id": employee_id,
            "name": f"{employee.get('first_name', '')} {employee.get('last_name', '')}".strip(),
            "employee_code": employee.get("employee_code", employee_id),
            "department": employee.get("department", ""),
            "designation": employee.get("designation", "")
        },
        "period": {
            "month": month,
            "year": year,
            "working_days": working_days
        },
        "attendance_summary": {
            "present_days": present_days,
            "absent_days": absent_days,
            "half_days": half_days,
            "late_arrivals": late_arrivals,
            "early_departures": early_departures,
            "paid_leave_days": paid_leave_days,
            "unpaid_leave_days": unpaid_leave_days,
            "effective_working_days": present_days + half_days * 0.5 + paid_leave_days
        },
        "leave_breakdown": leave_breakdown,
        "earnings_breakdown": earnings_breakdown,
        "deductions_breakdown": deductions_breakdown,
        "totals": {
            "gross_salary": payslip.get("gross_salary", 0) if payslip else gross_salary,
            "total_earnings": sum(earnings_breakdown.values()),
            "total_statutory_deductions": sum(deductions_breakdown["statutory"].values()),
            "total_attendance_deductions": sum(deductions_breakdown["attendance_based"].values()),
            "total_deductions": payslip.get("total_deductions", 0) if payslip else 0,
            "net_salary": payslip.get("net_salary", 0) if payslip else 0
        }
    }


# ==================== PAYROLL CONFIG & RULES ====================

def get_default_payroll_rules():
    """Return default payroll rules configuration"""
    return {
        # General Settings
        "working_days_per_month": 26,
        "pay_cycle": "monthly",  # monthly, bi-weekly, weekly
        "salary_credit_day": 1,  # Day of month for salary credit
        
        # Attendance Rules
        "attendance_rules": {
            "grace_period_minutes": 15,  # Minutes allowed after shift start
            "late_coming_deduction": {
                "enabled": True,
                "occurrences_before_deduction": 3,  # Free late comings per month
                "deduction_type": "half_day",  # half_day, hourly, fixed
                "deduction_amount": 0,  # Used if deduction_type is "fixed"
            },
            "early_leaving_deduction": {
                "enabled": True,
                "min_early_minutes": 30,  # Minutes early to count as early leaving
                "deduction_type": "half_day",
            },
            "half_day_rules": {
                "min_hours_for_full_day": 8,
                "min_hours_for_half_day": 4,
                "half_day_deduction_percent": 50,  # Percentage of daily wage deducted
            },
            "absent_deduction": {
                "deduction_type": "full_day",  # full_day, 1.5_day, 2_day
                "multiplier": 1.0,  # 1.0 = 1 day, 1.5 = 1.5 days, 2.0 = 2 days
            },
        },
        
        # Leave Rules for Payroll
        "leave_rules": {
            "CL": {"is_paid": True, "deduction_percent": 0},  # Casual Leave
            "SL": {"is_paid": True, "deduction_percent": 0},  # Sick Leave
            "EL": {"is_paid": True, "deduction_percent": 0},  # Earned Leave
            "PL": {"is_paid": True, "deduction_percent": 0},  # Privilege Leave
            "ML": {"is_paid": True, "deduction_percent": 0},  # Maternity Leave
            "PTL": {"is_paid": True, "deduction_percent": 0},  # Paternity Leave
            "CO": {"is_paid": True, "deduction_percent": 0},  # Compensatory Off
            "LWP": {"is_paid": False, "deduction_percent": 100},  # Leave Without Pay
            "default": {"is_paid": False, "deduction_percent": 100},
        },
        
        # Leave Policy Rules (Annual quotas, carry forward, Sunday rules)
        "leave_policy_rules": {
            "financial_year_start": "04-01",  # April 1st
            "annual_quotas": {
                "CL": 6,   # Casual Leave per year
                "SL": 6,   # Sick Leave per year
                "EL": 12,  # Earned Leave per year
            },
            "carry_forward": {
                "CL": False,  # CL lapses
                "SL": False,  # SL lapses
                "EL": True,   # EL carries forward
                "max_el_accumulation": 30,  # Max EL that can be accumulated
            },
            "sunday_leave_rules": {
                "enabled": True,
                "weekly_threshold": 2,     # If >2 leaves in a week, count 1 Sunday as leave
                "monthly_threshold": 6,    # If >6 leaves in a month, count 1 Sunday as leave
                "auto_apply": True,        # Auto-apply with HR warning
            },
        },
        
        # Overtime Rules
        "overtime_rules": {
            "enabled": True,
            "min_hours_for_ot": 1,  # Minimum OT hours to qualify
            "weekday_multiplier": 1.5,  # 1.5x for weekday OT
            "weekend_multiplier": 2.0,  # 2x for weekend OT
            "holiday_multiplier": 2.5,  # 2.5x for holiday OT
            "max_ot_hours_per_day": 4,
            "max_ot_hours_per_month": 50,
        },
        
        # PF (Provident Fund) Configuration
        "pf_rules": {
            "enabled": True,
            "employee_contribution_percent": 12.0,
            "employer_contribution_percent": 12.0,
            "wage_ceiling": 15000,  # Max basic for PF calculation
            "include_basic_only": True,  # PF on basic only or gross
            "admin_charges_percent": 0.5,
            "edli_charges_percent": 0.5,
        },
        
        # ESI (Employee State Insurance) Configuration
        "esi_rules": {
            "enabled": True,
            "employee_contribution_percent": 0.75,
            "employer_contribution_percent": 3.25,
            "wage_ceiling": 21000,  # Max gross for ESI eligibility
        },
        
        # Professional Tax Slabs
        "pt_rules": {
            "enabled": True,
            "state": "Maharashtra",  # State for PT calculation
            "slabs": [
                {"min": 0, "max": 10000, "amount": 0},
                {"min": 10001, "max": 15000, "amount": 150},
                {"min": 15001, "max": 999999999, "amount": 200},
            ],
        },
        
        # TDS (Tax Deducted at Source) Rules
        "tds_rules": {
            "enabled": True,
            "standard_deduction": 50000,
            "tax_slabs_old_regime": [
                {"min": 0, "max": 250000, "rate": 0},
                {"min": 250001, "max": 500000, "rate": 5},
                {"min": 500001, "max": 1000000, "rate": 20},
                {"min": 1000001, "max": 999999999, "rate": 30},
            ],
            "tax_slabs_new_regime": [
                {"min": 0, "max": 300000, "rate": 0},
                {"min": 300001, "max": 600000, "rate": 5},
                {"min": 600001, "max": 900000, "rate": 10},
                {"min": 900001, "max": 1200000, "rate": 15},
                {"min": 1200001, "max": 1500000, "rate": 20},
                {"min": 1500001, "max": 999999999, "rate": 30},
            ],
            "cess_percent": 4,
            "rebate_limit": 500000,
        },
        
        # Salary Components Rules
        "salary_components": {
            "basic_percent_of_ctc": 40,  # Basic as % of CTC
            "hra_percent_of_basic": 40,  # HRA as % of Basic
            "special_allowance_auto_calculate": True,  # Auto-calculate remaining as special allowance
        },
        
        # Bonus & Incentives
        "bonus_rules": {
            "statutory_bonus_enabled": True,
            "statutory_bonus_percent": 8.33,
            "statutory_bonus_ceiling": 7000,  # Monthly basic ceiling
            "performance_bonus_enabled": False,
            "annual_increment_month": 4,  # April
        },
        
        # Loan & Advance Deductions
        "loan_deduction_rules": {
            "max_emi_percent_of_salary": 50,  # Max EMI deduction as % of net salary
            "salary_advance_max_percent": 50,  # Max advance as % of salary
        },
        
        # Reimbursements
        "reimbursement_rules": {
            "process_with_salary": True,
            "max_pending_months": 3,
        },
    }


@router.get("/config")
async def get_payroll_config(request: Request):
    """Get payroll configuration with all rules"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    config = await db.payroll_config.find_one({"is_active": True}, {"_id": 0})
    
    # Return default rules if no config exists
    if not config:
        config = get_default_payroll_rules()
        config["config_id"] = "default"
        config["is_active"] = True
    
    return config


@router.get("/rules")
async def get_payroll_rules(request: Request):
    """Get all payroll rules (separate endpoint for clarity)"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Try new payroll_rules collection first
    rules = await db.payroll_rules.find_one({"is_active": True}, {"_id": 0})
    if rules:
        return rules
    
    # Fall back to payroll_config
    config = await db.payroll_config.find_one({"is_active": True}, {"_id": 0})
    if not config:
        return get_default_payroll_rules()
    return config


@router.put("/rules")
async def update_payroll_rules(data: dict, request: Request):
    """Update payroll rules (deduction percentages, etc.)"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get current rules
    current_rules = await db.payroll_rules.find_one({"is_active": True})
    
    update_data = {
        "epf_employee_percentage": float(data.get("epf_employee_percentage", 12)),
        "epf_employer_percentage": float(data.get("epf_employer_percentage", 12)),
        "epf_wage_ceiling": float(data.get("epf_wage_ceiling", 15000)),
        
        "esi_employee_percentage": float(data.get("esi_employee_percentage", 0.75)),
        "esi_employer_percentage": float(data.get("esi_employer_percentage", 3.25)),
        "esi_wage_ceiling": float(data.get("esi_wage_ceiling", 21000)),
        
        "sewa_percentage": float(data.get("sewa_percentage", 2)),
        "sewa_applicable": data.get("sewa_applicable", True),
        
        "lwf_employee": float(data.get("lwf_employee", 10)),
        "lwf_employer": float(data.get("lwf_employer", 20)),
        
        "default_working_days": int(data.get("default_working_days", 26)),
        "wfh_pay_percentage": float(data.get("wfh_pay_percentage", 50)),
        
        "late_deduction_enabled": data.get("late_deduction_enabled", True),
        "late_count_threshold": int(data.get("late_count_threshold", 3)),
        
        "salary_change_requires_approval": data.get("salary_change_requires_approval", True),
        
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "updated_by": user.get("user_id")
    }
    
    if current_rules:
        await db.payroll_rules.update_one(
            {"rules_id": current_rules.get("rules_id")},
            {"$set": update_data}
        )
    else:
        update_data["rules_id"] = f"rules_{uuid.uuid4().hex[:12]}"
        update_data["name"] = "Default Payroll Rules"
        update_data["is_active"] = True
        update_data["created_at"] = datetime.now(timezone.utc).isoformat()
        await db.payroll_rules.insert_one(update_data)
        update_data.pop("_id", None)
    
    return {"message": "Payroll rules updated successfully", "rules": update_data}


@router.post("/config")
async def update_payroll_config(data: dict, request: Request):
    """Update payroll configuration"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    data["config_id"] = f"cfg_{uuid.uuid4().hex[:12]}"
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    data["updated_by"] = user.get("user_id")
    data["is_active"] = True
    
    # Deactivate old config
    await db.payroll_config.update_many({}, {"$set": {"is_active": False}})
    
    await db.payroll_config.insert_one(data)
    data.pop('_id', None)
    return data


@router.put("/rules/{rule_section}")
async def update_payroll_rule_section(rule_section: str, data: dict, request: Request):
    """Update a specific section of payroll rules"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    valid_sections = [
        "attendance_rules", "leave_rules", "overtime_rules", 
        "pf_rules", "esi_rules", "pt_rules", "tds_rules",
        "salary_components", "bonus_rules", "loan_deduction_rules",
        "reimbursement_rules"
    ]
    
    if rule_section not in valid_sections:
        raise HTTPException(status_code=400, detail=f"Invalid rule section. Valid: {valid_sections}")
    
    # Get current config
    config = await db.payroll_config.find_one({"is_active": True}, {"_id": 0})
    if not config:
        config = get_default_payroll_rules()
    
    # Update the specific section
    config[rule_section] = data
    config["updated_at"] = datetime.now(timezone.utc).isoformat()
    config["updated_by"] = user.get("user_id")
    
    # Save updated config
    if config.get("config_id") and config["config_id"] != "default":
        await db.payroll_config.update_one(
            {"config_id": config["config_id"]},
            {"$set": {rule_section: data, "updated_at": config["updated_at"]}}
        )
    else:
        config["config_id"] = f"cfg_{uuid.uuid4().hex[:12]}"
        config["is_active"] = True
        await db.payroll_config.insert_one(config)
        config.pop('_id', None)
    
    return {"message": f"{rule_section} updated successfully", "data": data}


@router.get("/leave-type-rules")
async def get_leave_type_payroll_rules(request: Request):
    """Get payroll rules for each leave type"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get leave types
    leave_types = await db.leave_types.find({}, {"_id": 0}).to_list(50)
    
    # Get current payroll config
    config = await db.payroll_config.find_one({"is_active": True}, {"_id": 0})
    leave_rules = config.get("leave_rules", {}) if config else {}
    
    # Combine leave types with their payroll rules
    result = []
    for lt in leave_types:
        code = lt.get("code", "")
        rule = leave_rules.get(code, {"is_paid": False, "deduction_percent": 100})
        result.append({
            "leave_type_id": lt.get("leave_type_id"),
            "name": lt.get("name"),
            "code": code,
            "is_paid": rule.get("is_paid", False),
            "deduction_percent": rule.get("deduction_percent", 100),
        })
    
    return result


@router.put("/leave-type-rules")
async def update_leave_type_payroll_rules(data: List[dict], request: Request):
    """Update payroll rules for leave types"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get current config
    config = await db.payroll_config.find_one({"is_active": True}, {"_id": 0})
    if not config:
        config = get_default_payroll_rules()
    
    # Build leave rules dict
    leave_rules = {}
    for item in data:
        code = item.get("code")
        if code:
            leave_rules[code] = {
                "is_paid": item.get("is_paid", False),
                "deduction_percent": item.get("deduction_percent", 100),
            }
    
    config["leave_rules"] = leave_rules
    config["updated_at"] = datetime.now(timezone.utc).isoformat()
    config["updated_by"] = user.get("user_id")
    
    # Save
    if config.get("config_id") and config["config_id"] != "default":
        await db.payroll_config.update_one(
            {"config_id": config["config_id"]},
            {"$set": {"leave_rules": leave_rules, "updated_at": config["updated_at"]}}
        )
    else:
        config["config_id"] = f"cfg_{uuid.uuid4().hex[:12]}"
        config["is_active"] = True
        await db.payroll_config.insert_one(config)
        config.pop('_id', None)
    
    return {"message": "Leave type payroll rules updated", "leave_rules": leave_rules}


# ==================== LEAVE POLICY RULES ====================

@router.get("/leave-policy-rules")
async def get_leave_policy_rules(request: Request):
    """Get leave policy rules (quotas, carry forward, Sunday rules)"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "hr_executive", "finance"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    config = await db.payroll_config.find_one({"is_active": True}, {"_id": 0})
    
    # Default leave policy rules
    default_rules = {
        "financial_year_start": "04-01",
        "annual_quotas": {
            "CL": 6,
            "SL": 6,
            "EL": 12,
        },
        "carry_forward": {
            "CL": False,
            "SL": False,
            "EL": True,
            "max_el_accumulation": 30,
        },
        "sunday_leave_rules": {
            "enabled": True,
            "weekly_threshold": 2,
            "monthly_threshold": 6,
            "auto_apply": True,
        },
    }
    
    if config and config.get("leave_policy_rules"):
        return config["leave_policy_rules"]
    
    return default_rules


@router.put("/leave-policy-rules")
async def update_leave_policy_rules(data: dict, request: Request):
    """Update leave policy rules"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    config = await db.payroll_config.find_one({"is_active": True}, {"_id": 0})
    
    if not config:
        config = get_default_payroll_rules()
    
    # Update leave policy rules
    config["leave_policy_rules"] = {
        "financial_year_start": data.get("financial_year_start", "04-01"),
        "annual_quotas": data.get("annual_quotas", {"CL": 6, "SL": 6, "EL": 12}),
        "carry_forward": data.get("carry_forward", {"CL": False, "SL": False, "EL": True, "max_el_accumulation": 30}),
        "sunday_leave_rules": data.get("sunday_leave_rules", {
            "enabled": True,
            "weekly_threshold": 2,
            "monthly_threshold": 6,
            "auto_apply": True,
        }),
    }
    
    config["updated_at"] = datetime.now(timezone.utc).isoformat()
    config["updated_by"] = user.get("user_id")
    
    # Save
    if config.get("config_id") and config["config_id"] != "default":
        await db.payroll_config.update_one(
            {"config_id": config["config_id"]},
            {"$set": {
                "leave_policy_rules": config["leave_policy_rules"],
                "updated_at": config["updated_at"]
            }}
        )
    else:
        config["config_id"] = f"cfg_{uuid.uuid4().hex[:12]}"
        config["is_active"] = True
        await db.payroll_config.insert_one(config)
        config.pop('_id', None)
    
    return {"message": "Leave policy rules updated", "leave_policy_rules": config["leave_policy_rules"]}


# ==================== CUSTOM DEDUCTION RULES ====================

@router.get("/custom-rules")
async def get_custom_deduction_rules(request: Request):
    """Get custom deduction rules (e.g., late coming penalties)"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    rules = await db.custom_payroll_rules.find({"is_active": True}, {"_id": 0}).to_list(50)
    
    # Return default rules if none exist
    if not rules:
        rules = [
            {
                "rule_id": "default_late",
                "name": "Late Coming Penalty",
                "description": "Deduction for excessive late arrivals",
                "condition_type": "late_count",
                "condition_threshold": 3,
                "condition_operator": "greater_than",
                "action_type": "percentage_deduction",
                "action_value": 5,
                "is_active": True,
                "is_default": True
            },
            {
                "rule_id": "default_absent",
                "name": "Unauthorized Absence Penalty",
                "description": "Extra deduction for unapproved absences",
                "condition_type": "absent_without_leave",
                "condition_threshold": 2,
                "condition_operator": "greater_than",
                "action_type": "fixed_deduction",
                "action_value": 500,
                "is_active": True,
                "is_default": True
            }
        ]
    
    return rules


@router.post("/custom-rules")
async def create_custom_deduction_rule(data: dict, request: Request):
    """Create a new custom deduction rule"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    rule = {
        "rule_id": f"rule_{uuid.uuid4().hex[:12]}",
        "name": data.get("name", "Custom Rule"),
        "description": data.get("description", ""),
        "condition_type": data.get("condition_type"),  # late_count, absent_count, absent_without_leave, early_departure_count
        "condition_threshold": data.get("condition_threshold", 0),
        "condition_operator": data.get("condition_operator", "greater_than"),  # greater_than, equals, less_than
        "action_type": data.get("action_type"),  # percentage_deduction, fixed_deduction, half_day_deduction, full_day_deduction
        "action_value": data.get("action_value", 0),
        "apply_per_occurrence": data.get("apply_per_occurrence", False),  # If true, apply for each occurrence above threshold
        "is_active": True,
        "is_default": False,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user.get("user_id")
    }
    
    await db.custom_payroll_rules.insert_one(rule)
    rule.pop('_id', None)
    return rule


@router.put("/custom-rules/{rule_id}")
async def update_custom_deduction_rule(rule_id: str, data: dict, request: Request):
    """Update a custom deduction rule"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Don't allow editing default rules
    existing = await db.custom_payroll_rules.find_one({"rule_id": rule_id}, {"_id": 0})
    if existing and existing.get("is_default"):
        raise HTTPException(status_code=400, detail="Cannot edit default rules")
    
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    data["updated_by"] = user.get("user_id")
    
    await db.custom_payroll_rules.update_one(
        {"rule_id": rule_id},
        {"$set": data}
    )
    
    return await db.custom_payroll_rules.find_one({"rule_id": rule_id}, {"_id": 0})


@router.delete("/custom-rules/{rule_id}")
async def delete_custom_deduction_rule(rule_id: str, request: Request):
    """Delete a custom deduction rule"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Don't allow deleting default rules
    existing = await db.custom_payroll_rules.find_one({"rule_id": rule_id}, {"_id": 0})
    if existing and existing.get("is_default"):
        raise HTTPException(status_code=400, detail="Cannot delete default rules")
    
    await db.custom_payroll_rules.update_one(
        {"rule_id": rule_id},
        {"$set": {"is_active": False}}
    )
    
    return {"message": "Rule deleted"}


@router.put("/custom-rules/{rule_id}/toggle")
async def toggle_custom_rule(rule_id: str, request: Request):
    """Toggle a custom rule on/off"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    rule = await db.custom_payroll_rules.find_one({"rule_id": rule_id}, {"_id": 0})
    if not rule:
        raise HTTPException(status_code=404, detail="Rule not found")
    
    new_status = not rule.get("is_active", True)
    await db.custom_payroll_rules.update_one(
        {"rule_id": rule_id},
        {"$set": {"is_active": new_status}}
    )
    
    return {"message": f"Rule {'enabled' if new_status else 'disabled'}", "is_active": new_status}



# ==================== SEWA ADVANCE MANAGEMENT ====================

@router.get("/sewa-advances")
async def list_sewa_advances(
    request: Request,
    status: Optional[str] = None,
    employee_id: Optional[str] = None
):
    """List all SEWA advances"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    query = {}
    if status:
        query["status"] = status
    if employee_id:
        query["employee_id"] = employee_id
    
    advances = await db.sewa_advances.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    
    # Enrich with employee details
    for adv in advances:
        employee = await db.employees.find_one(
            {"employee_id": adv.get("employee_id")}, {"_id": 0}
        )
        if employee:
            adv["employee_name"] = f"{employee.get('first_name', '')} {employee.get('last_name', '')}".strip()
            adv["emp_code"] = employee.get("emp_code", "")
    
    return advances


@router.post("/sewa-advances")
async def create_sewa_advance(data: dict, request: Request):
    """
    Create a SEWA advance for an employee
    
    Required fields:
    - employee_id: Employee to assign advance to
    - total_amount: Total amount to be recovered
    - monthly_amount: Amount to deduct each month
    - duration_months: Expected duration (for reference)
    """
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    employee_id = data.get("employee_id")
    if not employee_id:
        raise HTTPException(status_code=400, detail="Employee ID is required")
    
    # Check if employee exists
    employee = await db.employees.find_one({"employee_id": employee_id})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Check for existing active advance
    existing = await db.sewa_advances.find_one({
        "employee_id": employee_id,
        "is_active": True
    })
    if existing:
        raise HTTPException(status_code=400, detail="Employee already has an active SEWA advance")
    
    total_amount = float(data.get("total_amount", 0))
    monthly_amount = float(data.get("monthly_amount", 0))
    
    if total_amount <= 0 or monthly_amount <= 0:
        raise HTTPException(status_code=400, detail="Total amount and monthly amount must be greater than 0")
    
    advance = {
        "advance_id": f"sewa_{uuid.uuid4().hex[:12]}",
        "employee_id": employee_id,
        "emp_code": employee.get("emp_code", ""),
        "employee_name": f"{employee.get('first_name', '')} {employee.get('last_name', '')}".strip(),
        "total_amount": total_amount,
        "monthly_amount": monthly_amount,
        "duration_months": int(data.get("duration_months", 0)) or int(total_amount / monthly_amount),
        "total_paid": 0,
        "remaining_amount": total_amount,
        "start_month": int(data.get("start_month", datetime.now().month)),
        "start_year": int(data.get("start_year", datetime.now().year)),
        "reason": data.get("reason", "SEWA Advance"),
        "status": "active",
        "is_active": True,
        "created_by": user.get("user_id"),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.sewa_advances.insert_one(advance)
    advance.pop("_id", None)
    return advance


@router.put("/sewa-advances/{advance_id}")
async def update_sewa_advance(advance_id: str, data: dict, request: Request):
    """Update a SEWA advance"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    advance = await db.sewa_advances.find_one({"advance_id": advance_id})
    if not advance:
        raise HTTPException(status_code=404, detail="SEWA advance not found")
    
    update_data = {}
    
    if "monthly_amount" in data:
        update_data["monthly_amount"] = float(data["monthly_amount"])
    if "total_amount" in data:
        update_data["total_amount"] = float(data["total_amount"])
        update_data["remaining_amount"] = float(data["total_amount"]) - float(advance.get("total_paid", 0))
    if "reason" in data:
        update_data["reason"] = data["reason"]
    
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    update_data["updated_by"] = user.get("user_id")
    
    await db.sewa_advances.update_one(
        {"advance_id": advance_id},
        {"$set": update_data}
    )
    
    return await db.sewa_advances.find_one({"advance_id": advance_id}, {"_id": 0})


@router.delete("/sewa-advances/{advance_id}")
async def delete_sewa_advance(advance_id: str, request: Request):
    """Deactivate a SEWA advance"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    await db.sewa_advances.update_one(
        {"advance_id": advance_id},
        {"$set": {
            "is_active": False,
            "status": "cancelled",
            "cancelled_at": datetime.now(timezone.utc).isoformat(),
            "cancelled_by": user.get("user_id")
        }}
    )
    
    return {"message": "SEWA advance cancelled"}


@router.post("/sewa-advances/{advance_id}/complete")
async def complete_sewa_advance(advance_id: str, request: Request):
    """Mark a SEWA advance as completed"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    await db.sewa_advances.update_one(
        {"advance_id": advance_id},
        {"$set": {
            "is_active": False,
            "status": "completed",
            "completed_at": datetime.now(timezone.utc).isoformat(),
            "completed_by": user.get("user_id")
        }}
    )
    
    return {"message": "SEWA advance marked as completed"}


# ==================== ONE-TIME DEDUCTIONS ====================

@router.get("/one-time-deductions")
async def list_one_time_deductions(
    request: Request,
    month: Optional[int] = None,
    year: Optional[int] = None,
    employee_id: Optional[str] = None
):
    """List one-time deductions"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    query = {"is_active": True}
    if month:
        query["month"] = month
    if year:
        query["year"] = year
    if employee_id:
        query["employee_id"] = employee_id
    
    deductions = await db.one_time_deductions.find(query, {"_id": 0}).to_list(500)
    
    # Enrich with employee details
    for ded in deductions:
        employee = await db.employees.find_one(
            {"employee_id": ded.get("employee_id")}, {"_id": 0}
        )
        if employee:
            ded["employee_name"] = f"{employee.get('first_name', '')} {employee.get('last_name', '')}".strip()
            ded["emp_code"] = employee.get("emp_code", "")
    
    return deductions


@router.post("/one-time-deductions")
async def create_one_time_deduction(data: dict, request: Request):
    """Create a one-time deduction for an employee (e.g., loan EMI, advance recovery)"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    employee_id = data.get("employee_id")
    if not employee_id:
        raise HTTPException(status_code=400, detail="Employee ID is required")
    
    # Check if employee exists
    employee = await db.employees.find_one({"employee_id": employee_id})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    deduction = {
        "deduction_id": f"otd_{uuid.uuid4().hex[:12]}",
        "employee_id": employee_id,
        "emp_code": employee.get("emp_code", ""),
        "employee_name": f"{employee.get('first_name', '')} {employee.get('last_name', '')}".strip(),
        "amount": float(data.get("amount", 0)),
        "month": int(data.get("month", datetime.now().month)),
        "year": int(data.get("year", datetime.now().year)),
        "reason": data.get("reason", "One-time deduction"),
        "category": data.get("category", "other"),  # loan_emi, advance_recovery, penalty, other
        "is_active": True,
        "created_by": user.get("user_id"),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.one_time_deductions.insert_one(deduction)
    deduction.pop("_id", None)
    return deduction


@router.delete("/one-time-deductions/{deduction_id}")
async def delete_one_time_deduction(deduction_id: str, request: Request):
    """Delete a one-time deduction"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    await db.one_time_deductions.update_one(
        {"deduction_id": deduction_id},
        {"$set": {"is_active": False}}
    )
    
    return {"message": "Deduction removed"}


# ==================== PAYSLIP EDITING (Before Lock) ====================

@router.put("/payslips/{payslip_id}")
async def update_payslip(payslip_id: str, data: dict, request: Request):
    """
    Update a payslip before payroll is locked
    
    HR can edit:
    - Attendance inputs (office_days, wfh_days, leave_days, etc.) and recalculate
    - Or directly edit individual values (gross_salary, deductions, net_salary)
    """
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    payslip = await db.payslips.find_one({"payslip_id": payslip_id})
    if not payslip:
        raise HTTPException(status_code=404, detail="Payslip not found")
    
    # Check if payroll is locked
    payroll = await db.payroll_runs.find_one({"payroll_id": payslip.get("payroll_id")})
    if payroll and payroll.get("status") == "locked":
        raise HTTPException(status_code=400, detail="Cannot edit payslip - payroll is locked")
    
    update_data = {}
    recalculate = data.get("recalculate", False)
    
    if recalculate:
        # Update attendance and recalculate
        attendance = data.get("attendance", {})
        if attendance:
            # Get employee salary structure
            employee_id = payslip.get("employee_id")
            emp_salary = await db.employee_salaries.find_one(
                {"employee_id": employee_id, "is_active": True}, {"_id": 0}
            )
            
            if not emp_salary:
                raise HTTPException(status_code=400, detail="No salary structure found for employee")
            
            # Get payroll config
            payroll_rules = await db.payroll_rules.find_one({"is_active": True}, {"_id": 0})
            config = await db.payroll_config.find_one({"is_active": True}, {"_id": 0})
            
            payroll_config = {
                "epf_employee_percentage": float((payroll_rules or {}).get("epf_employee_percentage", 12.0)),
                "epf_wage_ceiling": float((config or {}).get("pf_wage_ceiling", 15000)),
                "esi_employee_percentage": float((payroll_rules or {}).get("esi_employee_percentage", 0.75)),
                "esi_wage_ceiling": float((config or {}).get("esi_wage_ceiling", 21000)),
                "sewa_percentage": float((payroll_rules or {}).get("sewa_percentage", 2.0)),
                "wfh_pay_percentage": float((payroll_rules or {}).get("wfh_pay_percentage", 50.0)),
                "late_deduction_enabled": (payroll_rules or {}).get("late_deduction_enabled", True),
                "late_count_threshold": int((payroll_rules or {}).get("late_count_threshold", 2)),
            }
            
            # Get SEWA advance
            sewa_advance = await db.sewa_advances.find_one(
                {"employee_id": employee_id, "is_active": True}, {"_id": 0}
            )
            
            # Recalculate
            month = payslip.get("month")
            year = payslip.get("year")
            
            from routes.payroll_v2 import process_employee_salary
            
            new_payslip_data = process_employee_salary(
                employee_salary=emp_salary,
                attendance_data=attendance,
                payroll_config=payroll_config,
                month=month,
                year=year,
                sewa_advance_info=sewa_advance
            )
            
            # Update all calculated fields
            update_data = {
                "attendance": new_payslip_data["attendance"],
                "earnings": new_payslip_data["earnings"],
                "deductions": new_payslip_data["deductions"],
                "gross_salary": new_payslip_data["gross_salary"],
                "total_deductions": new_payslip_data["total_deductions"],
                "net_salary": new_payslip_data["net_payable"],
                "paid_days": new_payslip_data["attendance"]["total_earned_days"],
                "basic": new_payslip_data["earnings"]["basic_da_earned"],
                "hra": new_payslip_data["earnings"]["hra_earned"],
                "pf_employee": new_payslip_data["deductions"]["epf"],
                "esi_employee": new_payslip_data["deductions"]["esi"],
                "sewa": new_payslip_data["deductions"]["sewa"],
                "sewa_advance": new_payslip_data["deductions"]["sewa_advance"],
            }
    else:
        # Direct edit mode - update individual fields
        editable_fields = [
            "gross_salary", "total_deductions", "net_salary",
            "basic", "hra", "special_allowance",
            "pf_employee", "esi_employee", "sewa", "sewa_advance", "other_deduction",
            "office_days", "wfh_days", "leave_days", "late_count"
        ]
        
        for field in editable_fields:
            if field in data:
                update_data[field] = data[field]
        
        # Also allow updating nested attendance/earnings/deductions
        if "attendance" in data:
            update_data["attendance"] = data["attendance"]
        if "earnings" in data:
            update_data["earnings"] = data["earnings"]
        if "deductions" in data:
            update_data["deductions"] = data["deductions"]
    
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    update_data["updated_by"] = user.get("user_id")
    update_data["is_manually_edited"] = True
    
    await db.payslips.update_one(
        {"payslip_id": payslip_id},
        {"$set": update_data}
    )
    
    updated_payslip = await db.payslips.find_one({"payslip_id": payslip_id}, {"_id": 0})
    return updated_payslip


# ==================== PAYROLL EXPORT (Template Format) ====================

@router.get("/runs/{payroll_id}/export")
async def export_payroll_to_excel(payroll_id: str, request: Request):
    """
    Export payroll to Excel in the same format as the salary structure template
    """
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    payroll = await db.payroll_runs.find_one({"payroll_id": payroll_id}, {"_id": 0})
    if not payroll:
        raise HTTPException(status_code=404, detail="Payroll run not found")
    
    if payroll.get("status") not in ["processed", "locked"]:
        raise HTTPException(status_code=400, detail="Payroll must be processed before export")
    
    month, year = payroll["month"], payroll["year"]
    
    # Get all payslips
    payslips = await db.payslips.find(
        {"payroll_id": payroll_id}, {"_id": 0}
    ).to_list(1000)
    
    if not payslips:
        raise HTTPException(status_code=404, detail="No payslips found for this payroll")
    
    # Generate export data in template format
    export_data = generate_payroll_export_data(payslips, month, year)
    
    # Create Excel file
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Salary Structure"
        
        # Define headers (matching template)
        headers = [
            "Emp Code", "Name of Employees", "BASIC", "DA", "HRA", "Conveyance",
            "GRADE PAY", "OTHER ALLOW", "Med./Spl. Allow", "Total Salary (FIXED)",
            "Work from office", "Sunday + Holiday Leave Days", "Leave Days",
            "Work from Home @50%", "Late Deduction", "Basic+DA (Earned)",
            "HRA (Earned)", "Conveyance (Earned)", "GRADE PAY (Earned)",
            "OTHER ALLOW (Earned)", "Med./Spl. Allow (Earned)", "Total Earned Days",
            "Total Salary Earned", "EPF Employees", "ESI Employees", "SEWA",
            "Sewa Advance", "Other Deduction", "Total Deduction", "NET PAYABLE"
        ]
        
        # Write headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border
        
        # Write data rows
        for row_num, row_data in enumerate(export_data, 2):
            for col, header in enumerate(headers, 1):
                value = row_data.get(header, 0)
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = thin_border
                
                # Format numbers
                if col >= 3:  # Numeric columns
                    cell.number_format = '#,##0.00'
        
        # Add totals row
        total_row = len(export_data) + 2
        ws.cell(row=total_row, column=1, value="TOTALS")
        ws.cell(row=total_row, column=2, value=f"{len(export_data)} Employees")
        
        # Sum columns for totals
        sum_cols = [3, 4, 5, 6, 7, 8, 9, 10, 16, 17, 18, 19, 20, 21, 23, 24, 25, 26, 27, 28, 29, 30]
        for col in sum_cols:
            total = sum(row.get(headers[col-1], 0) or 0 for row in export_data)
            cell = ws.cell(row=total_row, column=col, value=total)
            cell.font = Font(bold=True)
            cell.number_format = '#,##0.00'
        
        # Auto-size columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        filename = f"Payroll_{month_names[month-1]}_{year}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl library not available for Excel export")


# ==================== SALARY BULK IMPORT (Template Format) ====================

@router.post("/import-salaries")
async def import_salaries_from_template(file: UploadFile = File(...), request: Request = None):
    """
    Import salary structures from the salary structure Excel template
    
    Expected columns: Emp Code, BASIC, DA, HRA, Conveyance, GRADE PAY, OTHER ALLOW, Med./Spl. Allow
    """
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Please upload an Excel file (.xlsx or .xls)")
    
    try:
        import openpyxl
        
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active
        
        # Get headers from first row
        headers = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            headers.append(str(val).strip() if val else "")
        
        # Map column indices
        col_map = {}
        expected_cols = {
            "emp code": "emp_code",
            "name of employees": "name",
            "basic": "basic",
            "da": "da",
            "hra": "hra",
            "conveyance": "conveyance",
            "grade pay": "grade_pay",
            "other allow": "other_allowance",
            "med./spl. allow": "medical_allowance",
            "total salary (fixed)": "total_fixed"
        }
        
        for idx, header in enumerate(headers):
            header_lower = header.lower().strip()
            for key, mapped in expected_cols.items():
                if key in header_lower:
                    col_map[mapped] = idx
                    break
        
        if "emp_code" not in col_map:
            raise HTTPException(status_code=400, detail="Emp Code column not found in file")
        
        imported = 0
        errors = []
        
        for row_num in range(2, ws.max_row + 1):
            try:
                emp_code = ws.cell(row=row_num, column=col_map["emp_code"] + 1).value
                if not emp_code:
                    continue
                
                emp_code = str(emp_code).strip()
                
                # Find employee by emp_code
                employee = await db.employees.find_one({"emp_code": emp_code})
                if not employee:
                    errors.append(f"Row {row_num}: Employee {emp_code} not found")
                    continue
                
                # Get values
                def get_val(key, default=0):
                    if key not in col_map:
                        return default
                    val = ws.cell(row=row_num, column=col_map[key] + 1).value
                    try:
                        return float(val) if val else default
                    except (ValueError, TypeError):
                        return default
                
                basic = get_val("basic", 0)
                da = get_val("da", 0)
                hra = get_val("hra", 0)
                conveyance = get_val("conveyance", 0)
                grade_pay = get_val("grade_pay", 0)
                other_allowance = get_val("other_allowance", 0)
                medical_allowance = get_val("medical_allowance", 0)
                
                total_fixed = basic + da + hra + conveyance + grade_pay + other_allowance + medical_allowance
                
                # Create/update salary structure
                salary_data = {
                    "salary_id": f"sal_{uuid.uuid4().hex[:12]}",
                    "employee_id": employee.get("employee_id"),
                    "emp_code": emp_code,
                    "employee_name": get_val("name", "") or f"{employee.get('first_name', '')} {employee.get('last_name', '')}".strip(),
                    "fixed_components": {
                        "basic": basic,
                        "da": da,
                        "hra": hra,
                        "conveyance": conveyance,
                        "grade_pay": grade_pay,
                        "other_allowance": other_allowance,
                        "medical_allowance": medical_allowance
                    },
                    "total_fixed": total_fixed,
                    "deduction_config": {
                        "epf_applicable": True,
                        "esi_applicable": total_fixed <= 21000,
                        "sewa_applicable": True,
                        "sewa_percentage": 2.0
                    },
                    "fixed_deductions": {
                        "sewa_advance": 0,
                        "other_deduction": 0
                    },
                    "is_active": True,
                    "imported_at": datetime.now(timezone.utc).isoformat(),
                    "imported_by": user.get("user_id")
                }
                
                # Deactivate old salary
                await db.employee_salaries.update_many(
                    {"employee_id": employee.get("employee_id"), "is_active": True},
                    {"$set": {"is_active": False}}
                )
                
                # Insert new salary
                await db.employee_salaries.insert_one(salary_data)
                imported += 1
                
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
        
        return {
            "message": f"Imported {imported} salary structures",
            "imported": imported,
            "errors": errors[:20] if errors else []  # Limit errors shown
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")


# ==================== PAYSLIP PDF DOWNLOAD ====================

@router.get("/payslip/{payslip_id}/pdf")
async def download_payslip_pdf(payslip_id: str, request: Request):
    """Download payslip as PDF - employees can download their own, HR can download any"""
    user = await get_current_user(request)
    
    # Find the payslip
    payslip = await db.payslips.find_one({"payslip_id": payslip_id}, {"_id": 0})
    if not payslip:
        raise HTTPException(status_code=404, detail="Payslip not found")
    
    # Check access - employees can only download their own
    if user.get("role") not in ["super_admin", "hr_admin", "finance", "hr_executive"]:
        if payslip.get("employee_id") != user.get("employee_id"):
            raise HTTPException(status_code=403, detail="Not authorized to download this payslip")
    
    # Generate PDF
    pdf_buffer = await generate_payslip_pdf(payslip)
    
    filename = f"payslip_{payslip.get('emp_code', payslip.get('employee_id'))}_{payslip.get('month', '')}_{payslip.get('year', '')}.pdf"
    
    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


@router.get("/my-payslip/{month}/{year}/pdf")
async def download_my_payslip_pdf(month: int, year: int, request: Request):
    """Download current user's payslip for a specific month/year as PDF"""
    user = await get_current_user(request)
    employee_id = user.get("employee_id")
    
    if not employee_id:
        raise HTTPException(status_code=400, detail="No employee profile linked")
    
    # Find the payslip
    payslip = await db.payslips.find_one({
        "employee_id": employee_id,
        "month": month,
        "year": year
    }, {"_id": 0})
    
    if not payslip:
        raise HTTPException(status_code=404, detail="Payslip not found for this period")
    
    # Generate PDF
    pdf_buffer = await generate_payslip_pdf(payslip)
    
    filename = f"payslip_{payslip.get('emp_code', employee_id)}_{month}_{year}.pdf"
    
    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


async def generate_payslip_pdf(payslip: dict) -> io.BytesIO:
    """Generate a professional payslip PDF"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=1*cm, leftMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, alignment=TA_CENTER, spaceAfter=6)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=10, alignment=TA_CENTER, spaceAfter=12)
    
    elements = []
    
    # Company Header
    elements.append(Paragraph("SHARDA DIESELS", title_style))
    elements.append(Paragraph("PAYSLIP", subtitle_style))
    
    # Month/Year
    month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June', 
                   'July', 'August', 'September', 'October', 'November', 'December']
    month_name = month_names[payslip.get('month', 1)] if 1 <= payslip.get('month', 1) <= 12 else ''
    elements.append(Paragraph(f"For the month of {month_name} {payslip.get('year', '')}", subtitle_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Employee Information Table
    employee_info = [
        ['Employee Code:', payslip.get('emp_code', '-'), 'Employee Name:', payslip.get('employee_name', '-')],
        ['Department:', payslip.get('department', '-'), 'Designation:', payslip.get('designation', '-')],
        ['Working Days:', str(payslip.get('working_days', 0)), 'Present Days:', str(payslip.get('present_days', payslip.get('paid_days', 0)))],
        ['LWP Days:', str(payslip.get('lwp_days', 0)), 'Late Count:', str(payslip.get('attendance', {}).get('late_count', 0) if isinstance(payslip.get('attendance'), dict) else 0)],
    ]
    
    emp_table = Table(employee_info, colWidths=[2.5*cm, 4*cm, 2.5*cm, 4*cm])
    emp_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.Color(0.9, 0.9, 0.95)),
        ('BACKGROUND', (2, 0), (2, -1), colors.Color(0.9, 0.9, 0.95)),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(emp_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Format currency
    def fmt_curr(val):
        try:
            return f"₹{float(val):,.2f}"
        except (ValueError, TypeError):
            return "₹0.00"
    
    # Get fixed components
    fc = payslip.get('fixed_components', {})
    earn = payslip.get('earnings', {})
    ded = payslip.get('deductions', {})
    
    # Earnings and Deductions side by side
    earnings_data = [['EARNINGS', 'Fixed', 'Earned']]
    earnings_data.append(['Basic + DA', fmt_curr(fc.get('basic', 0) + fc.get('da', 0)), fmt_curr(earn.get('basic_da_earned', payslip.get('basic', 0)))])
    earnings_data.append(['HRA', fmt_curr(fc.get('hra', 0)), fmt_curr(earn.get('hra_earned', payslip.get('hra', 0)))])
    earnings_data.append(['Conveyance', fmt_curr(fc.get('conveyance', 0)), fmt_curr(earn.get('conveyance_earned', 0))])
    earnings_data.append(['Grade Pay', fmt_curr(fc.get('grade_pay', 0)), fmt_curr(earn.get('grade_pay_earned', 0))])
    earnings_data.append(['Other Allowance', fmt_curr(fc.get('other_allowance', 0)), fmt_curr(earn.get('other_allowance_earned', 0))])
    earnings_data.append(['Medical Allowance', fmt_curr(fc.get('medical_allowance', 0)), fmt_curr(earn.get('medical_allowance_earned', payslip.get('special_allowance', 0)))])
    earnings_data.append(['Total Fixed', fmt_curr(fc.get('total_fixed', 0)), ''])
    earnings_data.append(['Late Deduction', '', fmt_curr(-1 * (earn.get('late_deduction', 0) or 0))])
    earnings_data.append(['GROSS SALARY', '', fmt_curr(payslip.get('gross_salary', 0))])
    
    deductions_data = [['DEDUCTIONS', 'Amount']]
    deductions_data.append(['EPF (Employee)', fmt_curr(ded.get('epf', payslip.get('pf_employee', 0)))])
    deductions_data.append(['ESI (Employee)', fmt_curr(ded.get('esi', payslip.get('esi_employee', 0)))])
    deductions_data.append(['SEWA', fmt_curr(ded.get('sewa', payslip.get('sewa', 0)))])
    deductions_data.append(['SEWA Advance', fmt_curr(ded.get('sewa_advance', payslip.get('sewa_advance', 0)))])
    deductions_data.append(['Other Deduction', fmt_curr(ded.get('other_deduction', payslip.get('other_deduction', 0)))])
    deductions_data.append(['', ''])
    deductions_data.append(['', ''])
    deductions_data.append(['', ''])
    deductions_data.append(['TOTAL DEDUCTIONS', fmt_curr(payslip.get('total_deductions', 0))])
    
    earn_table = Table(earnings_data, colWidths=[4.5*cm, 2.8*cm, 2.8*cm])
    earn_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.2, 0.4, 0.6)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.Color(0.85, 0.95, 0.85)),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('PADDING', (0, 0), (-1, -1), 5),
    ]))
    
    ded_table = Table(deductions_data, colWidths=[4.5*cm, 2.8*cm])
    ded_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.6, 0.2, 0.2)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.Color(0.95, 0.85, 0.85)),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('PADDING', (0, 0), (-1, -1), 5),
    ]))
    
    # Side by side tables
    combined_table = Table([[earn_table, ded_table]], colWidths=[10.2*cm, 7.5*cm])
    combined_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(combined_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Net Payable
    net_data = [
        ['NET PAYABLE', fmt_curr(payslip.get('net_salary', 0))]
    ]
    net_table = Table(net_data, colWidths=[10*cm, 7*cm])
    net_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.Color(0.1, 0.3, 0.5)),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 14),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ('PADDING', (0, 0), (-1, -1), 12),
    ]))
    elements.append(net_table)
    elements.append(Spacer(1, 0.5*inch))
    
    # Footer
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, alignment=TA_CENTER, textColor=colors.grey)
    elements.append(Paragraph("This is a computer-generated payslip and does not require a signature.", footer_style))
    elements.append(Paragraph(f"Generated on {datetime.now().strftime('%d-%b-%Y %H:%M')}", footer_style))
    
    doc.build(elements)
    buffer.seek(0)
    return buffer

