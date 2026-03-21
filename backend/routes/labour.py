"""Labour & Contract Labour Management API Routes"""
from fastapi import APIRouter, HTTPException, Request
from typing import List, Optional
from datetime import datetime, timezone
import uuid
from motor.motor_asyncio import AsyncIOMotorClient
import os

router = APIRouter(prefix="/labour", tags=["Labour Management"])

mongo_url = os.environ.get('MONGO_URL', 'mongodb://localhost:27017')
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ.get('DB_NAME', 'test_database')]


async def get_current_user(request: Request) -> dict:
    from server import get_current_user as auth_get_user
    return await auth_get_user(request)


# ==================== CONTRACTORS ====================

@router.get("/contractors")
async def list_contractors(
    request: Request,
    status: Optional[str] = None,
    department_id: Optional[str] = None
):
    """List contractors/agencies"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "hr_executive"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    query = {"is_active": True}
    if status:
        query["status"] = status
    if department_id:
        query["department_id"] = department_id
    
    contractors = await db.contractors.find(query, {"_id": 0}).to_list(100)
    return contractors


@router.post("/contractors")
async def create_contractor(data: dict, request: Request):
    """Register contractor/agency"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    contractor = {
        "contractor_id": f"CONT-{uuid.uuid4().hex[:8].upper()}",
        "name": data.get("name"),
        "company_name": data.get("company_name"),
        "contact_person": data.get("contact_person"),
        "email": data.get("email"),
        "phone": data.get("phone"),
        "address": data.get("address"),
        "gst_number": data.get("gst_number"),
        "pan_number": data.get("pan_number"),
        "department_id": data.get("department_id"),
        "contract_start": data.get("contract_start"),
        "contract_end": data.get("contract_end"),
        "contract_value": data.get("contract_value"),
        "status": "active",
        "is_active": True,
        "created_by": user["user_id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.contractors.insert_one(contractor)
    contractor.pop('_id', None)
    return contractor


@router.get("/contractors/{contractor_id}")
async def get_contractor(contractor_id: str, request: Request):
    """Get contractor details"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "hr_executive"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    contractor = await db.contractors.find_one({"contractor_id": contractor_id}, {"_id": 0})
    if not contractor:
        raise HTTPException(status_code=404, detail="Contractor not found")
    
    # Get associated workers
    workers = await db.contract_workers.find({"contractor_id": contractor_id}, {"_id": 0}).to_list(200)
    contractor["workers"] = workers
    contractor["worker_count"] = len(workers)
    
    return contractor


@router.put("/contractors/{contractor_id}")
async def update_contractor(contractor_id: str, data: dict, request: Request):
    """Update contractor"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.contractors.update_one({"contractor_id": contractor_id}, {"$set": data})
    return await db.contractors.find_one({"contractor_id": contractor_id}, {"_id": 0})


# ==================== CONTRACT WORKERS ====================

@router.get("/workers")
async def list_contract_workers(
    request: Request,
    contractor_id: Optional[str] = None,
    status: Optional[str] = None,
    department_id: Optional[str] = None
):
    """List contract workers"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "hr_executive", "manager"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    query = {"is_active": True}
    if contractor_id:
        query["contractor_id"] = contractor_id
    if status:
        query["status"] = status
    if department_id:
        query["department_id"] = department_id
    
    workers = await db.contract_workers.find(query, {"_id": 0}).to_list(500)
    return workers


@router.post("/workers")
async def create_contract_worker(data: dict, request: Request):
    """Add contract worker"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "hr_executive"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    worker = {
        "worker_id": f"CW-{uuid.uuid4().hex[:8].upper()}",
        "contractor_id": data.get("contractor_id"),
        "first_name": data.get("first_name"),
        "last_name": data.get("last_name"),
        "phone": data.get("phone"),
        "email": data.get("email"),
        "aadhaar_number": data.get("aadhaar_number"),
        "department_id": data.get("department_id"),
        "location_id": data.get("location_id"),
        "skill_category": data.get("skill_category"),
        "daily_rate": data.get("daily_rate"),
        "start_date": data.get("start_date"),
        "end_date": data.get("end_date"),
        "reporting_manager": data.get("reporting_manager"),
        "status": "active",
        "is_active": True,
        "created_by": user["user_id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.contract_workers.insert_one(worker)
    worker.pop('_id', None)
    return worker


# ==================== TEMPLATE DOWNLOAD & BULK UPLOAD ====================
# NOTE: These must be defined BEFORE /workers/{worker_id} to avoid route conflicts

TEMPLATE_COLUMNS = ["Sl No", "Employee Code", "Name", "Designation", "Date of Joining", "Ph.no", "Adhar no", "Contractor name"]


@router.get("/workers/template/download")
async def download_worker_template(request: Request):
    """Download Excel template for contract workers bulk upload"""
    from fastapi.responses import Response
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import io

    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "hr_executive"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    wb = Workbook()
    ws = wb.active
    ws.title = "Contract Workers List"

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col_idx, header in enumerate(TEMPLATE_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    col_widths = [8, 16, 25, 20, 16, 15, 16, 25]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    sample = [1, "CW001", "Sample Name", "Helper", "2026-01-15", "9876543210", "123456789012", "Contractor ABC"]
    for col_idx, val in enumerate(sample, 1):
        cell = ws.cell(row=2, column=col_idx, value=val)
        cell.border = thin_border
        cell.font = Font(color="999999", italic=True)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=contract_worker_template.xlsx"}
    )


@router.get("/workers/export")
async def export_workers(request: Request):
    """Export all contract workers as Excel file matching the template format"""
    from fastapi.responses import Response
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import io

    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "hr_executive"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    workers = await db.contract_workers.find({"is_active": True}, {"_id": 0}).to_list(1000)
    contractors_list = await db.contractors.find({"is_active": True}, {"_id": 0}).to_list(100)
    contractor_map = {c["contractor_id"]: c.get("name") or c.get("company_name", "") for c in contractors_list}

    wb = Workbook()
    ws = wb.active
    ws.title = "Contract Workers List"

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col_idx, header in enumerate(TEMPLATE_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    col_widths = [8, 16, 25, 20, 16, 15, 16, 25]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    for idx, worker in enumerate(workers, 1):
        row = [
            idx,
            worker.get("employee_code") or worker.get("worker_id", ""),
            worker.get("name", ""),
            worker.get("designation", ""),
            worker.get("joining_date", ""),
            worker.get("phone", ""),
            worker.get("aadhar_number", ""),
            contractor_map.get(worker.get("contractor_id"), "")
        ]
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=idx + 1, column=col_idx, value=val)
            cell.border = thin_border

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=contract_workers_export.xlsx"}
    )


@router.post("/workers/bulk-upload")
async def bulk_upload_workers(request: Request):
    """Bulk upload contract workers from Excel file matching the template format"""
    from openpyxl import load_workbook
    import io

    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "hr_executive"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    form = await request.form()
    file = form.get("file")
    if not file:
        raise HTTPException(status_code=400, detail="No file uploaded")

    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active

    headers = [str(cell.value or "").strip() for cell in ws[1]]

    col_map = {}
    for idx, h in enumerate(headers):
        hl = h.lower().strip()
        if hl in ("sl no", "sl.no", "s.no", "sno", "sl no."):
            col_map["sl_no"] = idx
        elif hl in ("employee code", "emp code", "emp_code", "employee_code"):
            col_map["employee_code"] = idx
        elif hl == "name":
            col_map["name"] = idx
        elif hl == "designation":
            col_map["designation"] = idx
        elif hl in ("date of joining", "doj", "joining date", "date_of_joining"):
            col_map["joining_date"] = idx
        elif hl in ("ph.no", "ph no", "phone", "phone no", "phone number", "mobile"):
            col_map["phone"] = idx
        elif hl in ("adhar no", "aadhar no", "aadhaar no", "aadhar number", "aadhaar number", "adhar no."):
            col_map["aadhar_number"] = idx
        elif hl in ("contractor name", "contractor", "agency", "contractor_name"):
            col_map["contractor_name"] = idx

    if "name" not in col_map:
        raise HTTPException(status_code=400, detail="Required column 'Name' not found in the uploaded file")

    contractors_list = await db.contractors.find({"is_active": True}, {"_id": 0}).to_list(100)
    contractor_lookup = {}
    for c in contractors_list:
        cname = (c.get("name") or c.get("company_name") or "").strip().lower()
        if cname:
            contractor_lookup[cname] = c["contractor_id"]

    created = 0
    skipped = 0
    errors = []
    now = datetime.now(timezone.utc).isoformat()

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(v is None for v in row):
            continue

        def get_val(key):
            idx = col_map.get(key)
            if idx is not None and idx < len(row):
                v = row[idx]
                return str(v).strip() if v is not None else ""
            return ""

        name = get_val("name")
        if not name:
            skipped += 1
            continue

        contractor_name_raw = get_val("contractor_name")
        contractor_id = None
        if contractor_name_raw:
            contractor_id = contractor_lookup.get(contractor_name_raw.lower())
            if not contractor_id:
                new_id = f"CONT-{uuid.uuid4().hex[:8].upper()}"
                new_contractor = {
                    "contractor_id": new_id,
                    "name": contractor_name_raw,
                    "company_name": contractor_name_raw,
                    "status": "active",
                    "is_active": True,
                    "created_by": user["user_id"],
                    "created_at": now
                }
                await db.contractors.insert_one(new_contractor)
                contractor_lookup[contractor_name_raw.lower()] = new_id
                contractor_id = new_id

        joining_date = get_val("joining_date")
        if joining_date:
            for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
                try:
                    joining_date = datetime.strptime(joining_date.split(" ")[0], fmt).strftime("%Y-%m-%d")
                    break
                except ValueError:
                    continue

        employee_code = get_val("employee_code")
        phone = get_val("phone")
        aadhar = get_val("aadhar_number")
        designation = get_val("designation")

        if employee_code:
            existing = await db.contract_workers.find_one(
                {"employee_code": employee_code, "is_active": True}
            )
            if existing:
                skipped += 1
                errors.append(f"Row {row_idx}: Employee Code '{employee_code}' already exists")
                continue

        worker = {
            "worker_id": f"CW-{uuid.uuid4().hex[:8].upper()}",
            "employee_code": employee_code,
            "name": name,
            "designation": designation,
            "phone": phone,
            "aadhar_number": aadhar,
            "joining_date": joining_date,
            "contractor_id": contractor_id,
            "status": "active",
            "is_active": True,
            "created_by": user["user_id"],
            "created_at": now
        }

        await db.contract_workers.insert_one(worker)
        created += 1

    return {
        "message": f"Uploaded {created} workers, {skipped} skipped",
        "created": created,
        "skipped": skipped,
        "errors": errors
    }


@router.get("/workers/{worker_id}")
async def get_contract_worker(worker_id: str, request: Request):
    """Get contract worker details"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "hr_executive", "manager"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    worker = await db.contract_workers.find_one({"worker_id": worker_id}, {"_id": 0})
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")
    
    # Get attendance
    attendance = await db.contract_worker_attendance.find(
        {"worker_id": worker_id}, {"_id": 0}
    ).sort("date", -1).to_list(30)
    worker["recent_attendance"] = attendance
    
    return worker


@router.put("/workers/{worker_id}")
async def update_contract_worker(worker_id: str, data: dict, request: Request):
    """Update contract worker"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "hr_executive"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.contract_workers.update_one({"worker_id": worker_id}, {"$set": data})
    return await db.contract_workers.find_one({"worker_id": worker_id}, {"_id": 0})


@router.put("/workers/{worker_id}/terminate")
async def terminate_contract_worker(worker_id: str, data: dict, request: Request):
    """Terminate contract worker"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    await db.contract_workers.update_one(
        {"worker_id": worker_id},
        {"$set": {
            "status": "terminated",
            "is_active": False,
            "termination_date": data.get("termination_date", datetime.now(timezone.utc).date().isoformat()),
            "termination_reason": data.get("reason"),
            "terminated_by": user["user_id"],
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    return {"message": "Worker terminated"}


# ==================== CONTRACT WORKER ATTENDANCE ====================

@router.get("/attendance")
async def list_contract_worker_attendance(
    request: Request,
    worker_id: Optional[str] = None,
    contractor_id: Optional[str] = None,
    date: Optional[str] = None
):
    """List contract worker attendance"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "hr_executive", "manager"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    query = {}
    if worker_id:
        query["worker_id"] = worker_id
    if contractor_id:
        query["contractor_id"] = contractor_id
    if date:
        query["date"] = date
    
    attendance = await db.contract_worker_attendance.find(query, {"_id": 0}).sort("date", -1).to_list(500)
    return attendance


@router.post("/attendance")
async def mark_contract_worker_attendance(data: dict, request: Request):
    """Mark contract worker attendance"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "hr_executive", "manager"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    attendance = {
        "attendance_id": f"cwa_{uuid.uuid4().hex[:12]}",
        "worker_id": data.get("worker_id"),
        "contractor_id": data.get("contractor_id"),
        "date": data.get("date", datetime.now(timezone.utc).date().isoformat()),
        "status": data.get("status", "present"),  # present, absent, half_day
        "in_time": data.get("in_time"),
        "out_time": data.get("out_time"),
        "hours_worked": data.get("hours_worked"),
        "overtime_hours": data.get("overtime_hours", 0),
        "remarks": data.get("remarks"),
        "marked_by": user["user_id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.contract_worker_attendance.insert_one(attendance)
    attendance.pop('_id', None)
    return attendance


@router.get("/attendance/daily-overview")
async def contract_worker_daily_attendance(
    request: Request,
    date: Optional[str] = None,
    contractor_id: Optional[str] = None
):
    """Get daily attendance overview for all contract workers"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "hr_executive", "manager"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    check_date = date or datetime.now(timezone.utc).date().isoformat()

    # Get all active workers
    wq = {"is_active": True}
    if contractor_id:
        wq["contractor_id"] = contractor_id
    workers = await db.contract_workers.find(wq, {"_id": 0}).to_list(2000)

    # Get attendance records for the date
    aq = {"date": check_date}
    if contractor_id:
        aq["contractor_id"] = contractor_id
    records = await db.contract_worker_attendance.find(aq, {"_id": 0}).to_list(2000)
    att_map = {r["worker_id"]: r for r in records}

    # Get contractor names
    contractors = await db.contractors.find({}, {"_id": 0, "contractor_id": 1, "name": 1}).to_list(200)
    contractor_names = {c["contractor_id"]: c["name"] for c in contractors}

    result = []
    present = 0
    absent = 0
    for w in workers:
        wid = w["worker_id"]
        att = att_map.get(wid)
        worker_name = w.get("name") or f"{w.get('first_name', '')} {w.get('last_name', '')}".strip() or wid
        entry = {
            "worker_id": wid,
            "name": worker_name,
            "employee_code": w.get("employee_code", ""),
            "contractor_id": w.get("contractor_id", ""),
            "contractor_name": contractor_names.get(w.get("contractor_id", ""), ""),
            "designation": w.get("designation", ""),
            "status": att["status"] if att else "absent",
            "in_time": att.get("in_time") if att else None,
            "out_time": att.get("out_time") if att else None,
            "hours_worked": att.get("hours_worked") if att else None,
            "source": att.get("source", "manual") if att else None,
        }
        if att:
            present += 1
        else:
            absent += 1
        result.append(entry)

    return {
        "date": check_date,
        "total_workers": len(workers),
        "present": present,
        "absent": absent,
        "workers": result
    }


@router.get("/attendance/monthly-summary")
async def contract_worker_monthly_summary(
    request: Request,
    month: Optional[str] = None,
    contractor_id: Optional[str] = None
):
    """Get monthly attendance summary for contract workers.
    month format: YYYY-MM (defaults to current month)"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "hr_executive", "manager"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    if not month:
        month = datetime.now(timezone.utc).strftime("%Y-%m")

    year, mon = month.split("-")
    year, mon = int(year), int(mon)

    # Calculate days in month
    import calendar
    days_in_month = calendar.monthrange(year, mon)[1]
    from_date = f"{month}-01"
    to_date = f"{month}-{days_in_month:02d}"

    # Get all active workers
    wq = {"is_active": True}
    if contractor_id:
        wq["contractor_id"] = contractor_id
    workers = await db.contract_workers.find(wq, {"_id": 0}).to_list(2000)

    # Get attendance for the month
    aq = {"date": {"$gte": from_date, "$lte": to_date}}
    if contractor_id:
        aq["contractor_id"] = contractor_id
    records = await db.contract_worker_attendance.find(aq, {"_id": 0}).to_list(10000)

    # Build per-worker attendance map: worker_id -> {date -> record}
    worker_att = {}
    for r in records:
        wid = r["worker_id"]
        if wid not in worker_att:
            worker_att[wid] = {}
        worker_att[wid][r["date"]] = {
            "status": r.get("status", "present"),
            "in_time": r.get("in_time"),
            "out_time": r.get("out_time"),
            "hours_worked": r.get("hours_worked"),
        }

    # Get contractor names
    contractors = await db.contractors.find({}, {"_id": 0, "contractor_id": 1, "name": 1}).to_list(200)
    contractor_names = {c["contractor_id"]: c["name"] for c in contractors}

    result = []
    for w in workers:
        wid = w["worker_id"]
        days = worker_att.get(wid, {})
        total_present = sum(1 for d in days.values() if d["status"] == "present")
        total_hours = sum(d.get("hours_worked") or 0 for d in days.values())
        worker_name = w.get("name") or f"{w.get('first_name', '')} {w.get('last_name', '')}".strip() or wid

        result.append({
            "worker_id": wid,
            "name": worker_name,
            "employee_code": w.get("employee_code", ""),
            "contractor_id": w.get("contractor_id", ""),
            "contractor_name": contractor_names.get(w.get("contractor_id", ""), ""),
            "designation": w.get("designation", ""),
            "days_present": total_present,
            "days_absent": days_in_month - total_present,
            "total_hours": round(total_hours, 1),
            "daily_attendance": days,
        })

    return {
        "month": month,
        "days_in_month": days_in_month,
        "total_workers": len(workers),
        "workers": result
    }


# ==================== SUMMARY & REPORTS ====================

@router.get("/summary")
async def get_labour_summary(request: Request):
    """Get labour management summary"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "hr_executive"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    today = datetime.now(timezone.utc).date().isoformat()
    
    total_contractors = await db.contractors.count_documents({"is_active": True})
    total_workers = await db.contract_workers.count_documents({"is_active": True})
    
    # Workers by department
    by_department = await db.contract_workers.aggregate([
        {"$match": {"is_active": True}},
        {"$group": {"_id": "$department_id", "count": {"$sum": 1}}}
    ]).to_list(20)
    
    # Today's attendance
    present_today = await db.contract_worker_attendance.count_documents({
        "date": today, "status": "present"
    })
    
    return {
        "total_contractors": total_contractors,
        "total_workers": total_workers,
        "present_today": present_today,
        "by_department": by_department
    }


# ==================== MONTHLY LABOUR RECORDS ====================

@router.get("/monthly-records")
async def list_monthly_records(
    request: Request,
    contractor_id: Optional[str] = None
):
    """List monthly labour records for a contractor"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "hr_executive"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    query = {}
    if contractor_id:
        query["contractor_id"] = contractor_id
    
    records = await db.contractor_monthly_records.find(query, {"_id": 0}).sort("month", -1).to_list(100)
    return records


@router.post("/monthly-records")
async def create_monthly_record(data: dict, request: Request):
    """Create monthly labour record"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "hr_executive"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    contractor_id = data.get("contractor_id")
    month = data.get("month")  # Format: YYYY-MM
    labour_count = data.get("labour_count")
    payment_amount = data.get("payment_amount")
    
    if not contractor_id or not month or not labour_count or not payment_amount:
        raise HTTPException(status_code=400, detail="Missing required fields")
    
    # Check if record already exists for this month
    existing = await db.contractor_monthly_records.find_one({
        "contractor_id": contractor_id,
        "month": month
    })
    
    if existing:
        raise HTTPException(status_code=400, detail=f"Record for {month} already exists. Please edit the existing record.")
    
    record = {
        "record_id": f"CMR-{uuid.uuid4().hex[:8].upper()}",
        "contractor_id": contractor_id,
        "month": month,
        "labour_count": int(labour_count),
        "payment_amount": float(payment_amount),
        "created_by": user["user_id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.contractor_monthly_records.insert_one(record)
    record.pop('_id', None)
    return record


@router.put("/monthly-records/{record_id}")
async def update_monthly_record(record_id: str, data: dict, request: Request):
    """Update monthly labour record"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "hr_executive"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    existing = await db.contractor_monthly_records.find_one({"record_id": record_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Record not found")
    
    update_data = {
        "labour_count": int(data.get("labour_count", existing.get("labour_count", 0))),
        "payment_amount": float(data.get("payment_amount", existing.get("payment_amount", 0))),
        "updated_by": user["user_id"],
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.contractor_monthly_records.update_one(
        {"record_id": record_id},
        {"$set": update_data}
    )
    
    return {"message": "Record updated"}


@router.delete("/monthly-records/{record_id}")
async def delete_monthly_record(record_id: str, request: Request):
    """Delete monthly labour record"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    result = await db.contractor_monthly_records.delete_one({"record_id": record_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Record not found")
    
    return {"message": "Record deleted"}


# ==================== WORKER DOCUMENTS ====================

@router.get("/workers/{worker_id}/documents")
async def list_worker_documents(worker_id: str, request: Request):
    """List documents for a contract worker"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "hr_executive"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    documents = await db.worker_documents.find(
        {"worker_id": worker_id},
        {"_id": 0}
    ).to_list(50)
    return documents


@router.post("/workers/{worker_id}/documents")
async def upload_worker_document(worker_id: str, request: Request):
    """Upload a document for a contract worker"""
    from fastapi import UploadFile
    import base64
    
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "hr_executive"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Check worker exists
    worker = await db.contract_workers.find_one({"worker_id": worker_id})
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")
    
    # Parse multipart form data
    form = await request.form()
    document_type = form.get("document_type")
    file = form.get("file")
    
    if not document_type or not file:
        raise HTTPException(status_code=400, detail="Document type and file are required")
    
    # Read file content and encode as base64
    file_content = await file.read()
    file_base64 = base64.b64encode(file_content).decode('utf-8')
    
    doc = {
        "document_id": f"doc_{uuid.uuid4().hex[:12]}",
        "worker_id": worker_id,
        "document_type": document_type,
        "file_name": file.filename,
        "file_data": file_base64,
        "file_url": f"/api/labour/workers/{worker_id}/documents/{document_type}",
        "content_type": file.content_type,
        "uploaded_by": user["user_id"],
        "uploaded_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Upsert - replace if same document type exists
    await db.worker_documents.update_one(
        {"worker_id": worker_id, "document_type": document_type},
        {"$set": doc},
        upsert=True
    )
    
    return {"message": "Document uploaded", "document_id": doc["document_id"]}


@router.get("/workers/{worker_id}/documents/{document_type}")
async def download_worker_document(worker_id: str, document_type: str, request: Request):
    """Download a specific document"""
    from fastapi.responses import Response
    import base64
    
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "hr_executive"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    doc = await db.worker_documents.find_one({
        "worker_id": worker_id,
        "document_type": document_type
    })
    
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    
    file_content = base64.b64decode(doc["file_data"])
    
    return Response(
        content=file_content,
        media_type=doc.get("content_type", "application/octet-stream"),
        headers={
            "Content-Disposition": f"attachment; filename={doc.get('file_name', 'document')}"
        }
    )


