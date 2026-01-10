"""Bulk Import API Routes"""
from fastapi import APIRouter, HTTPException, Request, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from typing import List, Optional
from datetime import datetime, timezone
import uuid
import io
import csv
import calendar
from motor.motor_asyncio import AsyncIOMotorClient
import os

router = APIRouter(prefix="/import", tags=["Bulk Import"])

mongo_url = os.environ.get('MONGO_URL', 'mongodb://localhost:27017')
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ.get('DB_NAME', 'test_database')]


async def get_current_user(request: Request) -> dict:
    from server import get_current_user as auth_get_user
    return await auth_get_user(request)


# ==================== EXCEL TEMPLATES ====================

@router.get("/templates/employees")
async def download_employee_template(request: Request):
    """Download employee import template as Excel"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    try:
        import xlsxwriter
        
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet('Employees')
        
        # Header format
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#4472C4',
            'font_color': 'white',
            'border': 1
        })
        
        headers = [
            "emp_code*", "first_name*", "last_name*", "email*", "phone", 
            "date_of_birth", "gender", "address", "city", "state", "pincode",
            "department_code", "designation_code", "location_code",
            "employment_type", "joining_date", "reporting_manager_email"
        ]
        
        # Required field format (red background)
        required_format = workbook.add_format({
            'bold': True,
            'bg_color': '#C00000',
            'font_color': 'white',
            'border': 1
        })
        
        for col, header in enumerate(headers):
            if '*' in header:
                worksheet.write(0, col, header, required_format)
            else:
                worksheet.write(0, col, header, header_format)
            worksheet.set_column(col, col, 15)
        
        # Add note row
        note_format = workbook.add_format({'italic': True, 'font_color': 'gray'})
        worksheet.write(1, 0, "* = Required fields. emp_code must be unique.", note_format)
        
        # Sample row (moved to row 3)
        sample = [
            "EMP001", "Rahul", "Sharma", "rahul.sharma@company.com", "+91 9876543210",
            "1990-05-15", "male", "123 Main Street", "Mumbai", "Maharashtra", "400001",
            "ENG", "SE", "HO-MUM", "permanent", "2024-01-15", "manager@company.com"
        ]
        for col, val in enumerate(sample):
            worksheet.write(2, col, val)
        
        workbook.close()
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=employee_import_template.xlsx",
                "Cache-Control": "no-cache, no-store, must-revalidate"
            }
        )
    except ImportError:
        raise HTTPException(status_code=500, detail="Excel library not available")


@router.get("/templates/salary")
async def download_salary_template(request: Request):
    """Download salary structure import template as Excel"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    try:
        import xlsxwriter
        
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet('Salary Structure')
        
        # Header formats
        header_blue = workbook.add_format({
            'bold': True, 'bg_color': '#4472C4', 'font_color': 'white', 
            'border': 1, 'align': 'center'
        })
        header_green = workbook.add_format({
            'bold': True, 'bg_color': '#70AD47', 'font_color': 'white', 
            'border': 1, 'align': 'center'
        })
        header_orange = workbook.add_format({
            'bold': True, 'bg_color': '#ED7D31', 'font_color': 'white', 
            'border': 1, 'align': 'center'
        })
        header_red = workbook.add_format({
            'bold': True, 'bg_color': '#C00000', 'font_color': 'white', 
            'border': 1, 'align': 'center'
        })
        header_purple = workbook.add_format({
            'bold': True, 'bg_color': '#7030A0', 'font_color': 'white', 
            'border': 1, 'align': 'center'
        })
        
        # All headers in order
        headers = [
            # Employee Info
            ("Emp Code", header_blue),
            ("Name of Employees", header_blue),
            # Fixed Salary
            ("BASIC", header_green),
            ("DA", header_green),
            ("HRA", header_green),
            ("Conveyance", header_green),
            ("GRADE PAY", header_green),
            ("OTHER ALLOW", header_green),
            ("Med./Spl. Allow", header_green),
            ("Total Salary (FIXED)", header_green),
            # Attendance
            ("Work from office", header_orange),
            ("Sunday + Holiday Leave Days", header_orange),
            ("Leave Days", header_orange),
            ("Work from Home @50%", header_orange),
            # Earned (calculated columns - user fills)
            ("Late Deduction", header_orange),
            ("Basic+DA (Earned)", header_orange),
            ("HRA (Earned)", header_orange),
            ("Conveyance (Earned)", header_orange),
            ("GRADE PAY (Earned)", header_orange),
            ("OTHER ALLOW (Earned)", header_orange),
            ("Med./Spl. Allow (Earned)", header_orange),
            ("Total Earned Days", header_orange),
            ("Total Salary Earned", header_orange),
            # Deductions
            ("EPF Employees", header_red),
            ("ESI Employees", header_red),
            ("SEWA", header_red),
            ("Sewa Advance", header_red),
            ("Other Deduction", header_red),
            ("Total Deduction", header_red),
            # Final
            ("NET PAYABLE", header_purple),
        ]
        
        for col, (header, fmt) in enumerate(headers):
            worksheet.write(0, col, header, fmt)
            worksheet.set_column(col, col, 18)
        
        # Sample row
        sample = [
            "EMP001", "Rahul Sharma",
            25000, 2500, 10000, 2000, 3000, 2000, 1500, 46000,  # Fixed
            22, 4, 2, 2,  # Attendance
            500, 22000, 8000, 1600, 2400, 1600, 1200, 22, 36800,  # Earned
            3000, 150, 500, 0, 0, 3650,  # Deductions
            33150  # Net
        ]
        for col, val in enumerate(sample):
            worksheet.write(1, col, val)
        
        workbook.close()
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=salary_structure_template.xlsx",
                "Cache-Control": "no-cache, no-store, must-revalidate"
            }
        )
    except ImportError:
        raise HTTPException(status_code=500, detail="Excel library not available")


@router.get("/templates/attendance")
async def download_attendance_template(request: Request, month: int = None, year: int = None):
    """Download attendance import template as Excel"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Default to current month if not specified
    now = datetime.now()
    month = month or now.month
    year = year or now.year
    
    try:
        import xlsxwriter
        
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet('Attendance')
        
        # Header format
        header_format = workbook.add_format({
            'bold': True, 'bg_color': '#4472C4', 'font_color': 'white',
            'border': 1, 'align': 'center'
        })
        day_header_format = workbook.add_format({
            'bold': True, 'bg_color': '#70AD47', 'font_color': 'white',
            'border': 1, 'align': 'center'
        })
        weekend_header_format = workbook.add_format({
            'bold': True, 'bg_color': '#ED7D31', 'font_color': 'white',
            'border': 1, 'align': 'center'
        })
        
        # Get days in month
        days_in_month = calendar.monthrange(year, month)[1]
        
        # Headers
        headers = ["SL NO", "Emp Code", "Name of Employees"]
        worksheet.write(0, 0, headers[0], header_format)
        worksheet.write(0, 1, headers[1], header_format)
        worksheet.write(0, 2, headers[2], header_format)
        worksheet.set_column(0, 0, 8)
        worksheet.set_column(1, 1, 12)
        worksheet.set_column(2, 2, 25)
        
        # Day columns
        for day in range(1, days_in_month + 1):
            col = day + 2
            # Check if weekend (Sunday)
            date = datetime(year, month, day)
            fmt = weekend_header_format if date.weekday() == 6 else day_header_format
            worksheet.write(0, col, str(day), fmt)
            worksheet.set_column(col, col, 5)
        
        # Add legend row
        legend_row = days_in_month + 5
        worksheet.write(legend_row, 0, "Legend:", header_format)
        worksheet.write(legend_row, 1, "P = Present", None)
        worksheet.write(legend_row, 2, "A = Absent", None)
        worksheet.write(legend_row, 3, "L = Leave", None)
        worksheet.write(legend_row, 4, "H = Holiday", None)
        worksheet.write(legend_row, 5, "WFH = Work from Home", None)
        
        # Sample row
        worksheet.write(1, 0, 1)
        worksheet.write(1, 1, "EMP001")
        worksheet.write(1, 2, "Rahul Sharma")
        for day in range(1, days_in_month + 1):
            date = datetime(year, month, day)
            if date.weekday() == 6:  # Sunday
                worksheet.write(1, day + 2, "H")
            else:
                worksheet.write(1, day + 2, "P")
        
        workbook.close()
        output.seek(0)
        
        month_name = calendar.month_name[month]
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=attendance_template_{month_name}_{year}.xlsx",
                "Cache-Control": "no-cache, no-store, must-revalidate"
            }
        )
    except ImportError:
        raise HTTPException(status_code=500, detail="Excel library not available")


@router.get("/templates/leave-balance")
async def download_leave_balance_template(request: Request):
    """Download leave balance import template as Excel"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    try:
        import xlsxwriter
        
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet('Leave Balance')
        
        # Header formats
        header_blue = workbook.add_format({
            'bold': True, 'bg_color': '#4472C4', 'font_color': 'white',
            'border': 1, 'align': 'center'
        })
        required_format = workbook.add_format({
            'bold': True, 'bg_color': '#C00000', 'font_color': 'white',
            'border': 1, 'align': 'center'
        })
        leave_format = workbook.add_format({
            'bold': True, 'bg_color': '#70AD47', 'font_color': 'white',
            'border': 1, 'align': 'center'
        })
        note_format = workbook.add_format({'italic': True, 'font_color': 'gray'})
        
        # Title row
        title_format = workbook.add_format({
            'bold': True, 'font_size': 12, 'align': 'left'
        })
        worksheet.write(0, 0, f"Leave Balance as on {datetime.now().strftime('%d %b %Y')}", title_format)
        
        # Headers (Row 2 - index 1)
        headers = [
            ("Emp ID*", required_format),
            ("Name", header_blue),
            ("Casual Leave (CL)", leave_format),
            ("Sick Leave (SL)", leave_format),
            ("Earned Leave (EL)", leave_format),
            ("Complementary Off", leave_format),
        ]
        
        for col, (header, fmt) in enumerate(headers):
            worksheet.write(1, col, header, fmt)
            worksheet.set_column(col, col, 20 if col <= 1 else 18)
        
        # Note row
        worksheet.write(2, 0, "* = Required. Emp ID must match employee code in system.", note_format)
        
        # Sample rows
        sample_data = [
            ("EMP001", "Rahul Sharma", 8, 6, 12, 2),
            ("EMP002", "Priya Singh", 10, 5, 15, 0),
        ]
        
        for row_idx, data in enumerate(sample_data, start=3):
            for col, val in enumerate(data):
                worksheet.write(row_idx, col, val)
        
        workbook.close()
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=leave_balance_template.xlsx",
                "Cache-Control": "no-cache, no-store, must-revalidate"
            }
        )
    except ImportError:
        raise HTTPException(status_code=500, detail="Excel library not available")


@router.post("/leave-balance")
async def import_leave_balance(request: Request, file: UploadFile = File(...)):
    """Bulk import leave balances from Excel"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    filename = file.filename.lower()
    content = await file.read()
    
    if filename.endswith('.xlsx'):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        
        # Find the header row (look for "Emp ID" in first few rows)
        header_row = None
        headers = []
        for row_num in range(1, min(6, ws.max_row + 1)):
            first_cell = str(ws.cell(row=row_num, column=1).value or "").strip().lower()
            if "emp" in first_cell and "id" in first_cell:
                header_row = row_num
                headers = [(ws.cell(row=row_num, column=col).value or "").strip() for col in range(1, ws.max_column + 1)]
                break
        
        if not header_row:
            raise HTTPException(status_code=400, detail="Could not find header row with 'Emp ID'")
        
        rows = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            # Skip empty rows and note rows
            first_cell = str(row[0] or "").strip() if row[0] else ""
            if not first_cell or first_cell.startswith("*") or first_cell.lower().startswith("required"):
                continue
            rows.append(dict(zip(headers, row)))
    elif filename.endswith('.csv'):
        decoded = content.decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(decoded))
        rows = list(reader)
    else:
        raise HTTPException(status_code=400, detail="Only CSV and Excel files are supported")
    
    # Column mapping - normalize various possible column names
    def find_column(row, *possible_names):
        for name in possible_names:
            for key in row.keys():
                if key and name.lower() in key.lower():
                    return row.get(key)
        return None
    
    # Leave type definitions - will create if not exists
    LEAVE_TYPES = [
        {"code": "CL", "name": "Casual Leave", "search_terms": ["casual", "cl"]},
        {"code": "SL", "name": "Sick Leave", "search_terms": ["sick", "sl"]},
        {"code": "EL", "name": "Earned Leave", "search_terms": ["earned", "el"]},
        {"code": "CO", "name": "Complementary Off", "search_terms": ["complementary", "comp off", "co"]},
    ]
    
    # Ensure leave types exist in the system
    current_year = datetime.now().year
    for lt in LEAVE_TYPES:
        existing = await db.leave_types.find_one({"code": lt["code"]})
        if not existing:
            await db.leave_types.insert_one({
                "leave_type_id": f"lt_{lt['code'].lower()}",
                "code": lt["code"],
                "name": lt["name"],
                "description": lt["name"],
                "annual_quota": 12,  # Default quota
                "is_paid": True,
                "is_active": True,
                "created_at": datetime.now(timezone.utc).isoformat()
            })
    
    # Get leave type IDs for mapping
    leave_type_map = {}
    all_leave_types = await db.leave_types.find({}, {"_id": 0}).to_list(20)
    for lt in all_leave_types:
        leave_type_map[lt["code"]] = lt["leave_type_id"]
    
    imported = 0
    errors = []
    updated_employees = []
    
    # Convert to number helper
    def to_float(val):
        if val is None or val == "":
            return 0.0
        try:
            return float(val)
        except:
            return 0.0
    
    for idx, row in enumerate(rows, start=header_row + 2 if header_row else 2):
        try:
            # Get employee ID
            emp_id = find_column(row, "emp id", "emp_id", "empid", "employee id", "employee_id")
            emp_id = str(emp_id or "").strip().replace('*', '')
            
            if not emp_id:
                errors.append({"row": idx, "error": "Missing Emp ID"})
                continue
            
            # Find employee by emp_code
            employee = await db.employees.find_one({"emp_code": emp_id}, {"_id": 0})
            if not employee:
                # Also try employee_id field
                employee = await db.employees.find_one({"employee_id": emp_id}, {"_id": 0})
            
            if not employee:
                errors.append({"row": idx, "error": f"Employee not found: {emp_id}"})
                continue
            
            employee_id = employee.get("employee_id")
            employee_name = f"{employee.get('first_name', '')} {employee.get('last_name', '')}".strip()
            
            # Parse leave balances from the row
            balances_to_import = []
            for lt in LEAVE_TYPES:
                value = find_column(row, *lt["search_terms"])
                if value is not None:
                    balances_to_import.append({
                        "code": lt["code"],
                        "leave_type_id": leave_type_map.get(lt["code"]),
                        "value": to_float(value)
                    })
            
            # Create/update individual leave balance records for each leave type
            for balance in balances_to_import:
                if not balance["leave_type_id"]:
                    continue
                
                balance_doc = {
                    "employee_id": employee_id,
                    "emp_code": emp_id,
                    "employee_name": employee_name,
                    "leave_type_id": balance["leave_type_id"],
                    "year": current_year,
                    "opening_balance": balance["value"],  # Imported value is the current/opening balance
                    "accrued": 0,
                    "used": 0,
                    "pending": 0,
                    "available": balance["value"],  # Available = opening - used
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                    "updated_by": user["user_id"],
                    "import_date": datetime.now(timezone.utc).isoformat(),
                }
                
                # Upsert - overwrite existing balance for this employee, leave type, and year
                await db.leave_balances.update_one(
                    {
                        "employee_id": employee_id, 
                        "leave_type_id": balance["leave_type_id"],
                        "year": current_year
                    },
                    {"$set": balance_doc},
                    upsert=True
                )
            
            imported += 1
            updated_employees.append(emp_id)
            
        except Exception as e:
            errors.append({"row": idx, "error": str(e)})
    
    return {
        "message": "Leave balance import completed",
        "imported": imported,
        "errors": errors,
        "total_rows": len(rows),
        "updated_employees": updated_employees[:10],  # Show first 10
        "year": current_year,
        "leave_types_created": [lt["code"] for lt in LEAVE_TYPES]
    }


# ==================== IMPORTS ====================

@router.post("/employees")
async def import_employees(request: Request, file: UploadFile = File(...)):
    """Bulk import employees from CSV or Excel"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    filename = file.filename.lower()
    content = await file.read()
    
    if filename.endswith('.xlsx'):
        # Parse Excel
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        # Normalize headers - remove asterisks and strip whitespace
        headers = [(cell.value or "").replace('*', '').strip() for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Skip note rows and empty rows
            first_cell = str(row[0] or "").strip() if row[0] else ""
            if first_cell.startswith("*") or first_cell.startswith("Required") or not any(row):
                continue
            rows.append(dict(zip(headers, row)))
    elif filename.endswith('.csv'):
        # Parse CSV
        decoded = content.decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(decoded))
        rows = list(reader)
    else:
        raise HTTPException(status_code=400, detail="Only CSV and Excel files are supported")
    
    # Get lookup data - handle missing fields gracefully
    departments = {}
    for d in await db.departments.find({}, {"_id": 0}).to_list(100):
        if "code" in d and "department_id" in d:
            departments[d["code"]] = d["department_id"]
    
    designations = {}
    for d in await db.designations.find({}, {"_id": 0}).to_list(100):
        if "code" in d and "designation_id" in d:
            designations[d["code"]] = d["designation_id"]
    
    locations = {}
    for l in await db.locations.find({}, {"_id": 0}).to_list(100):
        if "code" in l and "location_id" in l:
            locations[l["code"]] = l["location_id"]
    
    imported = 0
    errors = []
    default_password = "Welcome@123"
    
    # Import bcrypt for password hashing
    import bcrypt
    import re
    
    # Email validation pattern
    email_pattern = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')
    
    for idx, row in enumerate(rows, start=2):
        try:
            first_name = str(row.get("first_name") or "").strip()
            last_name = str(row.get("last_name") or "").strip()
            email = str(row.get("email") or "").strip().rstrip('.')  # Remove trailing periods
            
            if not first_name or not last_name or not email:
                errors.append({"row": idx, "error": "Missing required fields (first_name, last_name, email)"})
                continue
            
            # Validate email format
            if not email_pattern.match(email):
                errors.append({"row": idx, "error": f"Invalid email format: {email}"})
                continue
            
            # Check for duplicate email
            existing_email = await db.employees.find_one({"email": email.lower()})
            if existing_email:
                errors.append({"row": idx, "error": f"Email {email} already exists"})
                continue
            
            emp_code = str(row.get("emp_code") or "").strip()
            if not emp_code:
                # emp_code is required - do not auto-generate
                errors.append({"row": idx, "error": "Missing required field: emp_code (Employee Code is mandatory)"})
                continue
            
            # Check for duplicate emp_code
            existing_code = await db.employees.find_one({"emp_code": emp_code})
            if existing_code:
                errors.append({"row": idx, "error": f"Employee code {emp_code} already exists"})
                continue
            
            employee_id = f"EMP{uuid.uuid4().hex[:8].upper()}"
            
            employee = {
                "employee_id": employee_id,
                "emp_code": emp_code,
                "first_name": first_name,
                "last_name": last_name,
                "email": email.lower(),
                "phone": str(row.get("phone") or "").strip() or None,
                "date_of_birth": str(row.get("date_of_birth") or "").strip() or None,
                "gender": str(row.get("gender") or "").strip().lower() or None,
                "address": str(row.get("address") or "").strip() or None,
                "city": str(row.get("city") or "").strip() or None,
                "state": str(row.get("state") or "").strip() or None,
                "pincode": str(row.get("pincode") or "").strip() or None,
                "department_id": departments.get(str(row.get("department_code") or "").strip()),
                "designation_id": designations.get(str(row.get("designation_code") or "").strip()),
                "location_id": locations.get(str(row.get("location_code") or "").strip()),
                "employment_type": str(row.get("employment_type") or "permanent").strip().lower(),
                "joining_date": str(row.get("joining_date") or "").strip() or None,
                "status": "active",
                "is_active": True,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
            
            await db.employees.insert_one(employee)
            
            # Create user account with default password
            hashed_password = bcrypt.hashpw(default_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
            user_doc = {
                "user_id": f"user_{uuid.uuid4().hex[:12]}",
                "email": email.lower(),
                "password": hashed_password,
                "name": f"{first_name} {last_name}",
                "role": "employee",
                "roles": ["employee"],
                "permissions": [],
                "employee_id": employee_id,
                "department_id": employee.get("department_id"),
                "is_active": True,
                "must_change_password": True,  # Flag to prompt password change on first login
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
            
            # Check if user already exists
            existing_user = await db.users.find_one({"email": email.lower()})
            if not existing_user:
                await db.users.insert_one(user_doc)
            
            imported += 1
            
        except Exception as e:
            errors.append({"row": idx, "error": str(e)})
    
    return {
        "message": "Import completed",
        "imported": imported,
        "errors": errors,
        "total_rows": imported + len(errors),
        "info": f"User accounts created with default password: {default_password}" if imported > 0 else None
    }


@router.post("/attendance")
async def import_attendance(
    request: Request, 
    file: UploadFile = File(...),
    month: int = Form(...),
    year: int = Form(...)
):
    """Bulk import attendance from Excel"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    filename = file.filename.lower()
    content = await file.read()
    
    if not filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Only Excel files are supported for attendance import")
    
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        
        # Get headers (first row)
        headers = [cell.value for cell in ws[1]]
        
        # Find emp_code column index
        emp_code_idx = None
        name_idx = None
        for i, h in enumerate(headers):
            if h and "emp" in str(h).lower() and "code" in str(h).lower():
                emp_code_idx = i
            elif h and "name" in str(h).lower():
                name_idx = i
        
        if emp_code_idx is None:
            raise HTTPException(status_code=400, detail="Could not find 'Emp Code' column")
        
        days_in_month = calendar.monthrange(year, month)[1]
        
        imported = 0
        errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not any(row):
                    continue
                
                emp_code = str(row[emp_code_idx] or "").strip()
                if not emp_code:
                    continue
                
                # Find employee
                employee = await db.employees.find_one({
                    "$or": [
                        {"emp_code": emp_code},
                        {"employee_id": emp_code}
                    ]
                })
                
                if not employee:
                    errors.append({"row": row_idx, "error": f"Employee {emp_code} not found"})
                    continue
                
                employee_id = employee["employee_id"]
                
                # Process each day
                for day in range(1, days_in_month + 1):
                    col_idx = day + 2  # After SL NO, Emp Code, Name
                    if col_idx < len(row):
                        status_val = str(row[col_idx] or "").strip().upper()
                        
                        if status_val:
                            date_str = f"{year}-{str(month).zfill(2)}-{str(day).zfill(2)}"
                            
                            # Map status codes
                            status_map = {
                                "P": "present",
                                "A": "absent",
                                "L": "leave",
                                "H": "holiday",
                                "WFH": "wfh",
                                "W/O": "weekly_off",
                                "HD": "half_day"
                            }
                            status = status_map.get(status_val, status_val.lower())
                            
                            attendance_doc = {
                                "attendance_id": f"att_{uuid.uuid4().hex[:12]}",
                                "employee_id": employee_id,
                                "emp_code": emp_code,
                                "employee_name": employee.get("first_name", "") + " " + employee.get("last_name", ""),
                                "date": date_str,
                                "status": status,
                                "is_wfh": status == "wfh",
                                "source": "bulk_import",
                                "created_at": datetime.now(timezone.utc).isoformat()
                            }
                            
                            # Upsert
                            await db.attendance.update_one(
                                {"employee_id": employee_id, "date": date_str},
                                {"$set": attendance_doc},
                                upsert=True
                            )
                
                imported += 1
                
            except Exception as e:
                errors.append({"row": row_idx, "error": str(e)})
        
        return {
            "message": "Attendance import completed",
            "imported": imported,
            "errors": errors,
            "total_rows": imported + len(errors),
            "month": month,
            "year": year
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")


@router.post("/salary")
async def import_salary(request: Request, file: UploadFile = File(...)):
    """Bulk import salary structures from Excel"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    filename = file.filename.lower()
    content = await file.read()
    
    if filename.endswith('.xlsx'):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        headers = [str(cell.value or "").strip() for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row):
                rows.append(dict(zip(headers, row)))
    elif filename.endswith('.csv'):
        decoded = content.decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(decoded))
        rows = list(reader)
    else:
        raise HTTPException(status_code=400, detail="Only CSV and Excel files are supported")
    
    imported = 0
    errors = []
    
    # Column mapping (template column name -> internal field)
    column_map = {
        "Emp Code": "emp_code",
        "emp_code": "emp_code",
        "Name of Employees": "name",
        "BASIC": "basic",
        "DA": "da",
        "HRA": "hra",
        "Conveyance": "conveyance",
        "GRADE PAY": "grade_pay",
        "OTHER ALLOW": "other_allowance",
        "Med./Spl. Allow": "medical_allowance",
        "Total Salary (FIXED)": "total_fixed",
        "Work from office": "work_from_office_days",
        "Sunday + Holiday Leave Days": "holiday_days",
        "Leave Days": "leave_days",
        "Work from Home @50%": "wfh_days",
        "Late Deduction": "late_deduction",
        "Basic+DA (Earned)": "basic_da_earned",
        "HRA (Earned)": "hra_earned",
        "Conveyance (Earned)": "conveyance_earned",
        "GRADE PAY (Earned)": "grade_pay_earned",
        "OTHER ALLOW (Earned)": "other_allowance_earned",
        "Med./Spl. Allow (Earned)": "medical_allowance_earned",
        "Total Earned Days": "total_earned_days",
        "Total Salary Earned": "total_earned",
        "EPF Employees": "epf",
        "ESI Employees": "esi",
        "SEWA": "sewa",
        "Sewa Advance": "sewa_advance",
        "Other Deduction": "other_deduction",
        "Total Deduction": "total_deduction",
        "NET PAYABLE": "net_payable"
    }
    
    def get_val(row, key, default=0):
        """Get value from row, handling different column names"""
        for col_name, field_name in column_map.items():
            if field_name == key and col_name in row:
                val = row[col_name]
                if val is None or val == "":
                    return default
                try:
                    return float(val) if isinstance(default, (int, float)) else str(val)
                except:
                    return default
        return row.get(key, default)
    
    for idx, row in enumerate(rows, start=2):
        try:
            emp_code = get_val(row, "emp_code", "")
            if not emp_code:
                errors.append({"row": idx, "error": "Missing Emp Code"})
                continue
            
            # Find employee
            employee = await db.employees.find_one({
                "$or": [
                    {"emp_code": str(emp_code)},
                    {"employee_id": str(emp_code)}
                ]
            })
            
            if not employee:
                errors.append({"row": idx, "error": f"Employee {emp_code} not found"})
                continue
            
            employee_id = employee["employee_id"]
            
            # Build salary structure
            salary_doc = {
                "salary_id": f"sal_{uuid.uuid4().hex[:12]}",
                "employee_id": employee_id,
                "emp_code": str(emp_code),
                "employee_name": get_val(row, "name", f"{employee.get('first_name', '')} {employee.get('last_name', '')}"),
                
                # Fixed Components
                "fixed_components": {
                    "basic": get_val(row, "basic", 0),
                    "da": get_val(row, "da", 0),
                    "hra": get_val(row, "hra", 0),
                    "conveyance": get_val(row, "conveyance", 0),
                    "grade_pay": get_val(row, "grade_pay", 0),
                    "other_allowance": get_val(row, "other_allowance", 0),
                    "medical_allowance": get_val(row, "medical_allowance", 0),
                },
                "total_fixed": get_val(row, "total_fixed", 0),
                
                # Attendance
                "attendance": {
                    "work_from_office_days": get_val(row, "work_from_office_days", 0),
                    "holiday_days": get_val(row, "holiday_days", 0),
                    "leave_days": get_val(row, "leave_days", 0),
                    "wfh_days": get_val(row, "wfh_days", 0),
                },
                
                # Earned
                "earned_components": {
                    "basic_da": get_val(row, "basic_da_earned", 0),
                    "hra": get_val(row, "hra_earned", 0),
                    "conveyance": get_val(row, "conveyance_earned", 0),
                    "grade_pay": get_val(row, "grade_pay_earned", 0),
                    "other_allowance": get_val(row, "other_allowance_earned", 0),
                    "medical_allowance": get_val(row, "medical_allowance_earned", 0),
                },
                "late_deduction": get_val(row, "late_deduction", 0),
                "total_earned_days": get_val(row, "total_earned_days", 0),
                "total_earned": get_val(row, "total_earned", 0),
                
                # Deductions
                "deductions": {
                    "epf": get_val(row, "epf", 0),
                    "esi": get_val(row, "esi", 0),
                    "sewa": get_val(row, "sewa", 0),
                    "sewa_advance": get_val(row, "sewa_advance", 0),
                    "other": get_val(row, "other_deduction", 0),
                },
                "total_deduction": get_val(row, "total_deduction", 0),
                
                # Final
                "net_payable": get_val(row, "net_payable", 0),
                
                "effective_from": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
                "is_active": True,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
            
            # Deactivate existing salary
            await db.employee_salaries.update_many(
                {"employee_id": employee_id, "is_active": True},
                {"$set": {"is_active": False}}
            )
            
            await db.employee_salaries.insert_one(salary_doc)
            imported += 1
            
        except Exception as e:
            errors.append({"row": idx, "error": str(e)})
    
    return {
        "message": "Salary import completed",
        "imported": imported,
        "errors": errors,
        "total_rows": imported + len(errors)
    }


# ==================== EXPORT ====================

@router.get("/export/employees")
async def export_employees(request: Request):
    """Export all employees to Excel"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    employees = await db.employees.find({"is_active": True}, {"_id": 0}).to_list(10000)
    
    try:
        import xlsxwriter
        
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet('Employees')
        
        header_format = workbook.add_format({
            'bold': True, 'bg_color': '#4472C4', 'font_color': 'white', 'border': 1
        })
        
        headers = [
            "employee_id", "emp_code", "first_name", "last_name", "email", "phone",
            "date_of_birth", "gender", "address", "city", "state", "pincode",
            "department_id", "designation_id", "location_id",
            "employment_type", "joining_date", "status"
        ]
        
        for col, h in enumerate(headers):
            worksheet.write(0, col, h, header_format)
            worksheet.set_column(col, col, 15)
        
        for row_idx, emp in enumerate(employees, start=1):
            for col, h in enumerate(headers):
                worksheet.write(row_idx, col, emp.get(h, ""))
        
        workbook.close()
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=employees_export_{datetime.now().strftime('%Y%m%d')}.xlsx",
                "Cache-Control": "no-cache, no-store, must-revalidate"
            }
        )
    except ImportError:
        # Fallback to CSV
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for emp in employees:
            writer.writerow([emp.get(h, "") for h in headers])
        output.seek(0)
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=employees_export_{datetime.now().strftime('%Y%m%d')}.csv"}
        )


@router.get("/export/attendance")
async def export_attendance(request: Request, month: int = None, year: int = None):
    """Export attendance records to Excel"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    now = datetime.now()
    month = month or now.month
    year = year or now.year
    
    # Get date range
    start_date = f"{year}-{str(month).zfill(2)}-01"
    days_in_month = calendar.monthrange(year, month)[1]
    end_date = f"{year}-{str(month).zfill(2)}-{days_in_month}"
    
    attendance = await db.attendance.find({
        "date": {"$gte": start_date, "$lte": end_date}
    }, {"_id": 0}).to_list(50000)
    
    # Get employees
    employees = await db.employees.find({"is_active": True}, {"_id": 0}).to_list(10000)
    emp_map = {e["employee_id"]: e for e in employees}
    
    try:
        import xlsxwriter
        
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet('Attendance')
        
        header_format = workbook.add_format({
            'bold': True, 'bg_color': '#4472C4', 'font_color': 'white', 'border': 1, 'align': 'center'
        })
        
        headers = ["SL NO", "Emp Code", "Name"] + [str(d) for d in range(1, days_in_month + 1)]
        for col, h in enumerate(headers):
            worksheet.write(0, col, h, header_format)
        
        # Build attendance by employee
        att_by_emp = {}
        for att in attendance:
            eid = att["employee_id"]
            if eid not in att_by_emp:
                att_by_emp[eid] = {}
            day = int(att["date"].split("-")[2])
            status = att.get("status", "").upper()
            status_map = {"PRESENT": "P", "ABSENT": "A", "LEAVE": "L", "HOLIDAY": "H", "WFH": "WFH"}
            att_by_emp[eid][day] = status_map.get(status, status[:1] if status else "")
        
        row_idx = 1
        for emp in employees:
            emp_id = emp["employee_id"]
            worksheet.write(row_idx, 0, row_idx)
            worksheet.write(row_idx, 1, emp.get("emp_code", emp_id))
            worksheet.write(row_idx, 2, f"{emp.get('first_name', '')} {emp.get('last_name', '')}")
            
            emp_att = att_by_emp.get(emp_id, {})
            for day in range(1, days_in_month + 1):
                worksheet.write(row_idx, day + 2, emp_att.get(day, ""))
            
            row_idx += 1
        
        workbook.close()
        output.seek(0)
        
        month_name = calendar.month_name[month]
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=attendance_export_{month_name}_{year}.xlsx",
                "Cache-Control": "no-cache, no-store, must-revalidate"
            }
        )
    except ImportError:
        raise HTTPException(status_code=500, detail="Excel library not available")


@router.get("/export/salary")
async def export_salary(request: Request):
    """Export salary structures to Excel"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    salaries = await db.employee_salaries.find({"is_active": True}, {"_id": 0}).to_list(10000)
    
    try:
        import xlsxwriter
        
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet('Salary')
        
        header_format = workbook.add_format({
            'bold': True, 'bg_color': '#4472C4', 'font_color': 'white', 'border': 1
        })
        
        headers = [
            "Emp Code", "Name", "BASIC", "DA", "HRA", "Conveyance", "GRADE PAY", 
            "OTHER ALLOW", "Med./Spl. Allow", "Total Fixed",
            "EPF", "ESI", "SEWA", "Net Payable"
        ]
        
        for col, h in enumerate(headers):
            worksheet.write(0, col, h, header_format)
            worksheet.set_column(col, col, 15)
        
        for row_idx, sal in enumerate(salaries, start=1):
            fc = sal.get("fixed_components", {})
            ded = sal.get("deductions", {})
            worksheet.write(row_idx, 0, sal.get("emp_code", ""))
            worksheet.write(row_idx, 1, sal.get("employee_name", ""))
            worksheet.write(row_idx, 2, fc.get("basic", 0))
            worksheet.write(row_idx, 3, fc.get("da", 0))
            worksheet.write(row_idx, 4, fc.get("hra", 0))
            worksheet.write(row_idx, 5, fc.get("conveyance", 0))
            worksheet.write(row_idx, 6, fc.get("grade_pay", 0))
            worksheet.write(row_idx, 7, fc.get("other_allowance", 0))
            worksheet.write(row_idx, 8, fc.get("medical_allowance", 0))
            worksheet.write(row_idx, 9, sal.get("total_fixed", 0))
            worksheet.write(row_idx, 10, ded.get("epf", 0))
            worksheet.write(row_idx, 11, ded.get("esi", 0))
            worksheet.write(row_idx, 12, ded.get("sewa", 0))
            worksheet.write(row_idx, 13, sal.get("net_payable", 0))
        
        workbook.close()
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=salary_export_{datetime.now().strftime('%Y%m%d')}.xlsx",
                "Cache-Control": "no-cache, no-store, must-revalidate"
            }
        )
    except ImportError:
        raise HTTPException(status_code=500, detail="Excel library not available")


# ==================== INSURANCE IMPORT ====================

@router.get("/templates/insurance")
async def download_insurance_template(request: Request):
    """Download insurance import template as Excel"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "hr_executive"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    try:
        import xlsxwriter
        
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet('Insurance Data')
        
        # Header formats
        header_format = workbook.add_format({
            'bold': True, 'bg_color': '#4472C4', 'font_color': 'white',
            'border': 1, 'align': 'center'
        })
        required_format = workbook.add_format({
            'bold': True, 'bg_color': '#C00000', 'font_color': 'white',
            'border': 1, 'align': 'center'
        })
        note_format = workbook.add_format({'italic': True, 'font_color': 'gray'})
        
        # Title
        title_format = workbook.add_format({'bold': True, 'font_size': 12})
        worksheet.write(0, 0, "EMPLOYEE INSURANCE DATA", title_format)
        
        # Headers (Row 2)
        headers = [
            ("SL NO.", header_format),
            ("Employee Code*", required_format),
            ("Employee Name", header_format),
            ("ESIC (Yes/No)", header_format),
            ("Date", header_format),
            ("Amount", header_format),
            ("Insurance Company", header_format),
            ("Policy Number", header_format),
            ("Coverage Type", header_format),
            ("Accidental Insurance (Yes/No)", header_format),
            ("Start Date", header_format),
            ("End Date", header_format),
            ("Notes", header_format),
        ]
        
        for col, (header, fmt) in enumerate(headers):
            worksheet.write(1, col, header, fmt)
            worksheet.set_column(col, col, 18)
        
        # Note row
        worksheet.write(2, 0, "* = Required fields. Employee Code must exist in system. If ESIC=Yes, Date/Amount/Company/Policy are optional. Accidental Insurance: Yes/No", note_format)
        
        # Sample rows
        sample_data = [
            (1, "S0003", "Abritee Das Roy", "No", "2026-01-15", 50000, "LIC", "POL123456", "Health", "Yes", "2026-01-01", "2027-01-01", "Annual premium"),
            (2, "S0007", "Anup Kr Mishra", "Yes", "", "", "", "", "", "No", "", "", "ESIC covered employee"),
            (3, "S0010", "Ravi Kumar", "No", "2026-01-15", 75000, "HDFC Ergo", "POL789012", "Life", "No", "2026-01-01", "2027-01-01", ""),
        ]
        
        for row_idx, data in enumerate(sample_data, start=3):
            for col, val in enumerate(data):
                worksheet.write(row_idx, col, val)
        
        workbook.close()
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=insurance_template.xlsx",
                "Cache-Control": "no-cache, no-store, must-revalidate"
            }
        )
    except ImportError:
        raise HTTPException(status_code=500, detail="Excel library not available")


@router.post("/insurance")
async def import_insurance(request: Request, file: UploadFile = File(...)):
    """Bulk import insurance records from Excel"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "hr_executive"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    filename = file.filename.lower()
    content = await file.read()
    
    try:
        if filename.endswith('.xlsx'):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content))
            ws = wb.active
            
            # Find header row
            header_row = None
            headers = []
            for row_num in range(1, min(6, ws.max_row + 1)):
                first_cell = str(ws.cell(row=row_num, column=2).value or "").strip().lower()
                if "employee" in first_cell and "code" in first_cell:
                    header_row = row_num
                    headers = [(ws.cell(row=row_num, column=col).value or "").strip().replace('*', '') for col in range(1, ws.max_column + 1)]
                    break
            
            if not header_row:
                raise HTTPException(status_code=400, detail="Could not find header row with 'Employee Code'")
            
            rows = []
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                first_cell = str(row[1] or "").strip() if len(row) > 1 and row[1] else ""
                if not first_cell or first_cell.startswith("*") or first_cell.lower().startswith("required"):
                    continue
                rows.append(dict(zip(headers, row)))
        elif filename.endswith('.csv'):
            decoded = content.decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(decoded))
            rows = list(reader)
        else:
            raise HTTPException(status_code=400, detail="Only CSV and Excel files are supported")
        
        imported = 0
        errors = []
        
        for idx, row in enumerate(rows, start=header_row + 2 if header_row else 2):
            try:
                # Get employee code
                emp_code = None
                for key in row.keys():
                    if key and "employee" in key.lower() and "code" in key.lower():
                        emp_code = str(row[key] or "").strip()
                        break
                
                if not emp_code:
                    errors.append({"row": idx, "error": "Missing Employee Code"})
                    continue
                
                # Find employee
                employee = await db.employees.find_one({"emp_code": emp_code}, {"_id": 0})
                if not employee:
                    errors.append({"row": idx, "error": f"Employee {emp_code} not found"})
                    continue
                
                # Parse fields
                def get_field(row, *names):
                    for name in names:
                        for key in row.keys():
                            if key and name.lower() in key.lower():
                                return row.get(key)
                    return None
                
                # Parse ESIC field
                esic_val = get_field(row, "esic")
                esic = False
                if esic_val:
                    esic_str = str(esic_val).strip().lower()
                    esic = esic_str in ['yes', 'y', 'true', '1']
                
                insurance_date = get_field(row, "date")
                amount = get_field(row, "amount")
                company = get_field(row, "company", "insurance company")
                
                # If ESIC is No, then date, amount, and company are required
                if not esic and (not insurance_date or not amount or not company):
                    errors.append({"row": idx, "error": "Missing required fields (Date, Amount, or Insurance Company) - required when ESIC is No"})
                    continue
                
                # Convert date if needed
                if insurance_date:
                    if isinstance(insurance_date, datetime):
                        insurance_date = insurance_date.strftime("%Y-%m-%d")
                    else:
                        insurance_date = str(insurance_date)
                
                # Parse accidental insurance field
                accidental_val = get_field(row, "accidental")
                accidental_insurance = False
                if accidental_val:
                    accidental_str = str(accidental_val).strip().lower()
                    accidental_insurance = accidental_str in ['yes', 'y', 'true', '1']
                
                insurance_doc = {
                    "insurance_id": f"ins_{uuid.uuid4().hex[:12]}",
                    "employee_id": employee["employee_id"],
                    "emp_code": emp_code,
                    "employee_name": f"{employee.get('first_name', '')} {employee.get('last_name', '')}".strip(),
                    "esic": esic,
                    "insurance_date": insurance_date if not esic else None,
                    "amount": float(amount) if amount and not esic else None,
                    "insurance_company": str(company or "") if not esic else None,
                    "policy_number": str(get_field(row, "policy") or "") if not esic else None,
                    "coverage_type": str(get_field(row, "coverage", "type") or "health") if not esic else None,
                    "accidental_insurance": accidental_insurance,
                    "start_date": str(get_field(row, "start") or "") if get_field(row, "start") else None,
                    "end_date": str(get_field(row, "end") or "") if get_field(row, "end") else None,
                    "notes": str(get_field(row, "notes", "note") or ""),
                    "status": "active",
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "created_by": user["user_id"]
                }
                
                await db.insurance.insert_one(insurance_doc)
                imported += 1
                
            except Exception as e:
                errors.append({"row": idx, "error": str(e)})
        
        return {
            "message": "Insurance import completed",
            "imported": imported,
            "errors": errors,
            "total_rows": len(rows)
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")


# ==================== BUSINESS INSURANCE IMPORT ====================

@router.get("/templates/business-insurance")
async def download_business_insurance_template(request: Request):
    """Download business insurance import template as Excel"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "hr_executive"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    try:
        import xlsxwriter
        
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet('Business Insurance')
        
        # Header formats
        header_format = workbook.add_format({
            'bold': True, 'bg_color': '#4472C4', 'font_color': 'white',
            'border': 1, 'align': 'center'
        })
        required_format = workbook.add_format({
            'bold': True, 'bg_color': '#C00000', 'font_color': 'white',
            'border': 1, 'align': 'center'
        })
        note_format = workbook.add_format({'italic': True, 'font_color': 'gray'})
        
        # Headers matching user's template format
        headers = [
            ("Sl.No.", header_format),
            ("Name of insurance*", required_format),
            ("Vehicle no. if required", header_format),
            ("Name of insurance company*", required_format),
            ("Date of issuance", header_format),
            ("Due Date", header_format),
        ]
        
        for col, (header, fmt) in enumerate(headers):
            worksheet.write(0, col, header, fmt)
            worksheet.set_column(col, col, 25)
        
        # Note row
        worksheet.write(1, 0, "* = Required fields. Vehicle No. is optional. Dates in YYYY-MM-DD or DD/MM/YYYY format.", note_format)
        
        # Sample rows
        sample_data = [
            (1, "Fire Insurance", "", "New India Assurance", "2024-01-15", "2025-01-15"),
            (2, "Vehicle Insurance", "MH01AB1234", "ICICI Lombard", "2024-02-01", "2025-02-01"),
            (3, "Machinery Insurance", "", "HDFC Ergo", "2024-03-01", "2025-03-01"),
        ]
        
        for row_idx, data in enumerate(sample_data, start=2):
            for col, val in enumerate(data):
                worksheet.write(row_idx, col, val)
        
        workbook.close()
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=business_insurance_template.xlsx",
                "Cache-Control": "no-cache, no-store, must-revalidate"
            }
        )
    except ImportError:
        raise HTTPException(status_code=500, detail="Excel library not available")


@router.post("/business-insurance")
async def import_business_insurance(request: Request, file: UploadFile = File(...)):
    """Bulk import business insurance records from Excel"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "hr_executive"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    filename = file.filename.lower()
    content = await file.read()
    
    try:
        if filename.endswith('.xlsx'):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content))
            ws = wb.active
            
            # Find header row - look for "Name of insurance" in first few rows
            header_row = None
            headers = []
            for row_num in range(1, min(6, ws.max_row + 1)):
                for col in range(1, ws.max_column + 1):
                    cell_val = str(ws.cell(row=row_num, column=col).value or "").strip().lower()
                    if "name of insurance" in cell_val or "name_of_insurance" in cell_val:
                        header_row = row_num
                        headers = [(ws.cell(row=row_num, column=c).value or "").strip().replace('*', '') for c in range(1, ws.max_column + 1)]
                        break
                if header_row:
                    break
            
            if not header_row:
                # Try first row as default header
                header_row = 1
                headers = [(ws.cell(row=1, column=col).value or "").strip().replace('*', '') for col in range(1, ws.max_column + 1)]
            
            rows = []
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                # Skip note rows and empty rows
                first_val = str(row[0] or "").strip() if row[0] else ""
                second_val = str(row[1] or "").strip() if len(row) > 1 and row[1] else ""
                if (first_val.startswith("*") or first_val.lower().startswith("required") or 
                    second_val.startswith("*") or second_val.lower().startswith("required")):
                    continue
                if not any(row):
                    continue
                rows.append(dict(zip(headers, row)))
        elif filename.endswith('.csv'):
            decoded = content.decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(decoded))
            rows = list(reader)
        else:
            raise HTTPException(status_code=400, detail="Only CSV and Excel files are supported")
        
        imported = 0
        errors = []
        
        for idx, row in enumerate(rows, start=header_row + 2 if header_row else 2):
            try:
                # Parse fields flexibly
                def get_field(row, *names):
                    for name in names:
                        for key in row.keys():
                            if key and name.lower() in key.lower():
                                return row.get(key)
                    return None
                
                name_of_insurance = get_field(row, "name of insurance", "name_of_insurance", "insurance name")
                vehicle_no = get_field(row, "vehicle no", "vehicle_no", "vehicle number")
                insurance_company = get_field(row, "name of insurance company", "insurance company", "company")
                date_of_issuance = get_field(row, "date of issuance", "issuance date", "issue date", "start date")
                due_date = get_field(row, "due date", "due_date", "expiry date", "end date")
                
                # Clean up values
                name_of_insurance = str(name_of_insurance or "").strip() if name_of_insurance else None
                insurance_company = str(insurance_company or "").strip() if insurance_company else None
                vehicle_no = str(vehicle_no or "").strip() if vehicle_no else None
                
                if not name_of_insurance or not insurance_company:
                    errors.append({"row": idx, "error": "Missing required fields (Name of Insurance or Insurance Company)"})
                    continue
                
                # Convert dates if needed
                if date_of_issuance:
                    if isinstance(date_of_issuance, datetime):
                        date_of_issuance = date_of_issuance.strftime("%Y-%m-%d")
                    else:
                        date_of_issuance = str(date_of_issuance)
                
                if due_date:
                    if isinstance(due_date, datetime):
                        due_date = due_date.strftime("%Y-%m-%d")
                    else:
                        due_date = str(due_date)
                
                business_insurance_doc = {
                    "business_insurance_id": f"biz_ins_{uuid.uuid4().hex[:12]}",
                    "name_of_insurance": name_of_insurance,
                    "vehicle_no": vehicle_no if vehicle_no else None,
                    "insurance_company": insurance_company,
                    "date_of_issuance": date_of_issuance if date_of_issuance else None,
                    "due_date": due_date if due_date else None,
                    "notes": str(get_field(row, "notes", "note") or ""),
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "created_by": user["user_id"]
                }
                
                await db.business_insurance.insert_one(business_insurance_doc)
                imported += 1
                
            except Exception as e:
                errors.append({"row": idx, "error": str(e)})
        
        return {
            "message": "Business insurance import completed",
            "imported": imported,
            "errors": errors,
            "total_rows": len(rows)
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")


