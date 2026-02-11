"""Surveys & Suggestions API Routes"""
from fastapi import APIRouter, HTTPException, Request
from typing import List, Optional
from datetime import datetime, timezone
import uuid
from motor.motor_asyncio import AsyncIOMotorClient
import os

router = APIRouter(prefix="/helpdesk", tags=["Helpdesk"])

mongo_url = os.environ.get('MONGO_URL', 'mongodb://localhost:27017')
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ.get('DB_NAME', 'test_database')]


async def get_current_user(request: Request) -> dict:
    from server import get_current_user as auth_get_user
    return await auth_get_user(request)


# ==================== ANONYMOUS SUGGESTIONS ====================

@router.get("/suggestions")
async def list_suggestions(
    request: Request,
    status: Optional[str] = None
):
    """List suggestions - HR sees all, employees see their own"""
    user = await get_current_user(request)
    
    query = {}
    
    # Super admin sees all including submitter identity
    # HR sees all but anonymized
    # Employees see only their own
    if user.get("role") not in ["super_admin", "hr_admin", "hr_executive"]:
        query["submitted_by"] = user.get("employee_id")
    
    if status and status != "all":
        query["status"] = status
    
    suggestions = await db.suggestions.find(query, {"_id": 0}).sort("created_at", -1).to_list(200)
    
    # Anonymize for HR (not super_admin)
    if user.get("role") in ["hr_admin", "hr_executive"]:
        for s in suggestions:
            if s.get("is_anonymous"):
                s["submitted_by"] = None
                s["submitter_name"] = "Anonymous"
    
    return suggestions


@router.post("/suggestions")
async def create_suggestion(data: dict, request: Request):
    """Submit a suggestion (can be anonymous)"""
    user = await get_current_user(request)
    
    suggestion = {
        "suggestion_id": f"SUG-{datetime.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:6].upper()}",
        "submitted_by": user.get("employee_id"),
        "submitter_name": user.get("name"),
        "is_anonymous": data.get("is_anonymous", False),
        "category": data.get("category", "general"),
        "title": data.get("title"),
        "description": data.get("description"),
        "status": "submitted",  # submitted, under_review, acknowledged, implemented, rejected
        "hr_response": None,
        "hr_responded_by": None,
        "hr_responded_at": None,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.suggestions.insert_one(suggestion)
    suggestion.pop('_id', None)
    return suggestion


@router.get("/suggestions/{suggestion_id}")
async def get_suggestion(suggestion_id: str, request: Request):
    """Get suggestion details"""
    user = await get_current_user(request)
    
    suggestion = await db.suggestions.find_one({"suggestion_id": suggestion_id}, {"_id": 0})
    if not suggestion:
        raise HTTPException(status_code=404, detail="Suggestion not found")
    
    # Check access
    if user.get("role") not in ["super_admin", "hr_admin", "hr_executive"]:
        if suggestion.get("submitted_by") != user.get("employee_id"):
            raise HTTPException(status_code=403, detail="Not authorized")
    
    # Anonymize for HR (not super_admin)
    if user.get("role") in ["hr_admin", "hr_executive"] and suggestion.get("is_anonymous"):
        suggestion["submitted_by"] = None
        suggestion["submitter_name"] = "Anonymous"
    
    return suggestion


@router.put("/suggestions/{suggestion_id}/respond")
async def respond_to_suggestion(suggestion_id: str, data: dict, request: Request):
    """HR responds to a suggestion (visible to submitter only)"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "hr_executive"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    await db.suggestions.update_one(
        {"suggestion_id": suggestion_id},
        {"$set": {
            "hr_response": data.get("response"),
            "hr_responded_by": user.get("user_id"),
            "hr_responded_by_name": user.get("name"),
            "hr_responded_at": datetime.now(timezone.utc).isoformat(),
            "status": data.get("status", "acknowledged"),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    return {"message": "Response submitted"}


@router.put("/suggestions/{suggestion_id}/status")
async def update_suggestion_status(suggestion_id: str, data: dict, request: Request):
    """Update suggestion status"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "hr_executive"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    await db.suggestions.update_one(
        {"suggestion_id": suggestion_id},
        {"$set": {
            "status": data.get("status"),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    return {"message": "Status updated"}


# ==================== SURVEYS ====================

@router.get("/surveys")
async def list_surveys(
    request: Request,
    status: Optional[str] = None,
    survey_type: Optional[str] = None
):
    """List surveys - HR sees all, employees see surveys assigned to them"""
    user = await get_current_user(request)
    
    if user.get("role") in ["super_admin", "hr_admin", "hr_executive"]:
        # HR sees all surveys
        query = {}
        if status and status != "all":
            query["status"] = status
        if survey_type and survey_type != "all":
            query["survey_type"] = survey_type
        
        surveys = await db.surveys.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
    else:
        # Employees see surveys assigned to them
        employee_id = user.get("employee_id")
        employee = await db.employees.find_one({"employee_id": employee_id}, {"_id": 0})
        
        query = {
            "status": "active",
            "$or": [
                {"target_type": "all"},
                {"target_employees": employee_id},
                {"target_departments": employee.get("department_id") if employee else None},
                {"target_locations": employee.get("location") if employee else None}
            ]
        }
        
        surveys = await db.surveys.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
        
        # Filter out surveys the employee has already completed (if not editable)
        completed_ids = set()
        responses = await db.survey_responses.find(
            {"employee_id": employee_id}, {"survey_id": 1}
        ).to_list(100)
        for r in responses:
            completed_ids.add(r.get("survey_id"))
        
        # Mark surveys as completed or pending
        for s in surveys:
            s["my_status"] = "completed" if s["survey_id"] in completed_ids else "pending"
    
    return surveys


@router.post("/surveys")
async def create_survey(data: dict, request: Request):
    """Create a new survey (HR/Admin only)"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    survey = {
        "survey_id": f"SRV-{datetime.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:6].upper()}",
        "title": data.get("title"),
        "description": data.get("description", ""),
        "survey_type": data.get("survey_type", "custom"),  # poll, text, satisfaction, engagement, colleague_feedback, pulse, custom
        "is_anonymous": data.get("is_anonymous", False),
        "is_mandatory": data.get("is_mandatory", False),
        "allow_edit": data.get("allow_edit", True),  # Can employees edit responses until deadline
        
        # Targeting
        "target_type": data.get("target_type", "all"),  # all, selected, department, location
        "target_employees": data.get("target_employees", []),  # List of employee_ids
        "target_departments": data.get("target_departments", []),  # List of department_ids
        "target_locations": data.get("target_locations", []),  # List of locations
        
        # For colleague feedback
        "feedback_target_type": data.get("feedback_target_type"),  # hr_assigned, employee_choice
        "feedback_targets": data.get("feedback_targets", []),  # List of employee_ids to rate
        
        # Questions
        "questions": data.get("questions", []),
        
        # Scheduling
        "status": data.get("status", "draft"),  # draft, scheduled, active, closed
        "start_date": data.get("start_date"),
        "end_date": data.get("end_date"),
        "is_recurring": data.get("is_recurring", False),
        "recurrence_pattern": data.get("recurrence_pattern"),  # weekly, monthly, quarterly
        
        # Template
        "is_template": data.get("is_template", False),
        "template_name": data.get("template_name"),
        
        # Metadata
        "created_by": user.get("user_id"),
        "created_by_name": user.get("name"),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
        
        # Stats (updated as responses come in)
        "total_recipients": 0,
        "total_responses": 0,
        "response_rate": 0
    }
    
    # Calculate total recipients
    survey["total_recipients"] = await calculate_survey_recipients(survey)
    
    await db.surveys.insert_one(survey)
    survey.pop('_id', None)
    
    # Create notifications for active surveys
    if survey["status"] == "active":
        await create_survey_notifications(survey)
    
    return survey


async def calculate_survey_recipients(survey: dict) -> int:
    """Calculate total number of survey recipients"""
    target_type = survey.get("target_type", "all")
    
    if target_type == "all":
        count = await db.employees.count_documents({"is_active": True})
    elif target_type == "selected":
        count = len(survey.get("target_employees", []))
    elif target_type == "department":
        count = await db.employees.count_documents({
            "is_active": True,
            "department_id": {"$in": survey.get("target_departments", [])}
        })
    elif target_type == "location":
        count = await db.employees.count_documents({
            "is_active": True,
            "location": {"$in": survey.get("target_locations", [])}
        })
    else:
        count = 0
    
    return count


async def create_survey_notifications(survey: dict):
    """Create in-app notifications for survey recipients"""
    target_type = survey.get("target_type", "all")
    
    # Get target employees
    if target_type == "all":
        employees = await db.employees.find({"is_active": True}, {"employee_id": 1}).to_list(1000)
    elif target_type == "selected":
        employee_ids = survey.get("target_employees", [])
        employees = [{"employee_id": eid} for eid in employee_ids]
    elif target_type == "department":
        employees = await db.employees.find({
            "is_active": True,
            "department_id": {"$in": survey.get("target_departments", [])}
        }, {"employee_id": 1}).to_list(1000)
    elif target_type == "location":
        employees = await db.employees.find({
            "is_active": True,
            "location": {"$in": survey.get("target_locations", [])}
        }, {"employee_id": 1}).to_list(1000)
    else:
        employees = []
    
    # Create notifications
    notifications = []
    for emp in employees:
        notifications.append({
            "notification_id": f"notif_{uuid.uuid4().hex[:12]}",
            "employee_id": emp["employee_id"],
            "type": "survey",
            "title": f"New Survey: {survey.get('title')}",
            "message": survey.get("description", "Please complete this survey"),
            "link": f"/helpdesk?survey={survey['survey_id']}",
            "is_read": False,
            "created_at": datetime.now(timezone.utc).isoformat()
        })
    
    if notifications:
        await db.notifications.insert_many(notifications)


@router.get("/surveys/{survey_id}")
async def get_survey(survey_id: str, request: Request):
    """Get survey details"""
    user = await get_current_user(request)
    
    survey = await db.surveys.find_one({"survey_id": survey_id}, {"_id": 0})
    if not survey:
        raise HTTPException(status_code=404, detail="Survey not found")
    
    # Check if employee can access this survey
    if user.get("role") not in ["super_admin", "hr_admin", "hr_executive"]:
        employee_id = user.get("employee_id")
        employee = await db.employees.find_one({"employee_id": employee_id}, {"_id": 0})
        
        # Check targeting
        target_type = survey.get("target_type", "all")
        has_access = False
        
        if target_type == "all":
            has_access = True
        elif target_type == "selected" and employee_id in survey.get("target_employees", []):
            has_access = True
        elif target_type == "department" and employee and employee.get("department_id") in survey.get("target_departments", []):
            has_access = True
        elif target_type == "location" and employee and employee.get("location") in survey.get("target_locations", []):
            has_access = True
        
        if not has_access:
            raise HTTPException(status_code=403, detail="Not authorized to view this survey")
        
        # Check if already responded
        existing_response = await db.survey_responses.find_one(
            {"survey_id": survey_id, "employee_id": employee_id}, {"_id": 0}
        )
        survey["my_response"] = existing_response
    
    return survey


@router.put("/surveys/{survey_id}")
async def update_survey(survey_id: str, data: dict, request: Request):
    """Update survey (HR/Admin only)"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Don't allow updating closed surveys
    survey = await db.surveys.find_one({"survey_id": survey_id})
    if survey and survey.get("status") == "closed":
        raise HTTPException(status_code=400, detail="Cannot update a closed survey")
    
    update_data = {k: v for k, v in data.items() if k not in ["survey_id", "created_at", "created_by"]}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    # Recalculate recipients if targeting changed
    if "target_type" in data or "target_employees" in data or "target_departments" in data or "target_locations" in data:
        merged_survey = {**survey, **update_data}
        update_data["total_recipients"] = await calculate_survey_recipients(merged_survey)
    
    await db.surveys.update_one({"survey_id": survey_id}, {"$set": update_data})
    
    # Create notifications if survey just became active
    if data.get("status") == "active" and survey.get("status") != "active":
        merged_survey = {**survey, **update_data}
        await create_survey_notifications(merged_survey)
    
    return await db.surveys.find_one({"survey_id": survey_id}, {"_id": 0})


@router.delete("/surveys/{survey_id}")
async def delete_survey(survey_id: str, request: Request):
    """Delete a survey (HR/Admin only)"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Delete survey and all responses
    await db.survey_responses.delete_many({"survey_id": survey_id})
    await db.surveys.delete_one({"survey_id": survey_id})
    
    return {"message": "Survey deleted"}


@router.post("/surveys/{survey_id}/activate")
async def activate_survey(survey_id: str, request: Request):
    """Activate a draft/scheduled survey"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    survey = await db.surveys.find_one({"survey_id": survey_id})
    if not survey:
        raise HTTPException(status_code=404, detail="Survey not found")
    
    await db.surveys.update_one(
        {"survey_id": survey_id},
        {"$set": {
            "status": "active",
            "activated_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Create notifications
    await create_survey_notifications(survey)
    
    return {"message": "Survey activated"}


@router.post("/surveys/{survey_id}/close")
async def close_survey(survey_id: str, request: Request):
    """Close an active survey"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    await db.surveys.update_one(
        {"survey_id": survey_id},
        {"$set": {
            "status": "closed",
            "closed_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "Survey closed"}


@router.post("/surveys/{survey_id}/duplicate")
async def duplicate_survey(survey_id: str, request: Request):
    """Duplicate a survey as a new draft"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    original = await db.surveys.find_one({"survey_id": survey_id}, {"_id": 0})
    if not original:
        raise HTTPException(status_code=404, detail="Survey not found")
    
    # Create copy
    new_survey = {**original}
    new_survey["survey_id"] = f"SRV-{datetime.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:6].upper()}"
    new_survey["title"] = f"Copy of {original.get('title', 'Survey')}"
    new_survey["status"] = "draft"
    new_survey["total_responses"] = 0
    new_survey["response_rate"] = 0
    new_survey["created_by"] = user.get("user_id")
    new_survey["created_by_name"] = user.get("name")
    new_survey["created_at"] = datetime.now(timezone.utc).isoformat()
    new_survey["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.surveys.insert_one(new_survey)
    new_survey.pop('_id', None)
    
    return new_survey


# ==================== SURVEY RESPONSES ====================

@router.post("/surveys/{survey_id}/respond")
async def submit_survey_response(survey_id: str, data: dict, request: Request):
    """Submit response to a survey"""
    user = await get_current_user(request)
    employee_id = user.get("employee_id")
    
    # Get survey
    survey = await db.surveys.find_one({"survey_id": survey_id}, {"_id": 0})
    if not survey:
        raise HTTPException(status_code=404, detail="Survey not found")
    
    if survey.get("status") != "active":
        raise HTTPException(status_code=400, detail="Survey is not active")
    
    # Check deadline
    if survey.get("end_date"):
        end_date = datetime.fromisoformat(survey["end_date"].replace("Z", "+00:00"))
        if datetime.now(timezone.utc) > end_date:
            raise HTTPException(status_code=400, detail="Survey deadline has passed")
    
    # Check if already responded
    existing = await db.survey_responses.find_one(
        {"survey_id": survey_id, "employee_id": employee_id}
    )
    
    if existing and not survey.get("allow_edit", True):
        raise HTTPException(status_code=400, detail="You have already responded to this survey")
    
    response_data = {
        "response_id": existing.get("response_id") if existing else f"RSP-{uuid.uuid4().hex[:12]}",
        "survey_id": survey_id,
        "employee_id": employee_id,
        "employee_name": user.get("name") if not survey.get("is_anonymous") else None,
        "answers": data.get("answers", []),  # [{question_id, answer, rating, selected_options}]
        "feedback_target_id": data.get("feedback_target_id"),  # For colleague feedback
        "submitted_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    if existing:
        # Update existing response
        await db.survey_responses.update_one(
            {"response_id": existing["response_id"]},
            {"$set": response_data}
        )
    else:
        # Insert new response
        await db.survey_responses.insert_one(response_data)
        
        # Update survey stats
        total_responses = await db.survey_responses.count_documents({"survey_id": survey_id})
        response_rate = round((total_responses / survey.get("total_recipients", 1)) * 100, 1) if survey.get("total_recipients") else 0
        
        await db.surveys.update_one(
            {"survey_id": survey_id},
            {"$set": {
                "total_responses": total_responses,
                "response_rate": response_rate
            }}
        )
    
    response_data.pop('_id', None)
    return response_data


@router.get("/surveys/{survey_id}/responses")
async def get_survey_responses(survey_id: str, request: Request):
    """Get all responses for a survey (HR/Admin only)"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "hr_executive"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    survey = await db.surveys.find_one({"survey_id": survey_id}, {"_id": 0})
    if not survey:
        raise HTTPException(status_code=404, detail="Survey not found")
    
    responses = await db.survey_responses.find(
        {"survey_id": survey_id}, {"_id": 0}
    ).to_list(1000)
    
    # Anonymize if survey is anonymous
    if survey.get("is_anonymous"):
        for r in responses:
            r["employee_id"] = None
            r["employee_name"] = "Anonymous"
    
    return responses


@router.get("/surveys/{survey_id}/analytics")
async def get_survey_analytics(survey_id: str, request: Request):
    """Get analytics for a survey (HR/Admin only)"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "hr_executive"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    survey = await db.surveys.find_one({"survey_id": survey_id}, {"_id": 0})
    if not survey:
        raise HTTPException(status_code=404, detail="Survey not found")
    
    responses = await db.survey_responses.find(
        {"survey_id": survey_id}, {"_id": 0}
    ).to_list(1000)
    
    # Calculate analytics per question
    questions = survey.get("questions", [])
    question_analytics = []
    
    for q in questions:
        q_id = q.get("question_id")
        q_type = q.get("type")
        q_analytics = {
            "question_id": q_id,
            "question_text": q.get("text"),
            "type": q_type,
            "total_responses": 0,
            "analytics": {}
        }
        
        # Gather all answers for this question
        answers = []
        for r in responses:
            for a in r.get("answers", []):
                if a.get("question_id") == q_id:
                    answers.append(a)
        
        q_analytics["total_responses"] = len(answers)
        
        if q_type in ["multiple_choice", "single_choice", "yes_no"]:
            # Count options
            option_counts = {}
            for a in answers:
                selected = a.get("selected_options", [])
                if isinstance(selected, str):
                    selected = [selected]
                for opt in selected:
                    option_counts[opt] = option_counts.get(opt, 0) + 1
            
            q_analytics["analytics"] = {
                "option_counts": option_counts,
                "total": len(answers)
            }
        
        elif q_type in ["rating", "nps", "satisfaction"]:
            # Calculate average, distribution
            ratings = [a.get("rating", 0) for a in answers if a.get("rating") is not None]
            if ratings:
                q_analytics["analytics"] = {
                    "average": round(sum(ratings) / len(ratings), 2),
                    "min": min(ratings),
                    "max": max(ratings),
                    "distribution": {str(i): ratings.count(i) for i in set(ratings)}
                }
        
        elif q_type in ["text", "long_text"]:
            # Just collect responses
            q_analytics["analytics"] = {
                "responses": [a.get("answer", "") for a in answers if a.get("answer")]
            }
        
        question_analytics.append(q_analytics)
    
    return {
        "survey": survey,
        "summary": {
            "total_recipients": survey.get("total_recipients", 0),
            "total_responses": len(responses),
            "response_rate": round((len(responses) / survey.get("total_recipients", 1)) * 100, 1) if survey.get("total_recipients") else 0,
            "completion_status": "complete" if len(responses) >= survey.get("total_recipients", 0) else "in_progress"
        },
        "question_analytics": question_analytics
    }


# ==================== SURVEY TEMPLATES ====================

@router.get("/survey-templates")
async def list_survey_templates(request: Request):
    """List available survey templates"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "hr_executive"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get saved templates
    saved_templates = await db.surveys.find(
        {"is_template": True}, {"_id": 0}
    ).to_list(50)
    
    # Built-in templates
    builtin_templates = [
        {
            "template_id": "builtin_satisfaction",
            "template_name": "Employee Satisfaction Survey",
            "description": "Measure overall employee satisfaction with workplace",
            "survey_type": "satisfaction",
            "questions": [
                {"question_id": "q1", "type": "rating", "text": "How satisfied are you with your job overall?", "scale": 5},
                {"question_id": "q2", "type": "rating", "text": "How satisfied are you with your work-life balance?", "scale": 5},
                {"question_id": "q3", "type": "rating", "text": "How satisfied are you with your compensation?", "scale": 5},
                {"question_id": "q4", "type": "rating", "text": "How likely are you to recommend this company as a great place to work?", "scale": 10},
                {"question_id": "q5", "type": "long_text", "text": "What could we do to improve your experience?"}
            ]
        },
        {
            "template_id": "builtin_engagement",
            "template_name": "Employee Engagement Survey",
            "description": "Assess employee engagement and motivation",
            "survey_type": "engagement",
            "questions": [
                {"question_id": "q1", "type": "rating", "text": "I feel motivated to go beyond my basic job requirements", "scale": 5},
                {"question_id": "q2", "type": "rating", "text": "I understand how my work contributes to company goals", "scale": 5},
                {"question_id": "q3", "type": "rating", "text": "I have the tools and resources I need to do my job well", "scale": 5},
                {"question_id": "q4", "type": "rating", "text": "My manager supports my professional development", "scale": 5},
                {"question_id": "q5", "type": "single_choice", "text": "How often do you receive recognition for your work?", "options": ["Daily", "Weekly", "Monthly", "Rarely", "Never"]}
            ]
        },
        {
            "template_id": "builtin_pulse",
            "template_name": "Weekly Pulse Check",
            "description": "Quick weekly check-in with employees",
            "survey_type": "pulse",
            "questions": [
                {"question_id": "q1", "type": "rating", "text": "How would you rate your week overall?", "scale": 5},
                {"question_id": "q2", "type": "yes_no", "text": "Did you face any blockers this week?"},
                {"question_id": "q3", "type": "text", "text": "Anything you'd like to share with HR?"}
            ]
        },
        {
            "template_id": "builtin_onboarding",
            "template_name": "New Employee Onboarding Survey",
            "description": "Gather feedback from new hires about their onboarding experience",
            "survey_type": "custom",
            "questions": [
                {"question_id": "q1", "type": "rating", "text": "How would you rate your overall onboarding experience?", "scale": 5},
                {"question_id": "q2", "type": "rating", "text": "I received adequate training to do my job", "scale": 5},
                {"question_id": "q3", "type": "rating", "text": "My team made me feel welcome", "scale": 5},
                {"question_id": "q4", "type": "single_choice", "text": "Was your workspace ready on your first day?", "options": ["Yes, fully ready", "Partially ready", "Not ready"]},
                {"question_id": "q5", "type": "long_text", "text": "What could we improve about the onboarding process?"}
            ]
        },
        {
            "template_id": "builtin_exit",
            "template_name": "Exit Interview Survey",
            "description": "Understand why employees are leaving",
            "survey_type": "custom",
            "questions": [
                {"question_id": "q1", "type": "single_choice", "text": "Primary reason for leaving?", "options": ["Better opportunity", "Compensation", "Work-life balance", "Management", "Career growth", "Relocation", "Personal reasons", "Other"]},
                {"question_id": "q2", "type": "rating", "text": "How would you rate your overall experience working here?", "scale": 5},
                {"question_id": "q3", "type": "yes_no", "text": "Would you consider working here again in the future?"},
                {"question_id": "q4", "type": "yes_no", "text": "Would you recommend this company to others?"},
                {"question_id": "q5", "type": "long_text", "text": "Any feedback or suggestions for improvement?"}
            ]
        },
        {
            "template_id": "builtin_360",
            "template_name": "360 Degree Feedback",
            "description": "Colleague feedback survey for performance reviews",
            "survey_type": "colleague_feedback",
            "questions": [
                {"question_id": "q1", "type": "rating", "text": "How effectively does this person communicate?", "scale": 5},
                {"question_id": "q2", "type": "rating", "text": "How well does this person collaborate with others?", "scale": 5},
                {"question_id": "q3", "type": "rating", "text": "How reliable is this person in meeting deadlines?", "scale": 5},
                {"question_id": "q4", "type": "rating", "text": "How would you rate this person's problem-solving skills?", "scale": 5},
                {"question_id": "q5", "type": "long_text", "text": "What are this person's strengths?"},
                {"question_id": "q6", "type": "long_text", "text": "What areas could this person improve?"}
            ]
        }
    ]
    
    return {
        "builtin_templates": builtin_templates,
        "saved_templates": saved_templates
    }


@router.post("/survey-templates")
async def save_survey_as_template(data: dict, request: Request):
    """Save a survey as a reusable template"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    template = {
        "survey_id": f"TMPL-{uuid.uuid4().hex[:8]}",
        "is_template": True,
        "template_name": data.get("template_name"),
        "title": data.get("title"),
        "description": data.get("description"),
        "survey_type": data.get("survey_type", "custom"),
        "questions": data.get("questions", []),
        "created_by": user.get("user_id"),
        "created_by_name": user.get("name"),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.surveys.insert_one(template)
    template.pop('_id', None)
    return template


# ==================== NOTIFICATIONS ====================

@router.get("/notifications")
async def get_my_notifications(request: Request, unread_only: bool = False):
    """Get notifications for current user"""
    user = await get_current_user(request)
    employee_id = user.get("employee_id")
    
    query = {"employee_id": employee_id}
    if unread_only:
        query["is_read"] = False
    
    notifications = await db.notifications.find(
        query, {"_id": 0}
    ).sort("created_at", -1).to_list(50)
    
    return notifications


@router.put("/notifications/{notification_id}/read")
async def mark_notification_read(notification_id: str, request: Request):
    """Mark notification as read"""
    user = await get_current_user(request)
    
    await db.notifications.update_one(
        {"notification_id": notification_id, "employee_id": user.get("employee_id")},
        {"$set": {"is_read": True, "read_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"message": "Marked as read"}


@router.put("/notifications/mark-all-read")
async def mark_all_notifications_read(request: Request):
    """Mark all notifications as read"""
    user = await get_current_user(request)
    
    await db.notifications.update_many(
        {"employee_id": user.get("employee_id"), "is_read": False},
        {"$set": {"is_read": True, "read_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"message": "All notifications marked as read"}


# ==================== HELPER ENDPOINTS ====================

@router.get("/departments")
async def list_departments_for_targeting(request: Request):
    """Get departments for survey targeting"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "hr_executive"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    departments = await db.departments.find({}, {"_id": 0}).to_list(100)
    return departments


@router.get("/locations")
async def list_locations_for_targeting(request: Request):
    """Get unique locations for survey targeting"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "hr_executive"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get unique locations from employees
    pipeline = [
        {"$match": {"is_active": True, "location": {"$ne": None}}},
        {"$group": {"_id": "$location"}},
        {"$sort": {"_id": 1}}
    ]
    
    result = await db.employees.aggregate(pipeline).to_list(100)
    locations = [{"location": r["_id"]} for r in result if r["_id"]]
    
    return locations


@router.get("/employees-for-selection")
async def list_employees_for_selection(
    request: Request,
    department_id: Optional[str] = None,
    location: Optional[str] = None,
    search: Optional[str] = None
):
    """Get employees for survey targeting selection"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "hr_executive"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    query = {"is_active": True}
    
    if department_id:
        query["department_id"] = department_id
    if location:
        query["location"] = location
    if search:
        query["$or"] = [
            {"first_name": {"$regex": search, "$options": "i"}},
            {"last_name": {"$regex": search, "$options": "i"}},
            {"emp_code": {"$regex": search, "$options": "i"}}
        ]
    
    employees = await db.employees.find(
        query,
        {"_id": 0, "employee_id": 1, "emp_code": 1, "first_name": 1, "last_name": 1, "department_name": 1, "location": 1}
    ).to_list(500)
    
    return employees


# ==================== ENHANCED SURVEY ANALYTICS ====================

@router.get("/surveys/{survey_id}/analytics/detailed")
async def get_detailed_survey_analytics(survey_id: str, request: Request):
    """Get detailed analytics with department breakdown and timeline"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "hr_executive"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    survey = await db.surveys.find_one({"survey_id": survey_id}, {"_id": 0})
    if not survey:
        raise HTTPException(status_code=404, detail="Survey not found")
    
    responses = await db.survey_responses.find({"survey_id": survey_id}, {"_id": 0}).to_list(1000)
    
    # Build respondent info map
    respondent_ids = [r.get("employee_id") for r in responses]
    employees = await db.employees.find(
        {"$or": [{"employee_id": {"$in": respondent_ids}}, {"emp_code": {"$in": respondent_ids}}]},
        {"_id": 0, "employee_id": 1, "emp_code": 1, "first_name": 1, "last_name": 1, "department": 1, "department_name": 1}
    ).to_list(500)
    emp_map = {}
    for e in employees:
        emp_map[e.get("employee_id")] = e
        if e.get("emp_code"):
            emp_map[e.get("emp_code")] = e
    
    # Department breakdown
    dept_responses = {}
    for r in responses:
        emp = emp_map.get(r.get("employee_id"), {})
        dept = emp.get("department_name") or emp.get("department") or "Unknown"
        if dept not in dept_responses:
            dept_responses[dept] = {"count": 0, "respondents": []}
        dept_responses[dept]["count"] += 1
        dept_responses[dept]["respondents"].append(emp.get("first_name", ""))
    
    # Timeline: responses per day
    timeline = {}
    for r in responses:
        submitted = r.get("submitted_at", r.get("created_at", ""))
        if submitted:
            day = submitted[:10]
            timeline[day] = timeline.get(day, 0) + 1
    
    # Question analytics (enhanced)
    questions = survey.get("questions", [])
    question_analytics = []
    for q in questions:
        q_id = q.get("question_id")
        q_type = q.get("type")
        qa = {"question_id": q_id, "question_text": q.get("text"), "type": q_type, "total_responses": 0, "analytics": {}, "dept_breakdown": {}}
        
        answers = []
        answers_by_dept = {}
        for r in responses:
            emp = emp_map.get(r.get("employee_id"), {})
            dept = emp.get("department_name") or emp.get("department") or "Unknown"
            for a in r.get("answers", []):
                if a.get("question_id") == q_id:
                    answers.append(a)
                    if dept not in answers_by_dept:
                        answers_by_dept[dept] = []
                    answers_by_dept[dept].append(a)
        
        qa["total_responses"] = len(answers)
        
        if q_type in ["rating", "nps", "satisfaction"]:
            ratings = [a.get("rating", 0) for a in answers if a.get("rating") is not None]
            if ratings:
                qa["analytics"] = {
                    "average": round(sum(ratings) / len(ratings), 2),
                    "min": min(ratings),
                    "max": max(ratings),
                    "distribution": {str(i): ratings.count(i) for i in sorted(set(ratings))}
                }
                for dept, dept_ans in answers_by_dept.items():
                    dept_ratings = [a.get("rating", 0) for a in dept_ans if a.get("rating") is not None]
                    if dept_ratings:
                        qa["dept_breakdown"][dept] = round(sum(dept_ratings) / len(dept_ratings), 2)
        
        elif q_type in ["multiple_choice", "single_choice", "yes_no"]:
            option_counts = {}
            for a in answers:
                selected = a.get("selected_options", [])
                if isinstance(selected, str):
                    selected = [selected]
                for opt in selected:
                    option_counts[opt] = option_counts.get(opt, 0) + 1
            qa["analytics"] = {"option_counts": option_counts, "total": len(answers)}
        
        elif q_type in ["text", "long_text"]:
            text_responses = [a.get("answer", "") for a in answers if a.get("answer")]
            qa["analytics"] = {"responses": text_responses[:50], "total_text": len(text_responses)}
        
        question_analytics.append(qa)
    
    # Overall score (average of all rating questions)
    all_averages = [qa["analytics"].get("average", 0) for qa in question_analytics if qa["type"] in ["rating", "nps", "satisfaction"] and qa["analytics"].get("average")]
    overall_score = round(sum(all_averages) / len(all_averages), 2) if all_averages else None
    
    return {
        "survey": survey,
        "summary": {
            "total_recipients": survey.get("total_recipients", 0),
            "total_responses": len(responses),
            "response_rate": round((len(responses) / max(survey.get("total_recipients", 1), 1)) * 100, 1),
            "overall_score": overall_score,
            "status": survey.get("status"),
            "created": survey.get("created_at"),
            "started": survey.get("activated_at"),
            "closed": survey.get("closed_at"),
        },
        "department_breakdown": dept_responses,
        "response_timeline": dict(sorted(timeline.items())),
        "question_analytics": question_analytics
    }


@router.get("/surveys/{survey_id}/export")
async def export_survey_responses(survey_id: str, request: Request):
    """Export survey responses to Excel"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "hr_executive"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    survey = await db.surveys.find_one({"survey_id": survey_id}, {"_id": 0})
    if not survey:
        raise HTTPException(status_code=404, detail="Survey not found")
    
    responses = await db.survey_responses.find({"survey_id": survey_id}, {"_id": 0}).to_list(1000)
    
    # Get employee names
    resp_ids = [r.get("employee_id") for r in responses]
    employees = await db.employees.find(
        {"$or": [{"employee_id": {"$in": resp_ids}}, {"emp_code": {"$in": resp_ids}}]},
        {"_id": 0, "employee_id": 1, "emp_code": 1, "first_name": 1, "last_name": 1, "department": 1}
    ).to_list(500)
    emp_map = {e["employee_id"]: e for e in employees}
    for e in employees:
        if e.get("emp_code"):
            emp_map[e["emp_code"]] = e
    
    import openpyxl
    import io
    from fastapi.responses import StreamingResponse
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Survey Responses"
    
    questions = survey.get("questions", [])
    headers = ["Employee", "Department", "Submitted At"] + [q.get("text", f"Q{i+1}") for i, q in enumerate(questions)]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = openpyxl.styles.Font(bold=True)
    
    for row_idx, resp in enumerate(responses, 2):
        emp = emp_map.get(resp.get("employee_id"), {})
        name = f"{emp.get('first_name', '')} {emp.get('last_name', '')}".strip() or resp.get("employee_id")
        ws.cell(row=row_idx, column=1, value=name)
        ws.cell(row=row_idx, column=2, value=emp.get("department", ""))
        ws.cell(row=row_idx, column=3, value=resp.get("submitted_at", ""))
        
        answer_map = {a["question_id"]: a for a in resp.get("answers", [])}
        for col_idx, q in enumerate(questions, 4):
            ans = answer_map.get(q["question_id"], {})
            if q["type"] in ["rating", "nps", "satisfaction"]:
                ws.cell(row=row_idx, column=col_idx, value=ans.get("rating"))
            elif q["type"] in ["text", "long_text"]:
                ws.cell(row=row_idx, column=col_idx, value=ans.get("answer"))
            elif q["type"] in ["single_choice", "multiple_choice", "yes_no"]:
                selected = ans.get("selected_options", [])
                if isinstance(selected, list):
                    ws.cell(row=row_idx, column=col_idx, value=", ".join(selected))
                else:
                    ws.cell(row=row_idx, column=col_idx, value=str(selected))
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=survey_{survey_id}_responses.xlsx"}
    )


# ==================== 360-DEGREE FEEDBACK ====================

@router.post("/feedback-cycles")
async def create_feedback_cycle(data: dict, request: Request):
    """Create a new 360-degree feedback cycle"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "hr_executive"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    cycle = {
        "cycle_id": f"fc_{uuid.uuid4().hex[:12]}",
        "title": data.get("title", ""),
        "description": data.get("description", ""),
        "status": "draft",
        "start_date": data.get("start_date"),
        "end_date": data.get("end_date"),
        "questions": data.get("questions", [
            {"question_id": "fb_q1", "text": "How effective is this person at communication?", "type": "rating", "category": "Communication"},
            {"question_id": "fb_q2", "text": "How well does this person collaborate with the team?", "type": "rating", "category": "Teamwork"},
            {"question_id": "fb_q3", "text": "How would you rate their leadership abilities?", "type": "rating", "category": "Leadership"},
            {"question_id": "fb_q4", "text": "How reliable is this person in meeting deadlines?", "type": "rating", "category": "Reliability"},
            {"question_id": "fb_q5", "text": "How well does this person handle challenges?", "type": "rating", "category": "Problem Solving"},
            {"question_id": "fb_q6", "text": "What are this person's key strengths?", "type": "long_text", "category": "Strengths"},
            {"question_id": "fb_q7", "text": "What areas could this person improve on?", "type": "long_text", "category": "Areas for Improvement"},
        ]),
        "allow_self_nomination": data.get("allow_self_nomination", True),
        "min_reviewers": data.get("min_reviewers", 3),
        "anonymous": data.get("anonymous", True),
        "assignments": [],
        "created_by": user.get("user_id"),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    
    await db.feedback_cycles.insert_one(cycle)
    cycle.pop("_id", None)
    return cycle


@router.get("/feedback-cycles")
async def list_feedback_cycles(request: Request):
    """List feedback cycles"""
    user = await get_current_user(request)
    is_hr = user.get("role") in ["super_admin", "hr_admin", "hr_executive"]
    
    cycles = await db.feedback_cycles.find({}, {"_id": 0}).sort("created_at", -1).to_list(50)
    
    emp_id = user.get("employee_id")
    for cycle in cycles:
        if is_hr:
            total_assignments = len(cycle.get("assignments", []))
            completed = sum(1 for a in cycle.get("assignments", []) if a.get("status") == "completed")
            cycle["total_assignments"] = total_assignments
            cycle["completed_assignments"] = completed
        else:
            my_assignments = [a for a in cycle.get("assignments", []) if a.get("reviewer_id") == emp_id]
            cycle["my_assignments"] = len(my_assignments)
            cycle["my_completed"] = sum(1 for a in my_assignments if a.get("status") == "completed")
            cycle["my_pending"] = sum(1 for a in my_assignments if a.get("status") == "pending")
        
        cycle.pop("assignments", None)
    
    return cycles


@router.get("/feedback-cycles/{cycle_id}")
async def get_feedback_cycle(cycle_id: str, request: Request):
    """Get feedback cycle details"""
    user = await get_current_user(request)
    
    cycle = await db.feedback_cycles.find_one({"cycle_id": cycle_id}, {"_id": 0})
    if not cycle:
        raise HTTPException(status_code=404, detail="Feedback cycle not found")
    
    is_hr = user.get("role") in ["super_admin", "hr_admin", "hr_executive"]
    emp_id = user.get("employee_id")
    
    if not is_hr:
        cycle["my_assignments"] = [a for a in cycle.get("assignments", []) if a.get("reviewer_id") == emp_id]
        cycle.pop("assignments", None)
    
    return cycle


@router.put("/feedback-cycles/{cycle_id}")
async def update_feedback_cycle(cycle_id: str, data: dict, request: Request):
    """Update feedback cycle"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "hr_executive"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    update = {}
    for field in ["title", "description", "start_date", "end_date", "questions", "status", "allow_self_nomination", "min_reviewers", "anonymous"]:
        if field in data:
            update[field] = data[field]
    update["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.feedback_cycles.update_one({"cycle_id": cycle_id}, {"$set": update})
    return await db.feedback_cycles.find_one({"cycle_id": cycle_id}, {"_id": 0})


@router.post("/feedback-cycles/{cycle_id}/assign")
async def assign_reviewers(cycle_id: str, data: dict, request: Request):
    """Assign reviewers to an employee in a feedback cycle"""
    user = await get_current_user(request)
    is_hr = user.get("role") in ["super_admin", "hr_admin", "hr_executive"]
    emp_id = user.get("employee_id")
    
    cycle = await db.feedback_cycles.find_one({"cycle_id": cycle_id})
    if not cycle:
        raise HTTPException(status_code=404, detail="Cycle not found")
    
    target_employee = data.get("target_employee_id")
    reviewer_ids = data.get("reviewer_ids", [])
    
    if not is_hr and not cycle.get("allow_self_nomination"):
        raise HTTPException(status_code=403, detail="Self-nomination not allowed")
    
    if not is_hr:
        target_employee = emp_id
    
    existing = cycle.get("assignments", [])
    
    new_assignments = []
    for rid in reviewer_ids:
        if rid == target_employee:
            continue
        exists = any(a.get("target_employee_id") == target_employee and a.get("reviewer_id") == rid for a in existing)
        if not exists:
            new_assignments.append({
                "assignment_id": f"fa_{uuid.uuid4().hex[:8]}",
                "target_employee_id": target_employee,
                "reviewer_id": rid,
                "status": "pending",
                "created_at": datetime.now(timezone.utc).isoformat()
            })
    
    if new_assignments:
        await db.feedback_cycles.update_one(
            {"cycle_id": cycle_id},
            {"$push": {"assignments": {"$each": new_assignments}}}
        )
    
    return {"message": f"Assigned {len(new_assignments)} reviewers", "assignments_added": len(new_assignments)}


@router.post("/feedback-cycles/{cycle_id}/submit")
async def submit_feedback(cycle_id: str, data: dict, request: Request):
    """Submit feedback for an assigned review"""
    user = await get_current_user(request)
    emp_id = user.get("employee_id")
    
    assignment_id = data.get("assignment_id")
    answers = data.get("answers", [])
    
    cycle = await db.feedback_cycles.find_one({"cycle_id": cycle_id})
    if not cycle:
        raise HTTPException(status_code=404, detail="Cycle not found")
    
    assignment = next((a for a in cycle.get("assignments", []) if a.get("assignment_id") == assignment_id and a.get("reviewer_id") == emp_id), None)
    if not assignment:
        raise HTTPException(status_code=404, detail="Assignment not found")
    
    if assignment.get("status") == "completed":
        raise HTTPException(status_code=400, detail="Already submitted")
    
    feedback = {
        "feedback_id": f"fb_{uuid.uuid4().hex[:12]}",
        "cycle_id": cycle_id,
        "assignment_id": assignment_id,
        "target_employee_id": assignment["target_employee_id"],
        "reviewer_id": emp_id,
        "anonymous": cycle.get("anonymous", True),
        "answers": answers,
        "submitted_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.feedback_responses.insert_one(feedback)
    feedback.pop("_id", None)
    
    await db.feedback_cycles.update_one(
        {"cycle_id": cycle_id, "assignments.assignment_id": assignment_id},
        {"$set": {"assignments.$.status": "completed", "assignments.$.submitted_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return feedback


@router.get("/feedback-cycles/{cycle_id}/my-assignments")
async def get_my_feedback_assignments(cycle_id: str, request: Request):
    """Get my pending feedback assignments with target employee details"""
    user = await get_current_user(request)
    emp_id = user.get("employee_id")
    
    cycle = await db.feedback_cycles.find_one({"cycle_id": cycle_id}, {"_id": 0})
    if not cycle:
        raise HTTPException(status_code=404, detail="Cycle not found")
    
    my_assignments = [a for a in cycle.get("assignments", []) if a.get("reviewer_id") == emp_id]
    
    for a in my_assignments:
        emp = await db.employees.find_one(
            {"$or": [{"employee_id": a["target_employee_id"]}, {"emp_code": a["target_employee_id"]}]},
            {"_id": 0, "first_name": 1, "last_name": 1, "department": 1, "designation": 1}
        )
        if emp:
            a["target_name"] = f"{emp.get('first_name', '')} {emp.get('last_name', '')}".strip()
            a["target_department"] = emp.get("department", "")
            a["target_designation"] = emp.get("designation", "")
    
    return {"cycle": {"cycle_id": cycle["cycle_id"], "title": cycle["title"], "questions": cycle.get("questions", []), "anonymous": cycle.get("anonymous", True)}, "assignments": my_assignments}


@router.get("/feedback-cycles/{cycle_id}/analytics")
async def get_feedback_cycle_analytics(cycle_id: str, request: Request, employee_id: Optional[str] = None):
    """Get analytics for a feedback cycle (HR only) or for a specific employee"""
    user = await get_current_user(request)
    is_hr = user.get("role") in ["super_admin", "hr_admin", "hr_executive"]
    emp_id = user.get("employee_id")
    
    cycle = await db.feedback_cycles.find_one({"cycle_id": cycle_id}, {"_id": 0})
    if not cycle:
        raise HTTPException(status_code=404, detail="Cycle not found")
    
    query = {"cycle_id": cycle_id}
    if employee_id and is_hr:
        query["target_employee_id"] = employee_id
    elif not is_hr:
        query["target_employee_id"] = emp_id
    
    feedbacks = await db.feedback_responses.find(query, {"_id": 0}).to_list(500)
    
    if not is_hr:
        for f in feedbacks:
            if cycle.get("anonymous", True):
                f.pop("reviewer_id", None)
    
    questions = cycle.get("questions", [])
    question_summaries = []
    
    for q in questions:
        q_id = q.get("question_id")
        q_type = q.get("type")
        answers = []
        for f in feedbacks:
            for a in f.get("answers", []):
                if a.get("question_id") == q_id:
                    answers.append(a)
        
        summary = {
            "question_id": q_id,
            "question_text": q.get("text"),
            "category": q.get("category", ""),
            "type": q_type,
            "total_responses": len(answers),
        }
        
        if q_type in ["rating"]:
            ratings = [a.get("rating", 0) for a in answers if a.get("rating") is not None]
            if ratings:
                summary["average"] = round(sum(ratings) / len(ratings), 2)
                summary["distribution"] = {str(i): ratings.count(i) for i in range(1, 6)}
        elif q_type in ["long_text", "text"]:
            summary["responses"] = [a.get("answer", "") for a in answers if a.get("answer")]
        
        question_summaries.append(summary)
    
    # Per-employee summaries (HR only)
    employee_summaries = []
    if is_hr and not employee_id:
        assignments = cycle.get("assignments", [])
        target_ids = list(set(a.get("target_employee_id") for a in assignments))
        
        for tid in target_ids:
            emp = await db.employees.find_one(
                {"$or": [{"employee_id": tid}, {"emp_code": tid}]},
                {"_id": 0, "first_name": 1, "last_name": 1, "department": 1}
            )
            emp_feedbacks = [f for f in feedbacks if f.get("target_employee_id") == tid]
            assigned = sum(1 for a in assignments if a.get("target_employee_id") == tid)
            completed = sum(1 for a in assignments if a.get("target_employee_id") == tid and a.get("status") == "completed")
            
            rating_scores = []
            for f in emp_feedbacks:
                for a in f.get("answers", []):
                    if a.get("rating") is not None:
                        rating_scores.append(a["rating"])
            
            employee_summaries.append({
                "employee_id": tid,
                "name": f"{emp.get('first_name', '')} {emp.get('last_name', '')}".strip() if emp else tid,
                "department": emp.get("department", "") if emp else "",
                "total_assigned": assigned,
                "total_completed": completed,
                "avg_score": round(sum(rating_scores) / len(rating_scores), 2) if rating_scores else None,
                "total_feedbacks": len(emp_feedbacks)
            })
        
        employee_summaries.sort(key=lambda x: x.get("avg_score") or 0, reverse=True)
    
    total_assignments = len(cycle.get("assignments", []))
    completed_assignments = sum(1 for a in cycle.get("assignments", []) if a.get("status") == "completed")
    
    return {
        "cycle": {"cycle_id": cycle["cycle_id"], "title": cycle["title"], "status": cycle.get("status"), "anonymous": cycle.get("anonymous", True)},
        "summary": {
            "total_assignments": total_assignments,
            "completed": completed_assignments,
            "completion_rate": round((completed_assignments / max(total_assignments, 1)) * 100, 1),
        },
        "question_summaries": question_summaries,
        "employee_summaries": employee_summaries
    }


@router.delete("/feedback-cycles/{cycle_id}")
async def delete_feedback_cycle(cycle_id: str, request: Request):
    """Delete a feedback cycle"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin", "hr_executive"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    await db.feedback_cycles.delete_one({"cycle_id": cycle_id})
    await db.feedback_responses.delete_many({"cycle_id": cycle_id})
    return {"message": "Feedback cycle deleted"}


@router.get("/my-feedback-summary")
async def get_my_feedback_summary(request: Request):
    """Get aggregated feedback summary for the current employee"""
    user = await get_current_user(request)
    emp_id = user.get("employee_id")
    
    feedbacks = await db.feedback_responses.find({"target_employee_id": emp_id}, {"_id": 0}).to_list(200)
    
    if not feedbacks:
        return {"has_feedback": False, "message": "No feedback received yet"}
    
    # Aggregate scores by category
    category_scores = {}
    text_feedback = {"strengths": [], "improvements": []}
    
    for f in feedbacks:
        cycle = await db.feedback_cycles.find_one({"cycle_id": f["cycle_id"]}, {"_id": 0, "questions": 1, "anonymous": 1})
        if not cycle:
            continue
        q_map = {q["question_id"]: q for q in cycle.get("questions", [])}
        
        for a in f.get("answers", []):
            q = q_map.get(a.get("question_id"), {})
            cat = q.get("category", "General")
            
            if a.get("rating") is not None:
                if cat not in category_scores:
                    category_scores[cat] = []
                category_scores[cat].append(a["rating"])
            elif a.get("answer") and q.get("type") in ["text", "long_text"]:
                if "strength" in cat.lower():
                    text_feedback["strengths"].append(a["answer"])
                elif "improve" in cat.lower():
                    text_feedback["improvements"].append(a["answer"])
    
    categories = []
    for cat, scores in category_scores.items():
        categories.append({
            "category": cat,
            "average": round(sum(scores) / len(scores), 2),
            "count": len(scores)
        })
    
    overall = sum(s for scores in category_scores.values() for s in scores)
    total = sum(len(scores) for scores in category_scores.values())
    
    return {
        "has_feedback": True,
        "total_feedbacks": len(feedbacks),
        "overall_score": round(overall / total, 2) if total else None,
        "categories": sorted(categories, key=lambda x: x["average"], reverse=True),
        "text_feedback": text_feedback
    }

