"""SOP (Standard Operating Procedures) Management API Routes"""
from fastapi import APIRouter, HTTPException, Request, UploadFile, File
from fastapi.responses import Response
from typing import List, Optional
from datetime import datetime, timezone
import uuid
import base64
import re
import json
import asyncio
from motor.motor_asyncio import AsyncIOMotorClient
import os
from dotenv import load_dotenv

load_dotenv()

router = APIRouter(prefix="/sop", tags=["SOP Management"])

mongo_url = os.environ.get('MONGO_URL', 'mongodb://localhost:27017')
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ.get('DB_NAME', 'test_database')]


async def get_current_user(request: Request) -> dict:
    from server import get_current_user as auth_get_user
    return await auth_get_user(request)


async def match_employee_name(name: str, employees_cache: list) -> Optional[str]:
    """Match a name string to an employee_id"""
    if not name or not name.strip():
        return None
    
    name_clean = name.strip().lower()
    
    # Try exact match first
    for emp in employees_cache:
        full_name = f"{emp.get('first_name', '')} {emp.get('last_name', '')}".strip().lower()
        if full_name == name_clean:
            return emp.get('employee_id')
    
    # Try partial match (first name or last name)
    for emp in employees_cache:
        first = emp.get('first_name', '').lower()
        last = emp.get('last_name', '').lower()
        if first and first in name_clean:
            return emp.get('employee_id')
        if last and last in name_clean:
            return emp.get('employee_id')
    
    return None


async def parse_sop_with_ai(text_content: str, employees_list: list) -> dict:
    """Use AI to intelligently parse SOP content and extract structured data"""
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        
        api_key = os.environ.get('EMERGENT_LLM_KEY')
        if not api_key:
            return {"error": "No API key configured"}
        
        # Create employee name list for context
        employee_names = [f"{e.get('first_name', '')} {e.get('last_name', '')}".strip() for e in employees_list if e.get('first_name')]
        
        system_prompt = """You are an expert at parsing Standard Operating Procedure (SOP) documents from Indian manufacturing companies.
Extract ALL relevant information comprehensively and return it as valid JSON.

The JSON MUST have these fields (use null or empty array if not found):
{
  "sop_number": "The SOP ID/Number (e.g., SDPL/SOP/28, SDPL/SOP/HR/01)",
  "title": "The title or process name (e.g., EMPLOYEE LIFE CYCLE, EXPENSE BUDGETING)",
  "process_owner": "Full name and designation of process owner (e.g., NANDINI KUMARI ASSISTANT MANAGER HUMAN RESOURCE)",
  "created_by": "Name of document creator",
  "department": "Department name (HR, Finance, Production, Quality, etc.)",
  "version": "Version number",
  "revision_date": "Last revision date",
  "purpose": "COMPLETE purpose/objective text - do not truncate, include full description",
  "scope": "COMPLETE scope text - do not truncate",
  "input_requirements": "What inputs are needed to start this process",
  "output_deliverables": "What outputs/deliverables this process produces",
  "procedure_summary": "Brief summary of the overall procedure",
  "responsible_persons": ["Extract ALL roles/persons mentioned as responsible - look for RESP/IN-CHARGE column, include designations like HOD, ASSISTANT MANAGER HR, MD, ACCOUNTS, etc."],
  "reports": ["List of reports or documents mentioned"],
  "task_type": "Category: HR, Finance, Quality, Production, Safety, Audit, Maintenance, Admin, etc.",
  "stakeholders": "Extract from PROCESS STAKEHOLDERS field - include ALL stakeholders mentioned",
  "key_activities": ["Extract main activities/steps from the procedure section"],
  "process_flow_steps": [
    {
      "step_number": 1,
      "step_name": "Short name for the step",
      "description": "Detailed description of what happens in this step",
      "responsible": "Who is responsible (role/person)",
      "input": "Input for this step if mentioned",
      "output": "Output of this step if mentioned"
    }
  ]
}

IMPORTANT EXTRACTION RULES:
1. Look for "PROCESS STAKEHOLDERs" or "STAKEHOLDERS" row - extract ALL names/departments listed
2. Look for "RESP/IN-CHARGE" column in the process flow - extract all responsible parties
3. Look for "PURPOSE OF PROCESS" - copy the FULL text, not abbreviated
4. Look for "SCOPE" - copy the FULL text
5. For process_flow_steps: Look for the STANDARD OPERATING PROCEDURE section with step-by-step procedures
6. Each process step usually has: Step name, Description, and RESP/IN-CHARGE
7. Include ALL stakeholders even if they seem like roles (e.g., "POTENTIAL CANDIDATES", "EMPLOYEES", "ACCOUNTS DEPT")

Return ONLY valid JSON, no other text or explanation."""

        chat = LlmChat(
            api_key=api_key,
            session_id=f"sop_parse_{uuid.uuid4().hex[:8]}",
            system_message=system_prompt
        ).with_model("openai", "gpt-4o-mini")
        
        # Truncate content if too long - increased limit for better extraction
        max_chars = 20000
        if len(text_content) > max_chars:
            text_content = text_content[:max_chars] + "\n...[Content truncated]"
        
        user_message = UserMessage(
            text=f"""Parse this SOP document content and extract ALL information thoroughly.
Pay special attention to:
1. PROCESS STAKEHOLDERS section
2. RESP/IN-CHARGE column for each process step
3. Complete PURPOSE and SCOPE text
4. All process flow steps with their descriptions

SOP CONTENT:
---
{text_content}
---

Available employees for matching: {', '.join(employee_names[:50])}

Return comprehensive JSON with all extracted data."""
        )
        
        response = await chat.send_message(user_message)
        
        # Parse JSON from response
        try:
            # Try to extract JSON from response
            json_match = re.search(r'\{[\s\S]*\}', response)
            if json_match:
                parsed = json.loads(json_match.group())
                return parsed
        except json.JSONDecodeError:
            pass
        
        return {"error": "Failed to parse AI response", "raw_response": response[:500]}
        
    except Exception as e:
        return {"error": str(e)}


async def parse_sop_excel(file_content: bytes, employees_list: list = None) -> dict:
    """Parse SOP Excel file and extract metadata + content using AI"""
    import io
    import openpyxl
    
    result = {
        "sop_number": None,
        "title": None,
        "process_owner": None,
        "created_by": None,
        "department": None,
        "version": None,
        "purpose": None,
        "scope": None,
        "procedure_summary": None,
        "responsible_persons": [],
        "reports": [],
        "task_type": None,
        "stakeholders": [],
        "key_activities": [],
        "preview_data": [],
        "total_rows": 0,
        "total_cols": 0
    }
    
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_content))
        ws = wb.active
        
        result["total_rows"] = ws.max_row
        result["total_cols"] = ws.max_column
        
        # Extract all cell data
        all_text = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            row_data = [str(cell) if cell is not None else "" for cell in row]
            all_text.append(" | ".join([c for c in row_data if c.strip()]))
            
            if row_idx < 100:  # Preview first 100 rows
                result["preview_data"].append(row_data)
        
        # Combine all text for AI parsing
        full_text = "\n".join([t for t in all_text if t.strip()])
        
        # Use AI to parse the content
        if employees_list is None:
            employees_list = []
        
        ai_result = await parse_sop_with_ai(full_text, employees_list)
        
        if "error" not in ai_result:
            # Merge AI results into our result
            for key in ["sop_number", "title", "process_owner", "created_by", "department", 
                       "version", "purpose", "scope", "procedure_summary", "task_type",
                       "input_requirements", "output_deliverables", "revision_date"]:
                if ai_result.get(key):
                    result[key] = ai_result[key]
            
            for key in ["responsible_persons", "reports", "stakeholders", "key_activities", "process_flow_steps"]:
                if ai_result.get(key) and isinstance(ai_result[key], list):
                    result[key] = ai_result[key]
        else:
            result["ai_parse_error"] = ai_result.get("error")
            # Fallback to regex for basic fields
            sop_match = re.search(r'SDPL/SOP/[A-Z]*/?\d+|SOP[-/]\d+', full_text, re.IGNORECASE)
            if sop_match:
                result["sop_number"] = sop_match.group(0)
        
    except Exception as e:
        result["parse_error"] = str(e)
    
    return result


# ==================== SOP CRUD ====================

@router.get("/list")
async def list_sops(
    request: Request,
    department_id: Optional[str] = None,
    designation_id: Optional[str] = None,
    status: Optional[str] = None,
    search: Optional[str] = None,
    group_by: Optional[str] = None,
    owner_id: Optional[str] = None,
    involved_id: Optional[str] = None
):
    """List all SOPs with optional filters, search, and grouping"""
    await get_current_user(request)  # Auth check
    
    query = {"is_active": True}
    
    if department_id and department_id != 'all':
        query["departments"] = department_id
    
    if designation_id:
        query["designations"] = designation_id
    
    if status and status != 'all':
        query["status"] = status
    
    if owner_id:
        query["main_responsible"] = owner_id
    
    if involved_id:
        query["$or"] = [
            {"main_responsible": involved_id},
            {"also_involved": involved_id}
        ]
    
    # Text search across multiple fields
    if search:
        search_regex = {"$regex": search, "$options": "i"}
        query["$or"] = [
            {"title": search_regex},
            {"description": search_regex},
            {"sop_id": search_regex},
            {"sop_number": search_regex},
            {"process_owner": search_regex},
            {"purpose": search_regex},
            {"scope": search_regex},
            {"task_type": search_regex}
        ]
    
    sops = await db.sops.find(query, {"_id": 0, "file_data": 0}).sort("created_at", -1).to_list(500)
    
    # Enrich with department, designation, and employee names
    dept_ids = set()
    desig_ids = set()
    emp_ids = set()
    for sop in sops:
        dept_ids.update(sop.get("departments", []))
        desig_ids.update(sop.get("designations", []))
        emp_ids.update(sop.get("main_responsible", []))
        emp_ids.update(sop.get("also_involved", []))
    
    departments = await db.departments.find({"department_id": {"$in": list(dept_ids)}}, {"_id": 0}).to_list(100)
    designations = await db.designations.find({"designation_id": {"$in": list(desig_ids)}}, {"_id": 0}).to_list(100)
    employees = await db.employees.find(
        {"employee_id": {"$in": list(emp_ids)}},
        {"_id": 0, "employee_id": 1, "first_name": 1, "last_name": 1, "emp_code": 1}
    ).to_list(500)
    
    dept_map = {d["department_id"]: d["name"] for d in departments}
    desig_map = {d["designation_id"]: d["name"] for d in designations}
    emp_map = {e["employee_id"]: f"{e.get('first_name', '')} {e.get('last_name', '')}".strip() or e.get('emp_code', e['employee_id']) for e in employees}
    
    for sop in sops:
        sop["department_names"] = [dept_map.get(d, d) for d in sop.get("departments", [])]
        sop["designation_names"] = [desig_map.get(d, d) for d in sop.get("designations", [])]
        sop["main_responsible_names"] = [emp_map.get(e, e) for e in sop.get("main_responsible", [])]
        sop["also_involved_names"] = [emp_map.get(e, e) for e in sop.get("also_involved", [])]
    
    # Handle grouping
    if group_by:
        grouped = {}
        for sop in sops:
            if group_by == "department":
                keys = sop.get("department_names", ["Unassigned"]) or ["Unassigned"]
            elif group_by == "owner":
                keys = sop.get("main_responsible_names", ["Unassigned"]) or ["Unassigned"]
            elif group_by == "task_type":
                keys = [sop.get("task_type", "Uncategorized") or "Uncategorized"]
            elif group_by == "status":
                keys = [sop.get("status", "draft")]
            else:
                keys = ["All"]
            
            for key in keys:
                if key not in grouped:
                    grouped[key] = []
                grouped[key].append(sop)
        
        return {"grouped": True, "groups": grouped}
    
    return sops


@router.get("/my-sops")
async def get_my_sops(request: Request):
    """Get SOPs applicable to the current user - split by main responsible and also involved"""
    user = await get_current_user(request)
    employee_id = user.get("employee_id")
    
    if not employee_id:
        return {"main_responsible": [], "also_involved": []}
    
    # Get employee's designation and department
    employee = await db.employees.find_one(
        {"employee_id": employee_id},
        {"_id": 0, "designation_id": 1, "department_id": 1}
    )
    
    designation_id = employee.get("designation_id") if employee else None
    department_id = employee.get("department_id") if employee else None
    
    # Find SOPs where user is main responsible
    main_responsible_sops = await db.sops.find(
        {
            "is_active": True,
            "status": "published",
            "main_responsible": employee_id
        },
        {"_id": 0, "file_data": 0}
    ).sort("created_at", -1).to_list(50)
    
    # Find SOPs where user is also involved (directly or via designation/department)
    also_involved_query = {
        "is_active": True,
        "status": "published",
        "main_responsible": {"$ne": employee_id},  # Exclude ones where they're main responsible
        "$or": [
            {"also_involved": employee_id},  # Directly assigned
            {"designations": designation_id} if designation_id else {"_id": {"$exists": False}},
            {"departments": department_id} if department_id else {"_id": {"$exists": False}},
            {"designations": {"$size": 0}, "departments": {"$size": 0}, "also_involved": {"$size": 0}}  # For all
        ]
    }
    
    also_involved_sops = await db.sops.find(
        also_involved_query,
        {"_id": 0, "file_data": 0}
    ).sort("created_at", -1).to_list(50)
    
    return {
        "main_responsible": main_responsible_sops,
        "also_involved": also_involved_sops
    }


@router.get("/by-employee/{employee_id}")
async def get_sops_by_employee(employee_id: str, request: Request):
    """Get all SOPs where an employee is owner or involved"""
    await get_current_user(request)  # Auth check
    
    # SOPs where employee is main responsible (owner)
    owner_sops = await db.sops.find(
        {"is_active": True, "main_responsible": employee_id},
        {"_id": 0, "file_data": 0}
    ).sort("created_at", -1).to_list(100)
    
    # SOPs where employee is involved
    involved_sops = await db.sops.find(
        {"is_active": True, "also_involved": employee_id, "main_responsible": {"$ne": employee_id}},
        {"_id": 0, "file_data": 0}
    ).sort("created_at", -1).to_list(100)
    
    # Get employee's department and designation for indirect involvement
    employee = await db.employees.find_one(
        {"employee_id": employee_id},
        {"_id": 0, "department_id": 1, "designation_id": 1, "first_name": 1, "last_name": 1}
    )
    
    dept_sops = []
    desig_sops = []
    if employee:
        if employee.get("department_id"):
            dept_sops = await db.sops.find(
                {
                    "is_active": True,
                    "departments": employee["department_id"],
                    "main_responsible": {"$ne": employee_id},
                    "also_involved": {"$ne": employee_id}
                },
                {"_id": 0, "file_data": 0}
            ).to_list(50)
        
        if employee.get("designation_id"):
            desig_sops = await db.sops.find(
                {
                    "is_active": True,
                    "designations": employee["designation_id"],
                    "main_responsible": {"$ne": employee_id},
                    "also_involved": {"$ne": employee_id}
                },
                {"_id": 0, "file_data": 0}
            ).to_list(50)
    
    return {
        "employee": employee,
        "as_owner": owner_sops,
        "directly_involved": involved_sops,
        "via_department": dept_sops,
        "via_designation": desig_sops
    }


@router.post("/create")
async def create_sop(request: Request):
    """Create a new SOP with Excel file upload and auto-parsing"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    form = await request.form()
    
    title = form.get("title")
    description = form.get("description", "")
    departments = form.getlist("departments") or []
    designations = form.getlist("designations") or []
    main_responsible = form.getlist("main_responsible") or []
    also_involved = form.getlist("also_involved") or []
    task_type = form.get("task_type", "")
    file = form.get("file")
    
    sop_doc = {
        "sop_id": f"SOP-{uuid.uuid4().hex[:8].upper()}",
        "title": title or "Untitled SOP",
        "description": description,
        "departments": departments if isinstance(departments, list) else [departments] if departments else [],
        "designations": designations if isinstance(designations, list) else [designations] if designations else [],
        "main_responsible": main_responsible if isinstance(main_responsible, list) else [main_responsible] if main_responsible else [],
        "also_involved": also_involved if isinstance(also_involved, list) else [also_involved] if also_involved else [],
        "task_type": task_type,
        "status": "draft",
        "version": 1,
        "is_active": True,
        "created_by": user["user_id"],
        "created_by_name": user.get("name", user.get("email")),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Handle file upload and parse
    if file:
        file_content = await file.read()
        sop_doc["file_name"] = file.filename
        sop_doc["file_data"] = base64.b64encode(file_content).decode('utf-8')
        sop_doc["file_type"] = file.content_type
        sop_doc["file_size"] = len(file_content)
        
        # Get employees list for name matching
        employees_list = await db.employees.find(
            {"is_active": {"$ne": False}},
            {"_id": 0, "employee_id": 1, "first_name": 1, "last_name": 1}
        ).to_list(500)
        
        # Parse Excel with AI to extract metadata
        parsed = await parse_sop_excel(file_content, employees_list)
        
        sop_doc["preview_data"] = parsed.get("preview_data", [])
        sop_doc["total_rows"] = parsed.get("total_rows", 0)
        sop_doc["total_cols"] = parsed.get("total_cols", 0)
        
        # Store all parsed data
        if parsed.get("sop_number"):
            sop_doc["sop_number"] = parsed["sop_number"]
        
        if not title and parsed.get("title"):
            sop_doc["title"] = parsed["title"]
        
        if parsed.get("process_owner"):
            sop_doc["process_owner"] = parsed["process_owner"]
        
        if parsed.get("created_by"):
            sop_doc["document_created_by"] = parsed["created_by"]
        
        if parsed.get("department") and not departments:
            sop_doc["parsed_department"] = parsed["department"]
        
        if parsed.get("version"):
            sop_doc["document_version"] = parsed["version"]
        
        if parsed.get("purpose"):
            sop_doc["purpose"] = parsed["purpose"]
        
        if parsed.get("scope"):
            sop_doc["scope"] = parsed["scope"]
        
        if parsed.get("procedure_summary"):
            sop_doc["procedure_summary"] = parsed["procedure_summary"]
        
        if parsed.get("responsible_persons"):
            sop_doc["responsible_persons"] = parsed["responsible_persons"]
        
        if parsed.get("reports"):
            sop_doc["reports"] = parsed["reports"]
        
        if parsed.get("task_type") and not task_type:
            sop_doc["task_type"] = parsed["task_type"]
        
        if parsed.get("stakeholders"):
            sop_doc["stakeholders"] = parsed["stakeholders"]
        
        if parsed.get("key_activities"):
            sop_doc["key_activities"] = parsed["key_activities"]
        
        if parsed.get("input_requirements"):
            sop_doc["input_requirements"] = parsed["input_requirements"]
        
        if parsed.get("output_deliverables"):
            sop_doc["output_deliverables"] = parsed["output_deliverables"]
        
        if parsed.get("process_flow_steps"):
            sop_doc["process_flow_steps"] = parsed["process_flow_steps"]
        
        if parsed.get("revision_date"):
            sop_doc["revision_date"] = parsed["revision_date"]
        
        # Try to auto-match process owner to employee
        if parsed.get("process_owner") and not main_responsible:
            matched_id = await match_employee_name(parsed["process_owner"], employees_list)
            if matched_id:
                sop_doc["main_responsible"] = [matched_id]
        
        # Try to match responsible persons to employees
        if parsed.get("responsible_persons") and not also_involved:
            matched_involved = []
            for person in parsed["responsible_persons"]:
                matched = await match_employee_name(person, employees_list)
                if matched and matched not in sop_doc.get("main_responsible", []):
                    matched_involved.append(matched)
            if matched_involved:
                sop_doc["also_involved"] = matched_involved
    
    await db.sops.insert_one(sop_doc)
    sop_doc.pop("_id", None)
    sop_doc.pop("file_data", None)
    
    return sop_doc


@router.get("/{sop_id}")
async def get_sop(sop_id: str, request: Request):
    """Get SOP details"""
    await get_current_user(request)  # Auth check
    
    sop = await db.sops.find_one({"sop_id": sop_id}, {"_id": 0, "file_data": 0})
    if not sop:
        raise HTTPException(status_code=404, detail="SOP not found")
    
    # Get department and designation names
    if sop.get("departments"):
        departments = await db.departments.find(
            {"department_id": {"$in": sop["departments"]}},
            {"_id": 0, "department_id": 1, "name": 1}
        ).to_list(50)
        sop["department_names"] = [d["name"] for d in departments]
    
    if sop.get("designations"):
        designations = await db.designations.find(
            {"designation_id": {"$in": sop["designations"]}},
            {"_id": 0, "designation_id": 1, "name": 1}
        ).to_list(50)
        sop["designation_names"] = [d["name"] for d in designations]
    
    return sop


@router.put("/{sop_id}")
async def update_sop(sop_id: str, request: Request):
    """Update SOP details - all fields are editable"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    form = await request.form()
    
    update_data = {
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "updated_by": user["user_id"]
    }
    
    # All editable fields
    if form.get("title") is not None:
        update_data["title"] = form.get("title")
    if form.get("description") is not None:
        update_data["description"] = form.get("description")
    if form.get("status"):
        update_data["status"] = form.get("status")
    if form.get("task_type") is not None:
        update_data["task_type"] = form.get("task_type")
    if form.get("sop_number") is not None:
        update_data["sop_number"] = form.get("sop_number")
    if form.get("process_owner") is not None:
        update_data["process_owner"] = form.get("process_owner")
    if form.get("purpose") is not None:
        update_data["purpose"] = form.get("purpose")
    if form.get("scope") is not None:
        update_data["scope"] = form.get("scope")
    
    # Handle array fields
    departments = form.getlist("departments")
    if departments is not None:
        update_data["departments"] = [d for d in departments if d] if isinstance(departments, list) else [departments] if departments else []
    
    designations = form.getlist("designations")
    if designations is not None:
        update_data["designations"] = [d for d in designations if d] if isinstance(designations, list) else [designations] if designations else []
    
    main_responsible = form.getlist("main_responsible")
    if main_responsible is not None:
        update_data["main_responsible"] = [m for m in main_responsible if m] if isinstance(main_responsible, list) else [main_responsible] if main_responsible else []
    
    also_involved = form.getlist("also_involved")
    if also_involved is not None:
        update_data["also_involved"] = [a for a in also_involved if a] if isinstance(also_involved, list) else [also_involved] if also_involved else []
    
    # Handle new file upload
    file = form.get("file")
    if file and hasattr(file, 'read'):
        file_content = await file.read()
        update_data["file_name"] = file.filename
        update_data["file_data"] = base64.b64encode(file_content).decode('utf-8')
        update_data["file_type"] = file.content_type
        update_data["file_size"] = len(file_content)
        
        # Increment version
        existing = await db.sops.find_one({"sop_id": sop_id}, {"version": 1})
        update_data["version"] = (existing.get("version", 0) if existing else 0) + 1
        
        # Re-parse Excel
        parsed = await parse_sop_excel(file_content)
        update_data["preview_data"] = parsed.get("preview_data", [])
        update_data["total_rows"] = parsed.get("total_rows", 0)
        update_data["total_cols"] = parsed.get("total_cols", 0)
        
        # Update extracted fields if not explicitly set
        if parsed.get("sop_number") and "sop_number" not in update_data:
            update_data["sop_number"] = parsed["sop_number"]
        if parsed.get("process_owner") and "process_owner" not in update_data:
            update_data["process_owner"] = parsed["process_owner"]
    
    result = await db.sops.update_one({"sop_id": sop_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="SOP not found")
    
    # Return updated SOP
    updated = await db.sops.find_one({"sop_id": sop_id}, {"_id": 0, "file_data": 0})
    return updated


@router.put("/{sop_id}/publish")
async def publish_sop(sop_id: str, request: Request):
    """Publish a draft SOP and send notifications"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get SOP details for notification
    sop = await db.sops.find_one({"sop_id": sop_id}, {"_id": 0})
    if not sop:
        raise HTTPException(status_code=404, detail="SOP not found")
    
    result = await db.sops.update_one(
        {"sop_id": sop_id},
        {"$set": {
            "status": "published",
            "published_at": datetime.now(timezone.utc).isoformat(),
            "published_by": user["user_id"],
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="SOP not found")
    
    # Create notifications for all involved employees
    notification_recipients = set()
    notification_recipients.update(sop.get("main_responsible", []))
    notification_recipients.update(sop.get("also_involved", []))
    
    # Also notify employees in linked designations and departments
    if sop.get("designations"):
        employees_by_desig = await db.employees.find(
            {"designation_id": {"$in": sop["designations"]}, "is_active": True},
            {"_id": 0, "employee_id": 1}
        ).to_list(500)
        for emp in employees_by_desig:
            notification_recipients.add(emp["employee_id"])
    
    if sop.get("departments"):
        employees_by_dept = await db.employees.find(
            {"department_id": {"$in": sop["departments"]}, "is_active": True},
            {"_id": 0, "employee_id": 1}
        ).to_list(500)
        for emp in employees_by_dept:
            notification_recipients.add(emp["employee_id"])
    
    # Create notifications
    if notification_recipients:
        notifications = []
        for emp_id in notification_recipients:
            notifications.append({
                "notification_id": f"notif_{uuid.uuid4().hex[:12]}",
                "user_id": emp_id,
                "type": "sop_published",
                "title": "New SOP Published",
                "message": f"A new SOP '{sop['title']}' has been published that involves you.",
                "link": "/dashboard/sop",
                "sop_id": sop_id,
                "is_read": False,
                "created_at": datetime.now(timezone.utc).isoformat()
            })
        
        if notifications:
            await db.notifications.insert_many(notifications)
    
    return {"message": "SOP published", "notifications_sent": len(notification_recipients)}


@router.delete("/{sop_id}")
async def delete_sop(sop_id: str, request: Request):
    """Soft delete an SOP"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    result = await db.sops.update_one(
        {"sop_id": sop_id},
        {"$set": {"is_active": False, "deleted_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="SOP not found")
    
    return {"message": "SOP deleted"}


@router.post("/{sop_id}/reparse")
async def reparse_sop(sop_id: str, request: Request):
    """Re-parse an existing SOP file with improved AI extraction"""
    user = await get_current_user(request)
    if user.get("role") not in ["super_admin", "hr_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    sop = await db.sops.find_one({"sop_id": sop_id})
    if not sop:
        raise HTTPException(status_code=404, detail="SOP not found")
    
    if not sop.get("file_data"):
        raise HTTPException(status_code=400, detail="No file attached to this SOP")
    
    # Decode and re-parse the file
    file_content = base64.b64decode(sop["file_data"])
    
    # Get employees list for name matching
    employees_list = await db.employees.find(
        {"is_active": {"$ne": False}},
        {"_id": 0, "employee_id": 1, "first_name": 1, "last_name": 1}
    ).to_list(500)
    
    # Re-parse with improved AI
    parsed = await parse_sop_excel(file_content, employees_list)
    
    # Update SOP with new parsed data
    update_data = {
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "reparsed_at": datetime.now(timezone.utc).isoformat(),
        "reparsed_by": user["user_id"]
    }
    
    # Update all parsed fields
    for key in ["sop_number", "title", "process_owner", "document_created_by", "parsed_department",
                "document_version", "purpose", "scope", "procedure_summary", "task_type",
                "input_requirements", "output_deliverables", "revision_date"]:
        if parsed.get(key):
            update_data[key] = parsed[key]
    
    for key in ["responsible_persons", "reports", "stakeholders", "key_activities", "process_flow_steps"]:
        if parsed.get(key) and isinstance(parsed[key], list):
            update_data[key] = parsed[key]
    
    await db.sops.update_one({"sop_id": sop_id}, {"$set": update_data})
    
    # Return updated SOP
    updated = await db.sops.find_one({"sop_id": sop_id}, {"_id": 0, "file_data": 0})
    return updated



@router.get("/{sop_id}/download")
async def download_sop_file(sop_id: str, request: Request):
    """Download the SOP Excel file"""
    await get_current_user(request)  # Auth check
    
    sop = await db.sops.find_one({"sop_id": sop_id})
    if not sop:
        raise HTTPException(status_code=404, detail="SOP not found")
    
    if not sop.get("file_data"):
        raise HTTPException(status_code=404, detail="No file attached to this SOP")
    
    file_content = base64.b64decode(sop["file_data"])
    
    return Response(
        content=file_content,
        media_type=sop.get("file_type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        headers={
            "Content-Disposition": f"attachment; filename={sop.get('file_name', 'sop.xlsx')}"
        }
    )
