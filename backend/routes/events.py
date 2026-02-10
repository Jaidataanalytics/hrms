"""Employee Events & Celebrations API Routes"""
from fastapi import APIRouter, HTTPException, Request, UploadFile, File
from fastapi.responses import StreamingResponse
from typing import Optional
from datetime import datetime, timezone
import uuid
import io
from motor.motor_asyncio import AsyncIOMotorClient
import os

router = APIRouter(prefix="/events", tags=["Employee Events"])

mongo_url = os.environ.get('MONGO_URL', 'mongodb://localhost:27017')
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ.get('DB_NAME', 'test_database')]


async def get_current_user(request: Request) -> dict:
    from server import get_current_user as auth_get_user
    return await auth_get_user(request)


@router.get("")
async def list_events(request: Request, emp_code: Optional[str] = None, event_type: Optional[str] = None):
    """List all employee events, optionally filtered"""
    await get_current_user(request)
    query = {}
    if emp_code:
        query["emp_code"] = emp_code
    if event_type:
        query["event_type"] = event_type

    events = await db.employee_events.find(query, {"_id": 0}).sort("event_date", 1).to_list(500)

    for event in events:
        emp = await db.employees.find_one({"employee_id": event.get("emp_code")}, {"_id": 0, "first_name": 1, "last_name": 1, "department": 1})
        if emp:
            event["employee_name"] = f"{emp.get('first_name', '')} {emp.get('last_name', '')}".strip()
            event["department"] = emp.get("department", "")

    return events


@router.get("/today")
async def get_today_events(request: Request):
    """Get all celebrations happening today"""
    await get_current_user(request)
    today = datetime.now(timezone.utc)
    today_mm_dd = f"{today.month:02d}-{today.day:02d}"

    events = await db.employee_events.find({"_id": {"$exists": True}}, {"_id": 0}).to_list(1000)

    today_events = []
    for event in events:
        event_date_str = event.get("event_date", "")
        try:
            ed = datetime.strptime(event_date_str, "%Y-%m-%d")
            event_mm_dd = f"{ed.month:02d}-{ed.day:02d}"
            if event_mm_dd == today_mm_dd:
                emp = await db.employees.find_one({"employee_id": event.get("emp_code")}, {"_id": 0, "first_name": 1, "last_name": 1, "department": 1, "picture": 1})
                if emp:
                    event["employee_name"] = f"{emp.get('first_name', '')} {emp.get('last_name', '')}".strip()
                    event["department"] = emp.get("department", "")
                    event["picture"] = emp.get("picture")
                if event.get("event_type") == "work_anniversary":
                    years = today.year - ed.year
                    event["years"] = years
                today_events.append(event)
        except (ValueError, TypeError):
            continue

    return today_events


@router.get("/upcoming")
async def get_upcoming_events(request: Request, days: int = 30):
    """Get upcoming events in the next N days"""
    await get_current_user(request)
    today = datetime.now(timezone.utc)

    events = await db.employee_events.find({}, {"_id": 0}).to_list(1000)

    upcoming = []
    for event in events:
        event_date_str = event.get("event_date", "")
        try:
            ed = datetime.strptime(event_date_str, "%Y-%m-%d")
            this_year_date = ed.replace(year=today.year)
            if this_year_date < today.replace(hour=0, minute=0, second=0, microsecond=0):
                this_year_date = ed.replace(year=today.year + 1)
            delta = (this_year_date - today.replace(hour=0, minute=0, second=0, microsecond=0)).days
            if 0 < delta <= days:
                emp = await db.employees.find_one({"employee_id": event.get("emp_code")}, {"_id": 0, "first_name": 1, "last_name": 1, "department": 1})
                if emp:
                    event["employee_name"] = f"{emp.get('first_name', '')} {emp.get('last_name', '')}".strip()
                    event["department"] = emp.get("department", "")
                event["days_until"] = delta
                event["upcoming_date"] = this_year_date.strftime("%Y-%m-%d")
                upcoming.append(event)
        except (ValueError, TypeError):
            continue

    upcoming.sort(key=lambda x: x.get("days_until", 999))
    return upcoming


@router.post("")
async def create_event(data: dict, request: Request):
    """Create a new employee event"""
    user = await get_current_user(request)

    if not data.get("emp_code") or not data.get("event_type") or not data.get("event_date"):
        raise HTTPException(status_code=400, detail="emp_code, event_type, and event_date are required")

    valid_types = ["birthday", "work_anniversary", "marriage_anniversary", "custom"]
    if data["event_type"] not in valid_types:
        raise HTTPException(status_code=400, detail=f"event_type must be one of: {valid_types}")

    existing = await db.employee_events.find_one({
        "emp_code": data["emp_code"],
        "event_type": data["event_type"]
    })
    if existing and data["event_type"] != "custom":
        await db.employee_events.update_one(
            {"event_id": existing["event_id"]},
            {"$set": {"event_date": data["event_date"], "label": data.get("label", ""), "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
        updated = await db.employee_events.find_one({"event_id": existing["event_id"]}, {"_id": 0})
        return updated

    event = {
        "event_id": f"evt_{uuid.uuid4().hex[:12]}",
        "emp_code": data["emp_code"],
        "event_type": data["event_type"],
        "event_date": data["event_date"],
        "label": data.get("label", ""),
        "recurring": True,
        "created_by": user.get("user_id"),
        "created_at": datetime.now(timezone.utc).isoformat()
    }

    await db.employee_events.insert_one(event)
    event.pop("_id", None)
    return event


@router.put("/{event_id}")
async def update_event(event_id: str, data: dict, request: Request):
    """Update an event"""
    await get_current_user(request)

    event = await db.employee_events.find_one({"event_id": event_id})
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")

    update_data = {}
    if "event_date" in data:
        update_data["event_date"] = data["event_date"]
    if "label" in data:
        update_data["label"] = data["label"]
    if "event_type" in data:
        update_data["event_type"] = data["event_type"]
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()

    await db.employee_events.update_one({"event_id": event_id}, {"$set": update_data})
    return await db.employee_events.find_one({"event_id": event_id}, {"_id": 0})


@router.delete("/{event_id}")
async def delete_event(event_id: str, request: Request):
    """Delete an event"""
    await get_current_user(request)

    result = await db.employee_events.delete_one({"event_id": event_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Event not found")
    return {"message": "Event deleted"}


@router.post("/bulk-upload")
async def bulk_upload_events(request: Request, file: UploadFile = File(...)):
    """Bulk upload events from Excel template"""
    user = await get_current_user(request)

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files are supported")

    try:
        import openpyxl
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active

        headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]

        emp_code_idx = None
        event_type_idx = None
        event_date_idx = None
        label_idx = None

        for i, h in enumerate(headers):
            if "emp" in h and ("code" in h or "id" in h):
                emp_code_idx = i
            elif "type" in h:
                event_type_idx = i
            elif "date" in h:
                event_date_idx = i
            elif "label" in h or "name" in h or "description" in h:
                label_idx = i

        if emp_code_idx is None or event_type_idx is None or event_date_idx is None:
            raise HTTPException(status_code=400, detail="Template must have columns: Employee Code, Event Type, Date")

        created = 0
        updated = 0
        errors = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            try:
                emp_code = str(row[emp_code_idx].value).strip() if row[emp_code_idx].value else None
                event_type = str(row[event_type_idx].value).strip().lower().replace(" ", "_") if row[event_type_idx].value else None
                event_date_val = row[event_date_idx].value
                label = str(row[label_idx].value).strip() if label_idx is not None and row[label_idx].value else ""

                if not emp_code or not event_type or not event_date_val:
                    continue

                if isinstance(event_date_val, datetime):
                    event_date = event_date_val.strftime("%Y-%m-%d")
                else:
                    event_date = str(event_date_val).strip()
                    for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"]:
                        try:
                            event_date = datetime.strptime(event_date, fmt).strftime("%Y-%m-%d")
                            break
                        except ValueError:
                            continue

                type_mapping = {
                    "birthday": "birthday",
                    "work_anniversary": "work_anniversary",
                    "work anniversary": "work_anniversary",
                    "joining": "work_anniversary",
                    "joining_date": "work_anniversary",
                    "marriage_anniversary": "marriage_anniversary",
                    "marriage anniversary": "marriage_anniversary",
                    "wedding": "marriage_anniversary",
                    "wedding_anniversary": "marriage_anniversary",
                    "custom": "custom",
                }
                event_type = type_mapping.get(event_type, "custom")

                existing = await db.employee_events.find_one({
                    "emp_code": emp_code,
                    "event_type": event_type
                })

                if existing and event_type != "custom":
                    await db.employee_events.update_one(
                        {"event_id": existing["event_id"]},
                        {"$set": {"event_date": event_date, "label": label, "updated_at": datetime.now(timezone.utc).isoformat()}}
                    )
                    updated += 1
                else:
                    event = {
                        "event_id": f"evt_{uuid.uuid4().hex[:12]}",
                        "emp_code": emp_code,
                        "event_type": event_type,
                        "event_date": event_date,
                        "label": label,
                        "recurring": True,
                        "created_by": user.get("user_id"),
                        "created_at": datetime.now(timezone.utc).isoformat()
                    }
                    await db.employee_events.insert_one(event)
                    created += 1

            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")

        return {
            "message": f"Processed successfully. Created: {created}, Updated: {updated}",
            "created": created,
            "updated": updated,
            "errors": errors[:10]
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to process file: {str(e)}")


@router.get("/template")
async def download_template(request: Request):
    """Download the bulk upload template"""
    await get_current_user(request)

    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employee Events"

    headers = ["Employee Code", "Event Type", "Date (YYYY-MM-DD)", "Label/Description"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)

    examples = [
        ["EMP001", "birthday", "1990-05-15", ""],
        ["EMP002", "work_anniversary", "2020-03-01", "Joined as Engineer"],
        ["EMP003", "marriage_anniversary", "2018-11-20", ""],
        ["EMP004", "custom", "2026-06-10", "Promotion Day"],
    ]
    for row_idx, example in enumerate(examples, 2):
        for col, val in enumerate(example, 1):
            ws.cell(row=row_idx, column=col, value=val)

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 25

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=employee_events_template.xlsx"}
    )


@router.get("/my-events")
async def get_my_events(request: Request):
    """Get events for the current user"""
    user = await get_current_user(request)
    emp_id = user.get("employee_id")

    events = await db.employee_events.find({"emp_code": emp_id}, {"_id": 0}).to_list(50)
    return events


@router.post("/my-events")
async def add_my_event(data: dict, request: Request):
    """Employee adds their own event (marriage anniversary, custom)"""
    user = await get_current_user(request)
    emp_id = user.get("employee_id")

    allowed_types = ["marriage_anniversary", "custom"]
    event_type = data.get("event_type")
    if event_type not in allowed_types:
        raise HTTPException(status_code=400, detail=f"You can only add: {allowed_types}")

    if not data.get("event_date"):
        raise HTTPException(status_code=400, detail="event_date is required")

    existing = await db.employee_events.find_one({
        "emp_code": emp_id,
        "event_type": event_type
    })
    if existing and event_type != "custom":
        await db.employee_events.update_one(
            {"event_id": existing["event_id"]},
            {"$set": {"event_date": data["event_date"], "label": data.get("label", ""), "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
        return await db.employee_events.find_one({"event_id": existing["event_id"]}, {"_id": 0})

    event = {
        "event_id": f"evt_{uuid.uuid4().hex[:12]}",
        "emp_code": emp_id,
        "event_type": event_type,
        "event_date": data["event_date"],
        "label": data.get("label", ""),
        "recurring": True,
        "created_by": user.get("user_id"),
        "created_at": datetime.now(timezone.utc).isoformat()
    }

    await db.employee_events.insert_one(event)
    event.pop("_id", None)
    return event
